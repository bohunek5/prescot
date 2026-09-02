#!/usr/bin/env python3
"""
SKRYPT DOPISUJĄCY NA DOLE ARKUSZA ZŁĄCZKI I AKCESORIA LED ZE STANEM > 100
Zgodnie z poleceniem:
- Edycja bezpośrednia istniejącego pliku (nic nie usuwa, dopisuje na dole).
- Kody producenta oznaczone na zielono.
- Opisy WAPRO HTML bez powielania nazw w <h2> i bez kodów w punktach.
- Zastąpienie Economic -> Standard.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import xml.etree.ElementTree as ET
import json
import html
import re
import os

EXCEL_PATH = "/Users/karolbohdanowicz/Downloads/1294 EL-PLUS.xlsx"
EXCEL_COPY_PATH = "/Users/karolbohdanowicz/Downloads/PRESCOT_EL-PLUS_ZESTAWIENIE_KODOW_2026.xlsx"
XML_PATH = "/Users/karolbohdanowicz/my-ai-agents/prescot/tmp/sources/prescot-wapromag-live.xml"

def clean_str(val):
    return re.sub(r'\s+', ' ', str(val or '')).strip()

def clean_product_name(name):
    t = clean_str(name)
    t = re.sub(r'\bEconomic\b', 'Standard', t, flags=re.I)
    t = re.sub(r'\bECON\b', 'Standard', t, flags=re.I)
    return re.sub(r'\s+', ' ', t).strip()

def build_accessory_wapro_html(item):
    name = clean_product_name(item['name'])
    cat = item['cat']
    uname = name.upper()

    # 1. Złączki WAGO
    if 'WAGO' in uname:
        pins = "5" if "5X" in uname or "415" in uname else "3"
        heading = "Szybkie i bezpieczne łączenie przewodów instalacji LED"
        intro_p1 = f"Oryginalna szybkozłączka uniwersalna WAGO serii 221 na {pins} przewodów umożliwia natychmiastowe, pewne i beznarzędziowe łączenie wszystkich rodzajów przewodów (drut oraz linka) w instalacjach elektrycznych i oświetleniowych LED."
        intro_p2 = "Przezroczysta obudowa pozwala na natychmiastową kontrolę wzrokową prawidłowego osadzenia żyły, a ergonomiczne dźwignie zaciskowe gwarantują wielokrotne otwieranie i zamykanie bez utraty siły docisku sprężyny CAGE CLAMP."
        features = [
            f"Liczba torów / zacisków: {pins}-przewodowa",
            "Obsługiwane przekroje: Drut i linka od 0,2 mm² do 4,0 mm²",
            "Napięcie znamionowe: 450 V AC",
            "Prąd znamionowy: 32 A",
            "Konstrukcja: Krystalicznie przezroczysta obudowa ze zintegrowanym punktem pomiarowym",
            "Technologia zacisku: Opatentowana sprężyna dociskowa CAGE CLAMP"
        ]
        benefits = [
            "Beznarzędziowy montaż skracający czas prac elektroinstalacyjnych",
            "Pełna odporność na poluzowanie styków pod wpływem drgań i zmian temperatury",
            "Możliwość bezpiecznego łączenia przewodów jedno- i wielodrutowych o różnych przekrojach"
        ]
        applications = [
            "Połączenia w puszkach instalacyjnych, szafach sterowniczych, profilach LED i oprawach oświetleniowych"
        ]

    # 2. Złączki FAST / SDK
    elif any(w in uname for w in ['FAST', 'SDK']):
        conn = "dwutorowa (2X)" if "2X" in uname else ("jednotorowa (1X)" if "1X" in uname else "wielotorowa")
        heading = "Błyskawiczne zaciskowe łączenie przewodów LED"
        intro_p1 = f"Profesjonalna złączka zaciskowa FAST ({conn}) do szybkiego łączenia przewodów zasilających i sterujących instalacji oświetlenia LED bez konieczności lutowania."
        intro_p2 = "Zapewnia stabilny docisk miedzianych styków, eliminując straty napięcia i skracając czas montażu opraw meblowych i sufitowych."
        features = [
            f"Wariant połączenia: Złączka zaciskowa {conn}",
            "Przeznaczenie: Beznarzędziowe łączenie przewodów miedzianych",
            "Materiał obudowy: Tworzywo samogasnące o wysokiej wytrzymałości termicznej",
            "Montaż: Szybki montaż dociskowy bez użycia narzędzi"
        ]
        benefits = [
            "Brak konieczności używania lutownicy podczas montażu na drabinie",
            "Niewielkie gabaryty ułatwiające ukrycie w profilu aluminiowym lub puszce"
        ]
        applications = [
            "Instalacje zasilania taśm LED, opraw meblowych i oświetlenia podszafkowego"
        ]

    # 3. Łączniki ZM KLUŚ
    elif 'ŁĄCZNIK ZM' in uname or 'ZM-' in uname:
        angle = "135°" if "135" in uname else ("60°" if "60" in uname else "kątowy")
        heading = "Precyzyjne i stabilne łączenie profili aluminiowych LED"
        intro_p1 = f"Konstrukcyjny łącznik ZM marki KLUŚ ({angle}) służy do sztywnego i precyzyjnego łączenia profili aluminiowych w długie linie świetlne oraz figury geometryczne."
        intro_p2 = "Wykonany z wytrzymałej stali ocynkowanej, wyposażony w śruby dociskowe zapewniające idealne zlicowanie krawędzi profili bez powstawania szczelin i deformacji oprawy."
        features = [
            f"Typ łącznika: Łącznik konstrukcyjny do zamków małych ZM ({angle})",
            "Kompatybilność: Profile aluminiowe LED KLUŚ z gniazdem ZM",
            "Materiał: Stal o podwyższonej sprężystości, cynkowana galwanicznie",
            "Stabilizacja: Docisk za pomocą wkrętów dociskowych (imbusowych)"
        ]
        benefits = [
            "Idealnie proste lub kątowe prowadzenie linii światła bez widocznych przerw",
            "Wysoka sztywność mechaniczna całej konstrukcji oprawy liniowej"
        ]
        applications = [
            "Łączenie opraw wiszących, wpuszczanych i natynkowych w sufitach i ścianach"
        ]

    # 4. Zaślepki profili LED KLUŚ
    elif 'ZAŚLEPKA' in uname and ('KLUŚ' in uname or 'KLUS' in uname or 'PROFIL' in uname):
        with_hole = "z otworem na przewód" if "OTW" in uname or "Z OTWOREM" in uname else "pełna"
        color = "jasnoszary" if "SZAR" in uname else ("biały" if "BIAŁ" in uname else "czarny" if "CZARN" in uname else "standard")
        heading = "Estetyczne wykończenie i ochrona profili LED"
        intro_p1 = f"Zaślepka do profilu aluminiowego LED KLUŚ ({with_hole}, odcień: {color}) to element zamykający oprawę liniową, chroniący wnętrze profilu przed kurzem i zanieczyszczeniami."
        intro_p2 = "Precyzyjnie wyprofilowana krawędź idealnie licuje się ze ściankami profilu i osłoną, tworząc eleganckie, architektoniczne zwieńczenie oprawy."
        features = [
            f"Wariant wykonania: Zaślepka {with_hole}",
            f"Kolor wykończenia: {color.capitalize()}",
            "Materiał: Tworzywo sztuczne o podwyższonej odporności na promieniowanie UV",
            "Montaż: Wciskany w czoło profilu aluminiowego"
        ]
        benefits = [
            "Szczelne zamknięcie profilu zabezpieczające diody LED przed zabrudzeniem",
            "Profesjonalny wygląd oprawy bez widocznych ostrych krawędzi aluminium"
        ]
        applications = [
            "Zwieńczenie profili LED montowanych na sufitach, meblach i ścianach"
        ]

    # 5. Mocowniki, zawieszki i sprężyny KLUŚ
    elif any(w in uname for w in ['MOCOWNIK', 'SPRĘŻYNA', 'ZAWIESZKA']):
        elem = "Sprężyna montażowa" if "SPRĘŻYNA" in uname else ("Zawieszka mocująca" if "ZAWIESZKA" in uname else "Mocownik stalowy")
        heading = "Pewny i trwały montaż profili aluminiowych LED"
        intro_p1 = f"{elem} marki KLUŚ zapewnia szybki, estetyczny i stabilny montaż profili aluminiowych do sufitów, ścian lub zabudów kartonowo-gipsowych."
        intro_p2 = "Sprężysta konstrukcja ze stali nierdzewnej gwarantuje pewny docisk oprawy do podłoża oraz możliwość wygodnego demontażu serwisowego bez uszkodzenia tynku czy mebli."
        features = [
            f"Typ elementu: {elem}",
            "Kompatybilność: Dedykowany do systemowych profili oświetleniowych KLUŚ",
            "Materiał: Stal nierdzewna / stal sprężynowa ocynkowana",
            "Sposób mocowania: Montaż zatrzaskowy (KLIK) do rowka profilu"
        ]
        benefits = [
            "Niewidoczne mocowanie ukryte za korpusem profilu",
            "Błyskawiczny montaż i demontaż oprawy bez użycia skomplikowanych narzędzi"
        ]
        applications = [
            "Mocowanie opraw LED w sufitach podwieszanych, ścianach G-K i ciągach meblowych"
        ]

    # 6. Osłony profili KLUŚ
    elif 'OSŁONA' in uname:
        heading = "Równomierne rozproszenie światła i ochrona taśmy LED"
        intro_p1 = f"Wysokiej klasy osłona optyczna do profilu aluminiowego KLUŚ, zapewniająca perfekcyjne rozproszenie strumienia świetlnego i eliminację widoczności pojedynczych punktów LED."
        intro_p2 = "Wykonana ze szlachetnego poliwęglanu / tworzywa PMMA z filtrem UV, zachowuje pełną elastyczność i krystaliczną estetykę przez lata, nie ulegając żółknięciu."
        features = [
            "Materiał: Poliwęglan (PC) / PMMA z filtrem UV",
            "Funkcja optyczna: Równomierny rozsył światła i ochrona diod przed kurzem",
            "Montaż: Wygodny montaż na wcisk (KLIK) od góry profilu"
        ]
        benefits = [
            "Krystaliczny i jednolity strumień świetlny na całej długości profilu",
            "Odporność na żółknięcie pod wpływem słońca i światła diod LED"
        ]
        applications = [
            "Wykończenie linii światła LED w salonach, biurach, kuchniach i korytarzach"
        ]

    # 7. Przewody i osprzęt elektryczny LED
    else:
        heading = "Niezawodne okablowanie i osprzęt instalacji oświetleniowych LED"
        intro_p1 = f"Profesjonalny komponent elektroinstalacyjny Prescot ({name}) zaprojektowany do stabilnego zasilania i sterowania oprawami oraz taśmami LED."
        intro_p2 = "Wysokiej czystości miedź oraz trwała izolacja zapewniają bezpieczną eksploatację, niski opór elektryczny i wygodne układanie w zabudowach instalacyjnych."
        features = [
            "Przeznaczenie: Niskonapięciowe i sieciowe instalacje oświetlenia LED",
            "Materiał przewodnika: Miedź elektrolityczna o wysokiej czystości",
            "Trwałość: Wysoka odporność izolacji na starzenie i podwyższone temperatury",
            "Zgodność: Standardy bezpieczeństwa CE i normy instalatorskie"
        ]
        benefits = [
            "Pewny przesył energii bez spadków napięcia",
            "Wygodne układanie w peszlach, korytkach kablowych i profilach"
        ]
        applications = [
            "Podłączanie taśm LED, zasilaczy impulsowych, ściemniaczy i sterowników"
        ]

    points_b = '\n'.join([f"<p>- {html.escape(b)}</p>" for b in benefits])
    points_a = '\n'.join([f"<p>- {html.escape(a)}</p>" for a in applications])

    out = f"<section>\n<h2>{html.escape(heading)}</h2>\n<p>{html.escape(intro_p1)}</p>\n<p>{html.escape(intro_p2)}</p>\n"
    if points_b:
        out += f"<h3>Dlaczego warto:</h3>\n{points_b}\n"
    if points_a:
        out += f"<h3>Zastosowanie i miejsce montażu:</h3>\n{points_a}\n"
    out += "</section>"
    return out

def main():
    print(f"⏳ Wczytywanie istniejącego arkusza: {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    initial_row_count = ws.max_row
    print(f"📊 Liczba wierszy w pliku przed dodaniem: {initial_row_count} (nagłówek + {initial_row_count - 1} pozycji)")

    # Zbierz istniejące kody i EANy z pliku, żeby nie zdublować żadnej pozycji
    existing_eans = set()
    existing_codes = set()
    for r in range(2, ws.max_row + 1):
        c_prod = clean_str(ws.cell(r, 2).value).lower()
        c_mc = clean_str(ws.cell(r, 3).value).lower()
        ean = clean_str(ws.cell(r, 4).value)
        if ean: existing_eans.add(ean)
        if c_prod: existing_codes.add(c_prod)
        if c_mc: existing_codes.add(c_mc)

    # Wczytaj XML z chmury
    print("⏳ Wczytywanie bazy z chmury XML...")
    tree = ET.parse(XML_PATH)
    root = tree.getroot()

    items_to_append = []

    def is_target_accessory(name, cat):
        uname = name.upper()
        ucat = cat.upper()

        if any(w in ucat for w in ['KAJA', 'LAMPY BIURKOWE', 'KINKIETY', 'LAMPA PODŁOGOWA', 'ŻYRANDOL', 'PLAFON', 'ŚWIETLÓWKI LED', 'MODUŁY LED']):
            return False
        if any(w in uname for w in ['LAMPA BIURKOWA', 'KINKIET', 'LAMPKI CHOINKOWE', 'ŚWIETLÓWKA']):
            return False

        # 1. Złączki
        if any(w in uname for w in ['ZŁĄCZK', 'ZLACZK', 'SZYBKOZŁĄCZK', 'ŁĄCZNIK', 'LACZNIK', 'WAGO', 'FAST PROSTA', 'FAST KĄTOWA', 'SDK 5']):
            return True
        # 2. Zaślepki
        if 'ZAŚLEPK' in uname or 'ZASLEPK' in uname or 'ZAŚLEPKI' in ucat:
            return True
        # 3. Mocowniki, zawieszki, sprężyny KLUŚ
        if any(w in uname for w in ['MOCOWNIK', 'SPRĘŻYNA', 'ZAWIESZKA']) or (('UCHWYT' in uname) and ('KOSZULK' in uname or 'PROFIL' in uname or 'KLUŚ' in uname)):
            return True
        # 4. Osłony profili KLUŚ
        if 'OSŁONA' in uname and ('KLUŚ' in uname or 'KLUS' in uname or 'PROFIL' in uname):
            return True
        # 5. Przewody do LED i włączniki
        if any(w in uname for w in ['PRZEWÓD', 'PRZEWOD', 'WŁĄCZNIK']) and any(w in uname for w in ['SIF', 'TRANS', 'OMY', 'OZ-600', 'WŁĄCZNIK DO OBUDOWY', '3X0.75']):
            return True
        if 'OPRAWKA GU10' in uname:
            return True
        if 'OPRAWKA BRYZGOSZCZELNA' in uname:
            return True

        return False

    for o in root.iter('o'):
        attrs = {a.attrib.get('name'): a.text for a in o.findall('attrs/a')}
        ean = clean_str(attrs.get('EAN'))
        kp = clean_str(attrs.get('Kod_produktu'))
        k_prod = clean_str(attrs.get('Kod producenta'))
        name = clean_str(o.findtext('name'))
        cat = clean_str(o.findtext('cat'))
        price = float(o.attrib.get('price') or 0.0)
        try: stock = float(o.attrib.get('stock') or 0.0)
        except: stock = 0.0

        if stock <= 100:
            continue
        if (ean and ean in existing_eans) or (kp and kp.lower() in existing_codes) or (k_prod and k_prod.lower() in existing_codes):
            continue

        if is_target_accessory(name, cat):
            existing_eans.add(ean)
            if kp: existing_codes.add(kp.lower())
            if k_prod: existing_codes.add(k_prod.lower())

            img_url = ([i.attrib.get('url') for i in o.findall('imgs/main') if i.attrib.get('url')] or [i.attrib.get('url') for i in o.findall('imgs/i') if i.attrib.get('url')] or [''])[0]

            items_to_append.append({
                'ean': ean,
                'kod_produktu': kp,
                'kod_producenta': k_prod or kp,
                'name': clean_product_name(name),
                'cat': cat,
                'price': price,
                'stock': int(stock),
                'img': img_url
            })

    print(f"✨ Znaleziono {len(items_to_append)} złączek i akcesoriów LED ze stanem > 100 do dopisania na dole.")

    # Style do komórek
    font_data = Font(name="Calibri", size=10, color="000000")
    font_green_data = Font(name="Calibri", size=10, bold=True, color="1E4620")
    fill_green_data = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    start_row = ws.max_row + 1
    # Ustal kolejny numer Lp
    try:
        current_lp = int(ws.cell(ws.max_row, 1).value or 0)
    except:
        current_lp = ws.max_row - 1

    for idx, item in enumerate(items_to_append, 1):
        row_idx = start_row + idx - 1
        current_lp += 1

        opis_wapro = build_accessory_wapro_html(item)

        unit = "m" if "PRZEWÓD" in item['name'].upper() and not "WTYCZK" in item['name'].upper() else "szt."
        kod_cn = "8536 90 10" if "ZŁĄCZK" in item['name'].upper() else ("8544 49 95" if "PRZEWÓD" in item['name'].upper() else "3926 90 97")

        row_vals = [
            current_lp,
            item['kod_producenta'],  # Col 2 - ZIELONA
            item['kod_produktu'],    # Col 3 - STARY SYMBOL
            item['ean'],             # Col 4
            item['name'],            # Col 5
            item['cat'],             # Col 6
            item['price'],           # Col 7 - CENA KATALOGOWA NETTO
            None,                    # Col 8 - cena uwagi
            "PLN",                   # Col 9 - waluta
            23,                      # Col 10 - podatek
            item['stock'],           # Col 11 - ilość
            unit,                    # Col 12 - jednostka
            None,                    # Col 13 - jednostka uwagi
            None,                    # Col 14 - min zam
            None,                    # Col 15 - interwał
            None,                    # Col 16 - ETIM
            kod_cn,                  # Col 17 - kod cn
            None,                    # Col 18 - PKWiU
            "TAK",                   # Col 19 - split payment
            "Prescot Sp. z o.o., ul. Wileńska 1, 11-500 Giżycko, mail: komponenty@prescot.pl, tel: 877776482", # Col 20 - GPSR
            "24h",                   # Col 21 - czas real
            opis_wapro,              # Col 22 - OPIS WAPRO
            item['img'],             # Col 23 - zdjęcie
            None,                    # Col 24 - karta
            None                     # Col 25 - certyfikaty
        ]

        ws.append(row_vals)
        ws.row_dimensions[row_idx].height = 20

        # Zastosuj formatowanie
        for col_idx in range(1, len(row_vals) + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border_thin
            cell.font = font_data

            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2: # Kolumna 'kod producenta' - OZNACZONA NA ZIELONO
                cell.fill = fill_green_data
                cell.font = font_green_data
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx in [3, 4]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx == 7: # Cena katalogowa
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [9, 10, 11, 12, 19, 21]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    print(f"📈 Nowa liczba pozycji w arkuszu: {ws.max_row - 1}")

    # Zapisz bez naruszania struktury
    wb.save(EXCEL_PATH)
    wb.save(EXCEL_COPY_PATH)

    print(f"✅ Zapisano pomyślnie bezpośrednio w: {EXCEL_PATH}")
    print(f"✅ Zsynchronizowano plik: {EXCEL_COPY_PATH}")

if __name__ == "__main__":
    main()
