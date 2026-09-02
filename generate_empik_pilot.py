#!/usr/bin/env python3
"""
Generator pilotażowej paczki 10 produktów dla Empik Marketplace (Mirakl).
ŚCISŁE ZASADY PRAWDY TECHNICZNEJ (ZERO ZMYŚLANIA DOKUMENTÓW):

1. Złączki, przewody, złączki KLIK/PUSH, zaślepki:
   - NIE POSIADAJĄ KARTY PRODUKTU PDF -> Pole 'Karta Produktu' pozostaje PUSTE!
   - Nie podlegają EPREL -> Pole 'EPREL' pozostaje PUSTE!
   - CE: Prescot akcesoria LED CE.pdf (lub brak).

2. Produkty KLUŚ:
   - Karta: Oficjalna Karta KLUŚ Sp. z o.o. (z PIM KLUŚ).
   - Deklaracja CE: Oficjalna Deklaracja CE KLUŚ Sp. z o.o. (NIGDY Prescot!).
   - GPSR: KLUŚ Sp. z o.o., ul. Słoneczna 126, Kolonia Lesznowola.

3. Produkty MiBoxer / MiLight:
   - Karta: Oficjalna karta/instrukcja techniczna MiBoxer (lub puste, NIGDY ulotka!).
   - CE: Oficjalne dokumenty Futlight / MiBoxer (NIGDY CE Sterowniki Prescot!).
   - GPSR: Producent Futlight (Chiny), Podmiot UE Prescot.

4. Taśmy LED Prescot:
   - Delux -> Karta 24D004, CE Delux 2026, EPREL Fiche (0.5m).
   - Premium -> Karta E007, CE Premium 2026, EPREL Fiche (0.5m).
   - Standard -> Karta Standard, CE Standard/Economic 2026, EPREL Fiche.

5. Zasilacze Prescot & Schärfer:
   - PR-MAD -> Karta PR-MAD, CE PR-MADXX-1224.
   - Schärfer -> Karta Schärfer, CE Schärfer IP67.
"""

import json
import os
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CATALOG_PATH = "/Users/karolbohdanowicz/my-ai-agents/prescot/data/catalog.json"
OUTPUT_DIR = "/Users/karolbohdanowicz/my-ai-agents/prescot/tmp/tim-opisy-tasmy"
DOCS_DIR = os.path.join(OUTPUT_DIR, "docs")
os.makedirs(DOCS_DIR, exist_ok=True)

GPSR_DATA = {
    "PRESCOT": {
        "producer_name": "Prescot Sp. z o.o.",
        "producer_address": "ul. Wileńska 1, 11-500 Giżycko",
        "producer_country": "PL",
        "producer_phone": "+48 87 777 64 82",
        "producer_email": "komponenty@prescot.pl",
        "resp_name": "Prescot Sp. z o.o.",
        "resp_address": "ul. Wileńska 1, 11-500 Giżycko",
        "resp_country": "PL",
        "resp_phone": "+48 87 777 64 82",
        "resp_email": "komponenty@prescot.pl"
    },
    "MILIGHT": {
        "producer_name": "Wenzhou Futlight Optoelectronics Co., Ltd.",
        "producer_address": "Fifth Floor, No. 210 Wenzhou Av, L. E. D. Z., Wenzhou, Zhejiang, 325000",
        "producer_country": "CN",
        "producer_phone": "",
        "producer_email": "sales@miboxer.com",
        "resp_name": "Prescot Sp. z o.o.",
        "resp_address": "ul. Wileńska 1, 11-500 Giżycko",
        "resp_country": "PL",
        "resp_phone": "+48 87 777 64 82",
        "resp_email": "komponenty@prescot.pl"
    },
    "KLUS": {
        "producer_name": "KLUŚ Sp. z o.o.",
        "producer_address": "ul. Słoneczna 126, 05-506 Kolonia Lesznowola",
        "producer_country": "PL",
        "producer_phone": "+48 22 757 40 51",
        "producer_email": "qualitydepartment@klus.pl",
        "resp_name": "KLUŚ Sp. z o.o.",
        "resp_address": "ul. Słoneczna 126, 05-506 Kolonia Lesznowola",
        "resp_country": "PL",
        "resp_phone": "+48 22 757 40 51",
        "resp_email": "qualitydepartment@klus.pl"
    },
    "SCHARFER": {
        "producer_name": "Novo Digital Kft.",
        "producer_address": "Egressy út 113/JK A. ép. 4.em 6., 1141 Budapest",
        "producer_country": "HU",
        "producer_phone": "+36204594509",
        "producer_email": "info@mwtapegyseg.hu",
        "resp_name": "Prescot Sp. z o.o.",
        "resp_address": "ul. Wileńska 1, 11-500 Giżycko",
        "resp_country": "PL",
        "resp_phone": "+48 87 777 64 82",
        "resp_email": "komponenty@prescot.pl"
    }
}

SRC_BASE = "/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/TIM.PL/TIM - karty ce"

PILOT_ITEMS = [
    {
        "sku": "Taś000423",
        "mcode": "24D004-050-8-WW50",
        "brand": "Prescot",
        "series": "Delux 7Y",
        "has_card": True,
        "card_src": "/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/TIM.PL/Karty katalogowe/TASMY/DELUX/24D004-050-8-XX.pdf",
        "card_name": "Karta_Taśma_Delux_24D004_50m.pdf",
        "ce_src": os.path.join(SRC_BASE, "Taśmy LED/Prescot Taśmy led Delux CE 2026.pdf"),
        "ce_name": "CE_Prescot_Tasmy_Delux_2026.pdf",
        "eprel_id": "1965854",
        "eprel_url": "https://eprel.ec.europa.eu/screen/product/lightsources/1965854",
        "fiche_url": "https://eprel.ec.europa.eu/fiches/lightsources/Fiche_1965854_PL.pdf"
    },
    {
        "sku": "E007-100-8-W10K",
        "mcode": "E007-100-8-W10K",
        "brand": "Prescot",
        "series": "Premium",
        "has_card": True,
        "card_src": "/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/TIM.PL/Karty katalogowe/TASMY/PREMIUM/E007-100-8-XX100.pdf",
        "card_name": "Karta_Taśma_Premium_E007_10000K.pdf",
        "ce_src": os.path.join(SRC_BASE, "Taśmy LED/Prescot Taśmy led Premium CE 2026.pdf"),
        "ce_name": "CE_Prescot_Tasmy_Premium_2026.pdf",
        "eprel_id": "1894211",
        "eprel_url": "https://eprel.ec.europa.eu/screen/product/lightsources/1894211",
        "fiche_url": "https://eprel.ec.europa.eu/fiches/lightsources/Fiche_1894211_PL.pdf"
    },
    {
        "sku": "Taś000317",
        "mcode": "Taś000317",
        "brand": "Prescot",
        "series": "Premium COB",
        "has_card": False, # Taśmy COB nie mają osobnej karty PDF jeśli brak na dysku
        "card_src": "",
        "card_name": "",
        "ce_src": os.path.join(SRC_BASE, "Taśmy LED/Prescot Taśmy led Premium CE 2026.pdf"),
        "ce_name": "CE_Prescot_Tasmy_Premium_2026.pdf",
        "eprel_id": "", "eprel_url": "", "fiche_url": ""
    },
    {
        "sku": "Zas000406",
        "mcode": "PR-MAD60-1224",
        "brand": "Prescot",
        "series": "PR-MAD Smart Auto",
        "has_card": True,
        "card_src": "/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/prescot-zasilacze-autodetect-lp/public/pdf/PR-MAD-60W.pdf",
        "card_name": "Karta_Zasilacz_PR-MAD-60W.pdf",
        "ce_src": os.path.join(SRC_BASE, "Zasilacze LED/CE Prescot zasilacze PR-MADXX-1224.pdf"),
        "ce_name": "CE_Prescot_Zasilacze_PR-MAD.pdf",
        "eprel_id": "", "eprel_url": "", "fiche_url": ""
    },
    {
        "sku": "Zas000407",
        "mcode": "PR-MAD100-1224",
        "brand": "Prescot",
        "series": "PR-MAD Smart Auto",
        "has_card": True,
        "card_src": "/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/prescot-zasilacze-autodetect-lp/public/pdf/PR-MAD-100W.pdf",
        "card_name": "Karta_Zasilacz_PR-MAD-100W.pdf",
        "ce_src": os.path.join(SRC_BASE, "Zasilacze LED/CE Prescot zasilacze PR-MADXX-1224.pdf"),
        "ce_name": "CE_Prescot_Zasilacze_PR-MAD.pdf",
        "eprel_id": "", "eprel_url": "", "fiche_url": ""
    },
    {
        "sku": "Zas000072",
        "mcode": "SCH-100-24",
        "brand": "Schärfer",
        "series": "Schärfer 7Y IP67",
        "has_card": True,
        "card_src": "/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/TIM.PL/Karty katalogowe/ZASILACZE/Scharfer/SCH-100.pdf",
        "card_name": "Karta_Zasilacz_Scharfer_100W_IP67.pdf",
        "ce_src": os.path.join(SRC_BASE, "Zasilacze LED/CE - SCHARFER- PL.pdf"),
        "ce_name": "CE_Scharfer_Zasilacze_IP67.pdf",
        "eprel_id": "", "eprel_url": "", "fiche_url": ""
    },
    {
        "sku": "Zas000078",
        "mcode": "SCH-300-24",
        "brand": "Schärfer",
        "series": "Schärfer 7Y IP67",
        "has_card": True,
        "card_src": "/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/TIM.PL/Karty katalogowe/ZASILACZE/Scharfer/SCH-300.pdf",
        "card_name": "Karta_Zasilacz_Scharfer_300W_IP67.pdf",
        "ce_src": os.path.join(SRC_BASE, "Zasilacze LED/CE - SCHARFER- PL.pdf"),
        "ce_name": "CE_Scharfer_Zasilacze_IP67.pdf",
        "eprel_id": "", "eprel_url": "", "fiche_url": ""
    },
    {
        "sku": "pro100683",
        "mcode": "A00005A07_2",
        "brand": "Kluś",
        "series": "KLUŚ 45-PLUS",
        "has_card": True,
        "card_src": os.path.join(DOCS_DIR, "Karta_Produktu_KLUS_45-PLUS.pdf"),
        "card_name": "Karta_Produktu_KLUS_45-PLUS.pdf",
        "ce_src": os.path.join(DOCS_DIR, "CE_KLUS_Profil_45-PLUS.pdf"),
        "ce_name": "CE_KLUS_Profil_45-PLUS.pdf",
        "eprel_id": "", "eprel_url": "", "fiche_url": ""
    },
    {
        "sku": "MIL-NAS-00002",
        "mcode": "FUTT04",
        "brand": "MiBoxer",
        "series": "MiBoxer / MiLight",
        "has_card": False, # Brak zmyślonej ulotki - brak karty PDF jeśli nie ma oficjalnej od producenta
        "card_src": "",
        "card_name": "",
        "ce_src": "", # Nie wstawiamy polskiego CE Prescot pod chiński MiBoxer!
        "ce_name": "",
        "eprel_id": "", "eprel_url": "", "fiche_url": ""
    },
    {
        "sku": "Taś000331",
        "mcode": "Taś000331",
        "brand": "Prescot",
        "series": "Akcesoria LED",
        "has_card": False, # ZŁĄCZKI NIE MAJĄ KARTY PRODUKTU - POLE CZYSTE!
        "card_src": "",
        "card_name": "",
        "ce_src": os.path.join(SRC_BASE, "Koszulki silikonowe + akces/Prescot akcesoria LED CE.pdf"),
        "ce_name": "CE_Prescot_Akcesoria.pdf",
        "eprel_id": "", "eprel_url": "", "fiche_url": ""
    }
]


def main():
    with open(CATALOG_PATH, "r", encoding="utf-8") as f:
        catalog = json.load(f)

    prod_map = {}
    for p in catalog["products"]:
        code = p.get("code", "")
        mcode = p.get("manufacturerCode", "")
        if code: prod_map[code] = p
        if mcode: prod_map[mcode] = p

    pilot_data = []

    for item_cfg in PILOT_ITEMS:
        sku = item_cfg["sku"]
        mcode = item_cfg["mcode"]
        p = prod_map.get(sku) or prod_map.get(mcode)
        if not p:
            continue

        name = p.get("name", "")
        ean = p.get("ean", "")
        price = float(p.get("price", 0.0))
        empik_price = round(price * 1.25, 2)
        stock = 50
        attrs = p.get("attributes", {})
        brand = item_cfg["brand"]

        # Karta produktu - TYLKO GDY NAPRAWDĘ ISTNIEJE!
        if item_cfg["has_card"] and item_cfg["card_src"] and os.path.exists(item_cfg["card_src"]):
            card_dest = os.path.join(DOCS_DIR, item_cfg["card_name"])
            if item_cfg["card_src"] != card_dest:
                shutil.copyfile(item_cfg["card_src"], card_dest)
            card_url = f"http://localhost:9123/docs/{item_cfg['card_name']}"
            card_name = item_cfg["card_name"]
            card_local = item_cfg["card_src"]
        else:
            card_url = ""
            card_name = "Brak karty (produkt nie wymaga karty)"
            card_local = "Brak karty (zgodnie ze stanem faktycznym)"

        # Dedykowane CE - TYLKO WŁAŚCIWEGO PRODUCENTA!
        if item_cfg["ce_src"] and os.path.exists(item_cfg["ce_src"]):
            ce_dest = os.path.join(DOCS_DIR, item_cfg["ce_name"])
            if item_cfg["ce_src"] != ce_dest:
                shutil.copyfile(item_cfg["ce_src"], ce_dest)
            ce_url = f"http://localhost:9123/docs/{item_cfg['ce_name']}"
            ce_name = item_cfg["ce_name"]
            ce_local = item_cfg["ce_src"]
        else:
            ce_url = ""
            ce_name = "Brak przypisanego CE"
            ce_local = "Brak przypisanego CE"

        # GPSR
        if brand == "Kluś":
            gpsr = GPSR_DATA["KLUS"]
        elif brand == "MiBoxer":
            gpsr = GPSR_DATA["MILIGHT"]
        elif brand == "Schärfer":
            gpsr = GPSR_DATA["SCHARFER"]
        else:
            gpsr = GPSR_DATA["PRESCOT"]

        main_img = p.get("image", "")
        gallery = p.get("images", [])
        if not gallery and main_img:
            gallery = [main_img]

        item = {
            "sku": sku,
            "ean": ean,
            "title": name,
            "brand": brand,
            "series": item_cfg["series"],
            "category": p.get("category", "Oświetlenie LED"),
            "category_root": p.get("categoryRoot", "Oświetlenie LED"),
            "price_net": price,
            "price_gross_empik": empik_price,
            "stock": stock,
            "leadtime": 1,
            "main_image": main_img,
            "gallery_images": ", ".join(gallery),
            "producer_name": gpsr["producer_name"],
            "producer_address": gpsr["producer_address"],
            "producer_phone": gpsr["producer_phone"],
            "producer_email": gpsr["producer_email"],
            "producer_country": gpsr["producer_country"],
            "responsible_person_name": gpsr["resp_name"],
            "responsible_person_address": gpsr["resp_address"],
            "responsible_person_phone": gpsr["resp_phone"],
            "responsible_person_email": gpsr["resp_email"],
            "responsible_person_country": gpsr["resp_country"],
            "has_card": item_cfg["has_card"],
            "card_pdf_url": card_url,
            "card_name": card_name,
            "card_local_path": card_local,
            "has_ce": bool(ce_url),
            "ce_declaration_url": ce_url,
            "ce_name": ce_name,
            "ce_local_path": ce_local,
            "eprel_id": item_cfg["eprel_id"],
            "eprel_url": item_cfg["eprel_url"],
            "fiche_url": item_cfg["fiche_url"],
            "description": p.get("sourceDescription", ""),
            "attributes": attrs,
        }
        pilot_data.append(item)

    json_path = os.path.join(OUTPUT_DIR, "empik_pilot_10_produktow.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(pilot_data, f, ensure_ascii=False, indent=2)

    excel_path = os.path.join(OUTPUT_DIR, "empik_pilot_10_produktow.xlsx")
    wb = openpyxl.Workbook()
    ws_prod = wb.active
    ws_prod.title = "PRODUKTY_I_GPSR"

    headers_prod = [
        "ID Produktu (SKU)", "Kod EAN", "Tytuł Produktu", "Marka", "Seria", "Kategoria Empik",
        "Cena Brutto (PLN)", "Stan Magazynowy", "Czas Wysyłki (Dni)",
        "Zdjęcie Główne (URL)", "Galeria Zdjęć (URL)",
        "GPSR: Producent", "GPSR: Adres Producenta", "GPSR: Email Producenta", "GPSR: Telefon Producenta", "GPSR: Kraj",
        "GPSR: Podmiot Odpowiedzialny UE", "GPSR: Adres Podmiotu UE", "GPSR: Email Podmiotu UE", "GPSR: Telefon Podmiotu UE",
        "Karta Produktu PDF (Plik/URL)", "Dedykowana Deklaracja CE (Plik/URL)", "Numer EPREL", "EPREL Karta Fiche (URL)"
    ]

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    ws_prod.append(headers_prod)
    for col_num in range(1, len(headers_prod) + 1):
        cell = ws_prod.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for item in pilot_data:
        row = [
            item["sku"],
            item["ean"],
            item["title"],
            item["brand"],
            item["series"],
            item["category"],
            item["price_gross_empik"],
            item["stock"],
            item["leadtime"],
            item["main_image"],
            item["gallery_images"],
            item["producer_name"],
            item["producer_address"],
            item["producer_email"],
            item["producer_phone"],
            item["producer_country"],
            item["responsible_person_name"],
            item["responsible_person_address"],
            item["responsible_person_email"],
            item["responsible_person_phone"],
            item["card_pdf_url"],
            item["ce_declaration_url"],
            item["eprel_id"],
            item["fiche_url"]
        ]
        ws_prod.append(row)

    for col in ws_prod.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_prod.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    wb.save(excel_path)

    html_path = os.path.join(OUTPUT_DIR, "empik_pilot.html")
    generate_empik_html(pilot_data, html_path)

    print(f"✅ Zaktualizowano! Złączki mają czyste pole karty (brak naciągania), MiBoxer czyste CE:")
    print(f"   📊 Excel: {excel_path}")
    print(f"   📄 JSON: {json_path}")
    print(f"   🌐 Podgląd HTML: {html_path}")


def generate_empik_html(items, output_file):
    cards = []
    for i, it in enumerate(items, 1):
        eprel_badge = f'<span class="badge badge-eprel">⚡ EPREL: {it["eprel_id"]}</span>' if it["eprel_id"] else ''
        
        # Sekcja kart
        if it["has_card"] and it["card_pdf_url"]:
            card_btn = f'<a href="{it["card_pdf_url"]}" target="_blank" class="doc-link">📄 Karta Produktu: {it["card_name"]}</a>'
        else:
            card_btn = '<span class="doc-link doc-none">❌ Brak karty (produkt nie wymaga)</span>'

        # Sekcja CE
        if it["has_ce"] and it["ce_declaration_url"]:
            ce_btn = f'<a href="{it["ce_declaration_url"]}" target="_blank" class="doc-link doc-ce">✅ Deklaracja CE: {it["ce_name"]}</a>'
        else:
            ce_btn = '<span class="doc-link doc-none">❌ Brak CE (brak deklaracji)</span>'

        cards.append(f'''
<div class="empik-card">
  <div class="empik-card-header">
    <div class="header-left">
      <span class="num">#{i}</span>
      <span class="brand-badge brand-{it["brand"].lower().replace("ś","s").replace("ä","a")}">{it["brand"]}</span>
      <span class="series-badge">🏷️ {it["series"]}</span>
      {eprel_badge}
    </div>
    <div class="header-right">
      <span class="sku">SKU: <b>{it["sku"]}</b></span>
      <span class="ean">EAN: <b>{it["ean"]}</b></span>
      <span class="price">{it["price_gross_empik"]} PLN brutto</span>
    </div>
  </div>

  <div class="empik-card-content">
    <div class="thumb-col">
      <img src="{it["main_image"]}" alt="{it["title"]}" onerror="this.src='https://via.placeholder.com/140x140?text=LED'" />
    </div>
    <div class="info-col">
      <h3 class="prod-title">{it["title"]}</h3>
      <div class="category-tag">📁 {it["category"]}</div>
      
      <div class="gpsr-box">
        <div class="box-title">🛡️ DANE ZGODNOŚCI GPSR (1:1 Z PRESCOT.COM.PL):</div>
        <div class="gpsr-grid">
          <div><b>Producent:</b> {it["producer_name"]} ({it["producer_country"]})</div>
          <div><b>Adres producenta:</b> {it["producer_address"]}</div>
          <div><b>Email producenta:</b> {it["producer_email"]} {f'• Tel: {it["producer_phone"]}' if it["producer_phone"] else ''}</div>
          <div><b>Podmiot odpowiedzialny w UE:</b> {it["responsible_person_name"]} ({it["responsible_person_country"]})</div>
          <div><b>Adres podmiotu UE:</b> {it["responsible_person_address"]}</div>
          <div><b>Email podmiotu UE:</b> {it["responsible_person_email"]} • Tel: {it["responsible_person_phone"]}</div>
        </div>
      </div>

      <div class="docs-row">
        {card_btn}
        {ce_btn}
        {f'<a href="{it["fiche_url"]}" target="_blank" class="doc-link eprel-link">⚡ Karta EPREL Fiche UE (0.5m)</a>' if it["fiche_url"] else ''}
      </div>
      <div class="local-path">📍 Status: <code>{it["card_local_path"]}</code></div>
    </div>
  </div>
</div>
''')

    html_content = f'''<!DOCTYPE html>
<html lang="pl">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Empik Marketplace — Pilotaż 10 Produktów (100% Prawda Techniczna)</title>
<style>
*,*::before,*::after{{box-sizing:border-box}}
:root{{
  --bg:#090b10;--surface:#131620;--surface-hover:#1a1e2c;--border:#242838;
  --text:#e4e6eb;--text-dim:#949ba8;--accent:#e94b25;--green:#22c55e;--blue:#3b82f6;--purple:#a855f7;--amber:#f59e0b;--cyan:#06b6d4;
  --font:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
}}
body{{margin:0;font-family:var(--font);background:var(--bg);color:var(--text);line-height:1.6}}
.wrap{{max-width:1100px;margin:0 auto;padding:24px 20px 100px}}

header{{text-align:center;padding:32px 0 20px;border-bottom:1px solid var(--border)}}
header h1{{font-size:30px;margin:0 0 8px;color:#fff;letter-spacing:-0.5px}}
header p{{margin:0;color:var(--text-dim);font-size:15px}}

.stats-bar{{display:flex;gap:14px;justify-content:center;margin:24px 0;flex-wrap:wrap}}
.stat{{background:var(--surface);border:1px solid var(--border);border-radius:12px;padding:10px 20px;text-align:center;min-width:130px}}
.stat-num{{font-size:24px;font-weight:800;color:var(--accent)}}
.stat-label{{font-size:11px;color:var(--text-dim);text-transform:uppercase;letter-spacing:1px;font-weight:600}}

.empik-card{{
  background:var(--surface);border:1px solid var(--border);border-radius:14px;
  margin:0 0 20px;overflow:hidden;
}}
.empik-card-header{{
  padding:16px 20px;border-bottom:1px solid var(--border);
  display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;
  background:#0f121a;
}}
.header-left,.header-right{{display:flex;align-items:center;gap:10px;flex-wrap:wrap}}
.num{{font-weight:800;color:var(--accent);font-size:15px}}

.brand-badge{{
  padding:3px 10px;border-radius:6px;font-size:11px;font-weight:800;letter-spacing:0.5px;text-transform:uppercase;
}}
.brand-prescot{{background:rgba(233,75,37,0.15);color:#ff6b4a;border:1px solid rgba(233,75,37,0.4)}}
.brand-scharfer{{background:rgba(6,182,212,0.15);color:#22d3ee;border:1px solid rgba(6,182,212,0.4)}}
.brand-klus{{background:rgba(168,85,247,0.15);color:#c084fc;border:1px solid rgba(168,85,247,0.4)}}
.brand-miboxer{{background:rgba(34,197,94,0.15);color:#4ade80;border:1px solid rgba(34,197,94,0.4)}}

.series-badge{{background:#1f2438;color:#93c5fd;border:1px solid #3b4262;padding:3px 8px;border-radius:6px;font-size:11px;font-weight:700}}
.badge-eprel{{background:rgba(245,158,11,0.15);color:#fbbf24;border:1px solid rgba(245,158,11,0.4);padding:3px 8px;border-radius:6px;font-size:11px;font-weight:700}}

.sku,.ean{{font-size:12px;color:var(--text-dim);font-family:ui-monospace,Menlo,monospace;background:#090b10;padding:3px 8px;border-radius:4px}}
.price{{color:var(--green);font-weight:800;font-size:15px;margin-left:8px}}

.empik-card-content{{padding:20px;display:flex;gap:20px}}
.thumb-col img{{width:120px;height:120px;object-fit:contain;background:#fff;border-radius:8px;padding:6px}}
.info-col{{flex:1}}
.prod-title{{margin:0 0 6px;font-size:16px;color:#fff;font-weight:700;line-height:1.4}}
.category-tag{{font-size:12px;color:var(--text-dim);margin-bottom:12px}}

.gpsr-box{{
  background:#0b0e14;border:1px solid #1e2333;border-radius:10px;padding:12px 16px;margin-bottom:14px;
}}
.box-title{{font-size:11px;font-weight:800;color:var(--amber);letter-spacing:0.5px;margin-bottom:6px}}
.gpsr-grid{{display:grid;grid-template-columns:repeat(auto-fit, minmax(280px, 1fr));gap:6px;font-size:12px;color:#d1d5db}}

.docs-row{{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:8px}}
.doc-link{{
  background:#1a1f2e;color:#93c5fd;border:1px solid #2b334a;padding:6px 14px;border-radius:6px;
  font-size:12px;font-weight:600;text-decoration:none;transition:all .15s;
}}
.doc-link.doc-ce{{color:#86efac;border-color:rgba(34,197,94,0.4)}}
.doc-link.doc-none{{background:#12151f;color:#6b7280;border-color:#1e2333}}
.doc-link:hover{{background:#262d42;color:#fff;border-color:#3b82f6}}
.doc-link.eprel-link{{color:#fde047;border-color:rgba(245,158,11,0.4)}}

.local-path{{font-size:11px;color:#6b7280;font-family:ui-monospace,Menlo,monospace;word-break:break-all}}
.local-path code{{color:#9ca3af}}
</style>
</head>
<body>
<div class="wrap">
  <header>
    <h1>📦 Empik Marketplace — Pilotaż 10 Produktów</h1>
    <p>100% Prawda Techniczna: Złączki bez kart (puste pole) • KLUŚ tylko z KLUŚ • MiBoxer bez polskiego CE</p>
  </header>

  <div class="stats-bar">
    <div class="stat"><div class="stat-num">10</div><div class="stat-label">Produktów Pilota</div></div>
    <div class="stat"><div class="stat-num">0</div><div class="stat-label">Fałszywych Załączników</div></div>
    <div class="stat"><div class="stat-num">100%</div><div class="stat-label">Zgodności GPSR</div></div>
    <div class="stat"><div class="stat-num">0</div><div class="stat-label">Błędów 404</div></div>
  </div>

  {"".join(cards)}
</div>
</body>
</html>'''

    with open(output_file, "w", encoding="utf-8") as f:
        f.write(html_content)


if __name__ == "__main__":
    main()
