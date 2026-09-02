#!/usr/bin/env python3
import openpyxl
import re
import os

FILES = [
    "/Users/karolbohdanowicz/Downloads/1294 EL-PLUS.xlsx",
    "/Users/karolbohdanowicz/Downloads/PRESCOT_EL-PLUS_ZESTAWIENIE_KODOW_2026.xlsx"
]

def clean_file(fpath):
    if not os.path.exists(fpath):
        print(f"File not found: {fpath}")
        return

    print(f"Processing {fpath}...")
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active

    rows_to_delete = []

    for r in range(2, ws.max_row + 1):
        code_prod = str(ws.cell(r, 2).value or "").strip()
        code_art = str(ws.cell(r, 3).value or "").strip()
        ean = str(ws.cell(r, 4).value or "").strip()
        name = str(ws.cell(r, 5).value or "").strip()
        row_text = " ".join([str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1)])

        is_3in1 = bool(re.search(r'\b3\s*(?:in|w)\s*1\b', name, re.I) or 
                       re.search(r'\b3\s*(?:in|w)\s*1\b', code_prod, re.I) or 
                       re.search(r'\b3\s*(?:in|w)\s*1\b', code_art, re.I))

        is_sterowniczy = bool("sterownicz" in name.lower() or 
                              "sterownicz" in row_text.lower() and ("oz600" in code_prod.lower() or "oz600" in code_art.lower() or "przewód" in name.lower()))
        
        # Raw generic cables added at the end that are not LED accessories
        is_raw_cable = code_prod in ["OZ600225", "OZ600240", "SIF100CZ", "TR1050", "OMY2100B", "S3-3075-3-CZ"] or \
                       code_art in ["OZ600225", "OZ600240", "SIF100CZ", "TR1050", "OMY2100B", "S3-3075-3-CZ"]

        if is_3in1 or is_sterowniczy or is_raw_cable:
            reason = "3IN1" if is_3in1 else ("PRZEWÓD STEROWNICZY" if is_sterowniczy else "RAW CABLE")
            rows_to_delete.append((r, reason, code_prod, code_art, name))

    print(f"Found {len(rows_to_delete)} rows to delete:")
    for r, reason, cp, ca, n in rows_to_delete:
        print(f"  Row {r} [{reason}]: {cp} | {ca} | {n}")

    # Delete in reverse order
    for r, reason, cp, ca, n in sorted(rows_to_delete, key=lambda x: x[0], reverse=True):
        ws.delete_rows(r, 1)

    # Renumber Lp (Column A / Col 1)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).value = r - 1

    wb.save(fpath)
    print(f"Saved {fpath} with {ws.max_row - 1} remaining products (renumbered 1 to {ws.max_row - 1}).\n")

for f in FILES:
    clean_file(f)
