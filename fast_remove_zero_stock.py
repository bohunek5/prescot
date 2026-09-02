#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy
import re
import os

FILES = [
    "/Users/karolbohdanowicz/Downloads/1294 EL-PLUS.xlsx",
    "/Users/karolbohdanowicz/Downloads/PRESCOT_EL-PLUS_ZESTAWIENIE_KODOW_2026.xlsx"
]

def clean_file(fpath):
    if not os.path.exists(fpath):
        print(f"Brak pliku: {fpath}")
        return

    print(f"\n--- Przetwarzanie: {fpath} ---")
    wb = openpyxl.load_workbook(fpath)
    old_ws = wb.active
    sname = old_ws.title

    new_ws = wb.create_sheet(title="Towary_Cleaned")

    # Copy header row
    max_col = old_ws.max_column
    for c in range(1, max_col + 1):
        cell_src = old_ws.cell(1, c)
        cell_dst = new_ws.cell(1, c, cell_src.value)
        if cell_src.has_style:
            cell_dst.font = copy(cell_src.font)
            cell_dst.fill = copy(cell_src.fill)
            cell_dst.border = copy(cell_src.border)
            cell_dst.alignment = copy(cell_src.alignment)
            cell_dst.number_format = cell_src.number_format

    kept_cnt = 0
    deleted_cnt = 0

    for r in range(2, old_ws.max_row + 1):
        name = str(old_ws.cell(r, 5).value or "").strip()
        raw_stock = old_ws.cell(r, 11).value
        try:
            stock = float(raw_stock or 0)
        except:
            stock = 0.0

        uname = name.lower()
        has_wyc = bool(re.search(r'\bwyc\.?\b', uname, re.I) or 'wycofan' in uname)

        # Czy taśma LED?
        is_tape = False
        if uname.startswith('taśma') or uname.startswith('tasma'):
            if not any(w in uname for w in ['zaślepk', 'uchwyt', 'złączk', 'klips', 'profil', 'zestaw zaślepek']):
                is_tape = True

        should_delete = False
        if has_wyc:
            should_delete = True
        elif stock <= 0 and not is_tape:
            should_delete = True

        if should_delete:
            deleted_cnt += 1
            continue

        kept_cnt += 1
        new_row_idx = kept_cnt + 1

        for c in range(1, max_col + 1):
            cell_src = old_ws.cell(r, c)
            val = kept_cnt if c == 1 else cell_src.value
            cell_dst = new_ws.cell(new_row_idx, c, val)
            if cell_src.has_style:
                cell_dst.font = copy(cell_src.font)
                cell_dst.fill = copy(cell_src.fill)
                cell_dst.border = copy(cell_src.border)
                cell_dst.alignment = copy(cell_src.alignment)
                cell_dst.number_format = cell_src.number_format

    # Copy column widths
    for col_letter, col_dim in old_ws.column_dimensions.items():
        new_ws.column_dimensions[col_letter].width = col_dim.width

    # Replace sheet
    wb.remove(old_ws)
    new_ws.title = sname

    wb.save(fpath)
    print(f"✅ Zapisano {fpath}:")
    print(f"   Usunięto produktów: {deleted_cnt}")
    print(f"   Pozostało produktów: {kept_cnt}")
    print(f"   Przenumerowano Lp: 1 do {kept_cnt}")

for f in FILES:
    clean_file(f)
