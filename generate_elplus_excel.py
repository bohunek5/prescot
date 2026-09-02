#!/usr/bin/env python3
"""
GENERATOR ZESTAWIENIA SYMBOLI DLA EL-PLUS & MC (PRESCOT 2026) - V2.0 FINAL
1. Zestawienie starych symboli (kod produktu w MC) z nowymi (kod producenta / indeks handlowy).
2. Kolumna 'kod producenta' oznaczona na zielono.
3. W nagłówkach i opisach brak powielania nazwy produktu oraz brak indeksów/kodów w wypunktowaniach.
4. Zamiana słowa 'Economic' na 'Standard' we wszystkich nazwach i opisach.
5. Czyste, rozbudowane opisy WAPRO HTML dla wszystkich pozycji (stare opisy całkowicie zastąpione).
6. Aktualizacja cen netto z chmury WAPRO (prescot.xml).
7. Dołączenie nowych produktów z XML: Taśmy LED COB, IP67, IP62, WCOB, 48V, Zasilacze PR-MAD.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
import json
import html
import re
import os
import shutil

ELPLUS_ORIGINAL_PATH = "/Users/karolbohdanowicz/Downloads/1294 EL-PLUS.xlsx"
BACKUP_PATH = "/Users/karolbohdanowicz/Downloads/1294 EL-PLUS_BACKUP_ORIGINAL.xlsx"
OUTPUT_ELPLUS_PATH = "/Users/karolbohdanowicz/Downloads/1294 EL-PLUS.xlsx"
OUTPUT_NEW_FILE_PATH = "/Users/karolbohdanowicz/Downloads/PRESCOT_EL-PLUS_ZESTAWIENIE_KODOW_2026.xlsx"

XML_PATH = "/Users/karolbohdanowicz/my-ai-agents/prescot/tmp/sources/prescot-wapromag-live.xml"
CATALOG_PATH = "/Users/karolbohdanowicz/my-ai-agents/prescot/data/catalog.json"
SEO_PATH = "/Users/karolbohdanowicz/my-ai-agents/prescot/dist/data/seo-descriptions.json"

def clean_str(val):
    return re.sub(r'\s+', ' ', str(val or '')).strip()

def normalize_key(val):
    return clean_str(val).lower()

def clean_product_name(name):
    t = clean_str(name)
    t = re.sub(r'\bEconomic\b', 'Standard', t, flags=re.I)
    t = re.sub(r'\bECON\b', 'Standard', t, flags=re.I)
    t = re.sub(r'\s+', ' ', t).strip()
    return t

def render_clean_wapro_html(name, editorial):
    heading = editorial.get('sections', [{}])[0].get('heading') or "Profesjonalne oświetlenie liniowe LED"
    intro_paras = editorial.get('sections', [{}])[0].get('paragraphs', [editorial.get('meta_description', name)])
    if isinstance(intro_paras, str):
        intro_paras = [intro_paras]

    intro_html = '\n'.join([f"<p>{html.escape(clean_str(p))}</p>" for p in intro_paras if clean_str(p)])

    raw_features = editorial.get('sections', [{}, {}, {}])[2].get('paragraphs', []) if len(editorial.get('sections', [])) >= 3 else []
    clean_features = []
    for f in raw_features:
        s = clean_str(f)
        s_low = s.lower()
        if (s_low.startswith("kod:") or s_low.startswith("kod / indeks:") or 
            s_low.startswith("nazwa:") or s_low.startswith("model / oznaczenie:") or 
            s_low.startswith("producent:") or s_low.startswith("model:") or "kod produktu" in s_low):
            continue
        clean_features.append(s)

    benefits = editorial.get('benefits', [])
    clean_benefits = [b for b in benefits if not any(w in b.lower() for w in ['kod', 'ean', 'indeks'])]

    applications = editorial.get('applications', []) or (editorial.get('sections', [{}, {}])[1].get('paragraphs', []) if len(editorial.get('sections', [])) >= 2 else [])
    
    points_f = '\n'.join([f"<p>- {html.escape(clean_str(f).rstrip('.'))}</p>" for f in clean_features])
    points_b = '\n'.join([f"<p>- {html.escape(clean_str(b).rstrip('.'))}</p>" for b in clean_benefits])
    points_a = '\n'.join([f"<p>- {html.escape(clean_str(a).rstrip('.'))}</p>" for a in applications])

    out = f"<section>\n<h2>{html.escape(clean_str(heading))}</h2>\n{intro_html}\n"
    if points_b:
        out += f"<h3>Dlaczego warto:</h3>\n{points_b}\n"
    if points_a:
        out += f"<h3>Zastosowanie i miejsce montażu:</h3>\n{points_a}\n"
    out += "</section>"
    return out

def main():
    print("⏳ Wczytywanie bazy produktów i chmury XML...")
    
    # 1. Wczytaj catalog.json
    with open(CATALOG_PATH, 'r', encoding='utf-8') as f:
        catalog = json.load(f)
    cat_by_ean = {}
    cat_by_code = {}
    for p in catalog['products']:
        ean = clean_str(p.get('ean'))
        code = clean_str(p.get('code'))
        if ean: cat_by_ean[ean] = p
        if code: cat_by_code[normalize_key(code)] = p

    # 2. Wczytaj seo-descriptions.json (V13.0)
    with open(SEO_PATH, 'r', encoding='utf-8') as f:
        seo_data = json.load(f)
    seo_prods = seo_data.get('products', {})

    # 3. Wczytaj XML z chmury WAPRO
    tree = ET.parse(XML_PATH)
    root = tree.getroot()
    xml_by_ean = {}
    xml_by_kp = {}
    xml_by_kprod = {}
    xml_all_items = []

    for o in root.iter('o'):
        attrs = {a.attrib.get('name'): a.text for a in o.findall('attrs/a')}
        ean = clean_str(attrs.get('EAN'))
        kp = clean_str(attrs.get('Kod_produktu'))
        k_prod = clean_str(attrs.get('Kod producenta'))
        name = clean_str(o.findtext('name'))
        cat = clean_str(o.findtext('cat'))
        price = float(o.attrib.get('price') or 0.0)
        stock = clean_str(o.attrib.get('stock') or '0')
        
        imgs = [i.attrib.get('url') for i in o.findall('imgs/main') if i.attrib.get('url')]
        if not imgs:
            imgs = [i.attrib.get('url') for i in o.findall('imgs/i') if i.attrib.get('url')]
        img_url = imgs[0] if imgs else ''

        item = {
            'id': o.attrib.get('id'),
            'ean': ean,
            'kod_produktu': kp,
            'kod_producenta': k_prod,
            'name': name,
            'cat': cat,
            'price': price,
            'stock': stock,
            'img': img_url,
            'attrs': attrs,
            'o': o
        }
        xml_all_items.append(item)
        if ean: xml_by_ean[ean] = item
        if kp: xml_by_kp[normalize_key(kp)] = item
        if k_prod: xml_by_kprod[normalize_key(k_prod)] = item

    print(f"📦 Załadowano {len(xml_all_items)} produktów z XML WAPRO.")

    # 4. Wczytaj plik bazowy 1294 EL-PLUS.xlsx
    wb_orig = openpyxl.load_workbook(BACKUP_PATH if os.path.exists(BACKUP_PATH) else ELPLUS_ORIGINAL_PATH, data_only=True)
    ws_orig = wb_orig.active

    existing_eans = set()
    existing_symbols = set()
    rows_data = []

    price_updated_count = 0
    price_identical_count = 0

    for r in range(2, ws_orig.max_row + 1):
        old_sym = clean_str(ws_orig.cell(r, 2).value)
        ean = clean_str(ws_orig.cell(r, 3).value)
        name_raw = clean_str(ws_orig.cell(r, 4).value)
        name = clean_product_name(name_raw)

        id_kat = clean_str(ws_orig.cell(r, 5).value)
        cena_orig = ws_orig.cell(r, 6).value
        try: cena_num = float(cena_orig)
        except: cena_num = 0.0

        cena_uwagi = ws_orig.cell(r, 7).value
        waluta = ws_orig.cell(r, 8).value or "PLN"
        podatek = ws_orig.cell(r, 9).value or 23
        ilosc = ws_orig.cell(r, 10).value
        jednostka = ws_orig.cell(r, 11).value or "szt."
        jednostka_uwagi = ws_orig.cell(r, 12).value
        min_zam = ws_orig.cell(r, 13).value
        interwal = ws_orig.cell(r, 14).value
        klasa_etim = ws_orig.cell(r, 15).value
        kod_cn = ws_orig.cell(r, 16).value
        pkwiu = ws_orig.cell(r, 17).value
        split_pay = ws_orig.cell(r, 18).value or "TAK"
        gpsr_prod = ws_orig.cell(r, 19).value or "Prescot Sp. z o.o., ul. Wileńska 1, 11-500 Giżycko, mail: komponenty@prescot.pl, tel: 877776482"
        czas_real = ws_orig.cell(r, 20).value or "24h"
        zdjecie_orig = ws_orig.cell(r, 22).value
        karta_orig = ws_orig.cell(r, 23).value
        certyf_orig = ws_orig.cell(r, 24).value

        if ean: existing_eans.add(ean)
        if old_sym: existing_symbols.add(normalize_key(old_sym))

        # Dopasowanie w XML i catalog.json
        xml_item = xml_by_ean.get(ean) or xml_by_kp.get(normalize_key(old_sym)) or xml_by_kprod.get(normalize_key(old_sym))
        cat_item = cat_by_ean.get(ean) or cat_by_code.get(normalize_key(old_sym))

        # Kod producenta (nowy ładny indeks handlowy)
        kod_producenta_nowy = ""
        if cat_item and cat_item.get('manufacturerCode'):
            kod_producenta_nowy = cat_item.get('manufacturerCode')
        elif xml_item and xml_item.get('kod_producenta'):
            kod_producenta_nowy = xml_item.get('kod_producenta')
        elif cat_item and cat_item.get('code'):
            kod_producenta_nowy = cat_item.get('code')
        else:
            kod_producenta_nowy = old_sym

        # Kod produktu (stary symbol w MC)
        kod_produktu_stary = old_sym

        # Aktualizacja ceny netto z XML
        final_price = cena_num
        if xml_item and xml_item['price'] > 0:
            if abs(cena_num - xml_item['price']) > 0.001:
                price_updated_count += 1
            else:
                price_identical_count += 1
            final_price = xml_item['price']

        # Zdjęcie
        final_img = zdjecie_orig
        if not final_img and xml_item and xml_item['img']:
            final_img = xml_item['img']

        # NOWY OPIS WAPRO (stare opisy całkowicie zastąpione nowymi)
        seo_obj = seo_prods.get(f"ean:{ean}") or seo_prods.get(f"code:{old_sym}") or seo_prods.get(f"code:{kod_producenta_nowy}")
        if seo_obj and seo_obj.get('editorial'):
            final_opis = render_clean_wapro_html(name, seo_obj['editorial'])
        else:
            final_opis = f"<section>\n<h2>Profesjonalne oświetlenie LED</h2>\n<p>{name}</p>\n<h3>Zastosowanie i miejsce montażu:</h3>\n<p>- Przeznaczony do profesjonalnych instalacji oświetleniowych LED</p>\n</section>"

        rows_data.append({
            'kod_producenta': kod_producenta_nowy,
            'kod_produktu': kod_produktu_stary,
            'ean': ean,
            'name': name,
            'id_kategoria': id_kat,
            'cena': final_price,
            'cena_uwagi': cena_uwagi,
            'waluta': waluta,
            'podatek': podatek,
            'ilosc': ilosc,
            'jednostka': jednostka,
            'jednostka_uwagi': jednostka_uwagi,
            'min_zam': min_zam,
            'interwal': interwal,
            'klasa_etim': klasa_etim,
            'kod_cn': kod_cn,
            'pkwiu': pkwiu,
            'split_pay': split_pay,
            'gpsr_prod': gpsr_prod,
            'czas_real': czas_real,
            'opis': final_opis,
            'zdjecie': final_img,
            'karta': karta_orig,
            'certyfikaty': certyf_orig,
            'is_new': False
        })

    print(f"📊 Przetworzono {len(rows_data)} pozycji pierwotnych.")

    # 5. Wykrywanie nowych produktów do dodania z XML
    def is_target_new_product(item):
        uname = item['name'].upper()
        ucode = (item['kod_producenta'] + ' ' + item['kod_produktu']).upper()
        cat = item['cat']
        is_tape = 'TAŚMA' in uname or 'TASMA' in uname or 'Taśmy' in cat

        if 'COB' in uname or 'COB' in ucode or 'WCOB' in uname or 'WCOB' in ucode:
            return True, 'COB'
        if ('48V' in uname or '48V' in ucode) and is_tape:
            return True, '48V'
        if 'PR-MAD' in uname or 'PR-MAD' in ucode or ('MAD' in ucode and 'Zasilacz' in item['name']):
            return True, 'PR-MAD'
        if is_tape and ('IP67' in uname or 'IP67' in ucode or 'HERMETYCZ' in uname):
            return True, 'IP67'
        if is_tape and any(w in uname or w in ucode for w in ['IP62', 'IP63', 'IP65']):
            return True, 'IP62'
        return False, ''

    new_added_count = 0
    for item in xml_all_items:
        ean = item['ean']
        kp = item['kod_produktu']
        k_prod = item['kod_producenta']
        
        # Pomiń jeśli już w pliku
        if (ean and ean in existing_eans) or (kp and normalize_key(kp) in existing_symbols) or (k_prod and normalize_key(k_prod) in existing_symbols):
            continue

        is_match, tag = is_target_new_product(item)
        if is_match:
            new_added_count += 1
            existing_eans.add(ean)
            if kp: existing_symbols.add(normalize_key(kp))
            if k_prod: existing_symbols.add(normalize_key(k_prod))

            name = clean_product_name(item['name'])
            seo_obj = seo_prods.get(f"ean:{ean}") or seo_prods.get(f"code:{kp}") or seo_prods.get(f"code:{k_prod}")
            if seo_obj and seo_obj.get('editorial'):
                final_opis = render_clean_wapro_html(name, seo_obj['editorial'])
            else:
                final_opis = f"<section>\n<h2>Profesjonalne oświetlenie LED</h2>\n<p>{name}</p>\n<h3>Zastosowanie i miejsce montażu:</h3>\n<p>- Przeznaczony do profesjonalnych instalacji oświetleniowych LED</p>\n</section>"

            cat_str = item['cat'] or ("Taśmy LED" if "Taśma" in name else "Zasilacze LED")

            rows_data.append({
                'kod_producenta': k_prod or kp,
                'kod_produktu': kp or k_prod,
                'ean': ean,
                'name': name,
                'id_kategoria': cat_str,
                'cena': item['price'],
                'cena_uwagi': None,
                'waluta': "PLN",
                'podatek': 23,
                'ilosc': item['stock'] or None,
                'jednostka': "szt." if "Zasilacz" in name else "m",
                'jednostka_uwagi': None,
                'min_zam': None,
                'interwal': None,
                'klasa_etim': None,
                'kod_cn': "8504 40 82" if "Zasilacz" in name else "9405 42 31",
                'pkwiu': None,
                'split_pay': "TAK",
                'gpsr_prod': "Prescot Sp. z o.o., ul. Wileńska 1, 11-500 Giżycko, mail: komponenty@prescot.pl, tel: 877776482",
                'czas_real': "24h",
                'opis': final_opis,
                'zdjecie': item['img'],
                'karta': None,
                'certyfikaty': None,
                'is_new': True
            })

    print(f"✨ Dodano {new_added_count} nowych produktów z chmury (COB, IP67, IP62, WCOB, 48V, PR-MAD).")
    print(f"📈 Łączna liczba produktów w arkuszu: {len(rows_data)}")

    # 6. Tworzenie nowego profesjonalnego skoroszytu Excel
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "Towary"

    headers = [
        "Lp",
        "kod producenta",     # Kolumna 2 - NOWY INDEKS HANDLOWY (ZIELONA)
        "kod produktu",       # Kolumna 3 - STARY SYMBOL W MC
        "Kod EAN",            # Kolumna 4
        "nazwa",              # Kolumna 5
        "idKategoria",        # Kolumna 6
        "cena katalogowa",    # Kolumna 7
        "cena uwagi",         # Kolumna 8
        "waluta",             # Kolumna 9
        "podatek",            # Kolumna 10
        "ilość",              # Kolumna 11
        "jednostka",          # Kolumna 12
        "jednostka uwagi",    # Kolumna 13
        "minimum zamówieniowe", # Kolumna 14
        "interwał zamówienia",  # Kolumna 15
        "klasa ETIM",         # Kolumna 16
        "kod cn",             # Kolumna 17
        "PKWiU",              # Kolumna 18
        "split payment",      # Kolumna 19
        "gpsr producent",     # Kolumna 20
        "Dostep. czas realizacji", # Kolumna 21
        "Opis",               # Kolumna 22
        "zdjęcie",            # Kolumna 23
        "karta katalogowa",   # Kolumna 24
        "certyfikaty"         # Kolumna 25
    ]

    # Style
    font_header_default = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_header_green = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10, color="000000")
    font_green_data = Font(name="Calibri", size=10, bold=True, color="1E4620")

    fill_header_default = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Granatowy
    fill_header_green = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")   # Ciemnozielony
    fill_green_data = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")     # Jasnozielone tło dla kodów producenta

    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(1, col_idx)
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col_idx == 2:
            cell.fill = fill_header_green
            cell.font = font_header_green
        else:
            cell.fill = fill_header_default
            cell.font = font_header_default

    ws.row_dimensions[1].height = 28

    for idx, r in enumerate(rows_data, 1):
        row_vals = [
            idx,
            r['kod_producenta'],
            r['kod_produktu'],
            r['ean'],
            r['name'],
            r['id_kategoria'],
            r['cena'],
            r['cena_uwagi'],
            r['waluta'],
            r['podatek'],
            r['ilosc'],
            r['jednostka'],
            r['jednostka_uwagi'],
            r['min_zam'],
            r['interwal'],
            r['klasa_etim'],
            r['kod_cn'],
            r['pkwiu'],
            r['split_pay'],
            r['gpsr_prod'],
            r['czas_real'],
            r['opis'],
            r['zdjecie'],
            r['karta'],
            r['certyfikaty']
        ]
        ws.append(row_vals)
        row_idx = idx + 1
        ws.row_dimensions[row_idx].height = 20

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border_thin
            cell.font = font_data

            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2: # Kolumna 'kod producenta' - OZNACZONA NA ZIELONO
                cell.fill = fill_green_data
                cell.font = font_green_data
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx in [3, 4]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_idx == 7: # Cena katalogowa
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [9, 10, 11, 12, 19, 21]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    col_widths = {
        1: 6,   # Lp
        2: 25,  # kod producenta (ZIELONA)
        3: 22,  # kod produktu (STARY W MC)
        4: 16,  # EAN
        5: 45,  # nazwa
        6: 35,  # kategoria
        7: 16,  # cena
        8: 12,  # cena uwagi
        9: 8,   # waluta
        10: 8,  # podatek
        11: 10, # ilość
        12: 10, # jednostka
        13: 15, # jednostka uwagi
        14: 15, # min zam
        15: 15, # interwal
        16: 14, # ETIM
        17: 14, # CN
        18: 14, # PKWiU
        19: 14, # split pay
        20: 35, # GPSR
        21: 15, # czas real
        22: 40, # Opis
        23: 35, # Zdjęcie
        24: 20, # Karta
        25: 20  # Certyfikaty
    }

    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    wb_out.save(OUTPUT_ELPLUS_PATH)
    wb_out.save(OUTPUT_NEW_FILE_PATH)

    print(f"✅ Zapisano gotowy plik: {OUTPUT_ELPLUS_PATH}")
    print(f"✅ Zapisano gotowy plik dodatkowy: {OUTPUT_NEW_FILE_PATH}")

if __name__ == "__main__":
    main()
