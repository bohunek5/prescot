#!/usr/bin/env python3
"""Build the TIM Scharfer remediation workbook from captured live evidence."""

from __future__ import annotations

import json
import re
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/karolbohdanowicz/my-ai-agents/prescot")
OUTPUT = Path("/Users/karolbohdanowicz/Downloads/TIM_SCHARFER_RAPORT_2026-09-01.xlsx")
XML_PATH = ROOT / "tmp/sources/prescot-live-2026-08-31.xml"
PRE_AUDIT = ROOT / "exports/tim/remediation/scharfer-document-assets-audit-2026-09-01.json"
POST_AUDIT = ROOT / "exports/tim/remediation/scharfer-document-assets-postverify-2026-09-01.json"
FINAL_AUDIT = ROOT / "exports/tim/remediation/scharfer-final-live-readonly-2026-09-01.json"
EAN_BLOCK = ROOT / "exports/tim/remediation/scharfer-pilot-validation-capture-2026-09-01.json"
LIVE_REPORTS = [
    ROOT / "exports/tim/remediation/scharfer-document-refresh-pilot-live-2026-09-01.json",
    ROOT / "exports/tim/remediation/scharfer-document-refresh-batch10-live-2026-09-01.json",
    ROOT / "exports/tim/remediation/scharfer-document-refresh-rest9-live-2026-09-01.json",
]
DESCRIPTION_REPORTS = [
    ROOT / "exports/tim/remediation/scharfer-description-pilot-live-2026-09-01.json",
    ROOT / "exports/tim/remediation/scharfer-description-batch10-live-2026-09-01.json",
    ROOT / "exports/tim/remediation/scharfer-description-rest9-live-2026-09-01.json",
]

NEW_EAN = {
    "SCH-18-12": "5999863091001", "SCH-18-24": "5999863091018",
    "SCH-20-12": "5999863091025", "SCH-20-24": "5999863091032",
    "SCH-30-12": "5999863091049", "SCH-30-24": "5999863091063",
    "SCH-45-12": "5999863091056", "SCH-45-24": "5999863091070",
    "SCH-60-12": "5999863091087", "SCH-60-24": "5999863091094",
    "SCH-100-12": "5999863091100", "SCH-100-24": "5999863091117",
    "SCH-150-12": "5999863091124", "SCH-150-24": "5999863091131",
    "SCH-200-12": "5999863091148", "SCH-200-24": "5999863091155",
    "SCH-300-12": "5999863091162", "SCH-300-24": "5999863091179",
    "SCH-400-12": "5999863091186", "SCH-400-24": "5999863091193",
}


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def numeric(value):
    if isinstance(value, dict):
        value = value.get("value")
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def xml_by_model():
    root = ET.parse(XML_PATH).getroot()
    result = {}
    for offer in root.findall("o"):
        attrs = {
            str(node.attrib.get("name") or "").strip(): str(node.text or "").strip()
            for node in list(offer.find("attrs") or [])
        }
        model = attrs.get("Kod producenta", "") or attrs.get("Kod_produktu", "")
        if model not in NEW_EAN:
            continue
        result[model] = {
            "ean": attrs.get("EAN", ""),
            "price": numeric(offer.attrib.get("price")),
            "stock": numeric(offer.attrib.get("stock")) or 0,
            "url": str(offer.attrib.get("url") or ""),
            "name": str(offer.findtext("name") or "").strip(),
        }
    return result


def style(ws, url_columns=()):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="8B174D")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column in url_columns:
            cell = row[column - 1]
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = min(60, max(11, max(len(str(cell.value or "")) for cell in list(column)[:200]) + 2))
        ws.column_dimensions[letter].width = width


xml_rows = xml_by_model()
if set(xml_rows) != set(NEW_EAN):
    raise RuntimeError(f"XML nie zawiera pełnych 20 modeli Scharfer: {sorted(set(NEW_EAN) - set(xml_rows))}")
for model, expected in NEW_EAN.items():
    if xml_rows[model]["ean"] != expected:
        raise RuntimeError(f"{model}: EAN XML {xml_rows[model]['ean']} != oczekiwany {expected}")

pre = read_json(PRE_AUDIT)
post = read_json(POST_AUDIT)
final_audit = read_json(FINAL_AUDIT)
block = read_json(EAN_BLOCK)
live_results = []
for path in LIVE_REPORTS:
    report = read_json(path)
    if report.get("fatalError") or report.get("blockedWrites"):
        raise RuntimeError(f"Nieczysty raport zapisu: {path}")
    live_results.extend(report["results"])
if len(live_results) != 20 or any(row.get("status") != "saved_and_verified" for row in live_results):
    raise RuntimeError("Brak 20 poprawnie zapisanych i zweryfikowanych kart Scharfer.")
description_results = []
for path in DESCRIPTION_REPORTS:
    report = read_json(path)
    if report.get("fatalError") or report.get("blockedWrites"):
        raise RuntimeError(f"Nieczysty raport opisu: {path}")
    description_results.extend(report["results"])
if len(description_results) != 20 or any(row.get("status") != "saved_and_verified" for row in description_results):
    raise RuntimeError("Brak 20 poprawnie zapisanych i zweryfikowanych opisów Scharfer.")

live_by_model = {row["model"]: row for row in live_results}
post_by_model = {row["model"]: row for row in post["products"]}
final_by_id = {int(row["id"]): row for row in final_audit["products"]}
if pre["uniqueAssets"] != 40 or pre["downloadableAssets"] != 0 or post["unavailableAssets"] != 0:
    raise RuntimeError("Audyt dokumentów przed/po nie zgadza się z oczekiwanym wynikiem.")

wb = Workbook()
ws = wb.active
ws.title = "PODSUMOWANIE"
ws.append(["Pozycja", "Wynik", "Znaczenie"])
summary = [
    ("Aktywne karty Scharfer objęte pracą", 20, "Wszystkie istniejące aktywne karty 12 V i 24 V, moce 18–400 W"),
    ("Stare relacje dokumentów", "40/40 niedostępne", "Każde CE i każda karta zwracały HTTP 500"),
    ("Nowe relacje dokumentów", "40/40 dostępne", "20× CE + 20× karta katalogowa; 11 unikalnych PDF-ów"),
    ("Produkty zapisane i sprawdzone", "20/20", "Pilot 1, następnie 10, następnie 9"),
    ("Opisy poprawione i sprawdzone", "20/20", "Indeks handlowy SCH, bez EAN i bez wewnętrznych kodów PRE"),
    ("Ceny", "NIE ZMIENIONO", "Cena każdej karty porównana przed i po zapisie"),
    ("Nazwy, status, workflow", "NIE ZMIENIONO", "Każda karta pozostała aktywna i opublikowana"),
    ("Nowe EAN w XML", "20/20 potwierdzone", "EAN dostawcy 599986...; zgodne z potwierdzonym zestawem"),
    ("Zmiana EAN w TIM", "ZABLOKOWANA PRZEZ TIM", "Kod kreskowy nie może być edytowany po nadaniu indeksu TIM"),
    ("Stan magazynowy w TIM", "0 na starych EAN", "Feed używa nowych EAN, więc nie dopasowuje się do kart ze starym zablokowanym EAN"),
    ("EPREL", "NIE DOTYCZY", "Scharfer to zasilacze LED / oddzielny osprzęt sterujący, nie źródła światła"),
    ("Osobna instrukcja", "BRAK ŹRÓDŁA", "Wskazówki montażowe są w karcie technicznej; nie utworzono sztucznego duplikatu"),
    ("Stare duplikaty", 3, "Aktywne, lecz bez EAN, zdjęcia, opisu i dokumentów; nie wycofano bez osobnej decyzji"),
    ("Następny krok", "Migracja EAN po stronie TIM", "Bez duplikowania kart i bez obchodzenia blokady administratora"),
]
for row in summary:
    ws.append(row)
style(ws)

ws = wb.create_sheet("SCHARFER_20")
headers = [
    "PIM ID", "Model / indeks handlowy", "Nazwa TIM", "Indeks TIM", "EAN obecnie TIM", "EAN docelowy XML",
    "Cena TIM netto przed", "Cena TIM netto po", "Cena XML netto", "Cena zgodna", "Stan XML", "CE działa",
    "Karta działa", "Opis ma indeks SCH", "Opis ma EAN", "Aktywny", "Opublikowany", "EPREL", "Bloker / następny krok", "URL produktu",
]
ws.append(headers)
for model in NEW_EAN:
    live = live_by_model[model]
    before = live["before"]["identity"]
    after = live["after"]["identity"]
    xml = xml_rows[model]
    post_row = post_by_model[model]
    final_row = final_by_id[live["id"]]
    ce_ok = all(item["audit"]["downloadStatus"] == 200 for item in post_row["relations"]["certifications"])
    sheet_ok = all(item["audit"]["downloadStatus"] == 200 for item in post_row["relations"]["dataSheet"])
    price_equal = xml["price"] is not None and abs(before["price"] - xml["price"]) < 0.0001
    ws.append([
        live["id"], model, before["name"], before["timIndex"], before["ean"], NEW_EAN[model], before["price"],
        after["price"], xml["price"], "TAK" if price_equal else "NIE", xml["stock"], "TAK" if ce_ok else "NIE",
        "TAK" if sheet_ok else "NIE", "TAK" if model in final_row["descriptionHtml"] else "NIE",
        "TAK" if re.search(r"\b\d{13}\b", final_row["descriptionHtml"]) else "NIE",
        "TAK" if after["state"] == "active" else "NIE",
        "TAK" if after["published"] else "NIE", "NIE DOTYCZY — zasilacz LED",
        "TIM musi administracyjnie zmienić stary EAN na nowy; nie tworzyć duplikatu", xml["url"],
    ])
style(ws, url_columns=(20,))

ws = wb.create_sheet("DUPLIKATY_3")
ws.append(["PIM ID", "Model", "Nazwa TIM", "EAN", "Stan", "Zdjęcie", "Opis", "CE", "Karta", "Decyzja"])
duplicates = [row for row in final_audit["products"] if row["expectedBrand"] == "Scharfer" and not row["ean"]]
if len(duplicates) != 3:
    raise RuntimeError(f"Oczekiwano 3 duplikatów bez EAN, jest {len(duplicates)}")
for row in duplicates:
    ws.append([
        row["id"], row["model"], row["timName"], row["ean"], row["state"], "TAK" if row["mainPhoto"] else "NIE",
        "TAK" if row["descriptionHtml"] else "NIE", row["certifications"], row["dataSheet"],
        "Nie wycofano — potrzebna osobna zgoda na zmianę statusu/archiwizację",
    ])
style(ws)

ws = wb.create_sheet("DOWODY_I_BLOKERY")
ws.append(["Temat", "Dowód / wynik", "Plik źródłowy"])
validation = block.get("results", [{}])[0]
message_body = validation.get("saveResponseBody", "")
try:
    message = json.loads(message_body).get("message", message_body)
except (TypeError, json.JSONDecodeError):
    message = message_body
ws.append(["Blokada EAN TIM", message, str(EAN_BLOCK)])
ws.append(["Audyt przed", f"Unikalne aktywa: {pre['uniqueAssets']}; dostępne: {pre['downloadableAssets']}; niedostępne: {pre['unavailableAssets']}", str(PRE_AUDIT)])
ws.append(["Audyt po", f"Unikalne aktywa: {post['uniqueAssets']}; dostępne: {post['downloadableAssets']}; niedostępne: {post['unavailableAssets']}", str(POST_AUDIT)])
for path in LIVE_REPORTS:
    report = read_json(path)
    ws.append(["Zapis dokumentów", f"Zapisano: {sum(row['status'] == 'saved_and_verified' for row in report['results'])}; zablokowane operacje: {len(report['blockedWrites'])}; błąd: {report['fatalError'] or 'brak'}", str(path)])
for path in DESCRIPTION_REPORTS:
    report = read_json(path)
    ws.append(["Zapis opisów", f"Zapisano: {sum(row['status'] == 'saved_and_verified' for row in report['results'])}; zablokowane operacje: {len(report['blockedWrites'])}; błąd: {report['fatalError'] or 'brak'}", str(path)])
style(ws)

ws = wb.create_sheet("ŹRÓDŁA")
ws.append(["Źródło", "Zastosowanie"])
sources = [
    (str(XML_PATH), "Cena netto, stan i aktualny EAN dostawcy"),
    ("/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/scharfer 2026.07/scharfer nowe EAN - karty - tim - allegro - sklep.xls", "Potwierdzenie nowej węgierskiej serii EAN"),
    ("/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/TIM.PL/Karty katalogowe/ZASILACZE/Scharfer/Karty PL", "Źródłowe karty techniczne PL"),
    ("/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/TIM.PL/TIM - karty ce/Zasilacze LED/CE - SCHARFER- PL.pdf", "Deklaracja CE Scharfer"),
    (str(ROOT / "output/pdf/scharfer-tim-current-ean-2026-09-01"), "Karty pod aktywne, obecnie zablokowane EAN TIM"),
    (str(ROOT / "output/pdf/scharfer-new-ean-2026-09-01"), "Karty przygotowane pod docelowe nowe EAN dostawcy"),
]
for row in sources:
    ws.append(row)
style(ws)

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(OUTPUT)
