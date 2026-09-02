#!/usr/bin/env python3
"""Build the final, traceable TIM/KLUS progress workbook for 2026-09-01."""

from __future__ import annotations

import json
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/karolbohdanowicz/my-ai-agents/prescot")
DATA = ROOT / "exports/tim/remediation"
OUTPUT = Path("/Users/karolbohdanowicz/Downloads/TIM_RAPORT_POSTEPU_2026-09-01.xlsx")

PATHS = {
    "buffer": DATA / "buffer-current-live-final-progress-2026-09-01.json",
    "activation_queue": DATA / "klus-buffer-activation-queue-2026-09-01.json",
    "activation_verify": DATA / "klus-buffer-activation-final-postverify-2026-09-01.json",
    "document_queue": DATA / "klus-buffer-official-documents-safe-queue-2026-09-01.json",
    "document_verify": DATA / "klus-buffer-official-documents-safe-postverify-2026-09-01.json",
    "active_map": DATA / "klus-active-official-document-map-2026-09-01.json",
    "active_document_queue": DATA / "klus-active-official-documents-queue-2026-09-01.json",
    "active_document_verify": DATA / "klus-active-documents-final-postverify-2026-09-01.json",
    "active_description_queue": DATA / "klus-active-description-queue-2026-09-01.json",
    "active_description_write": DATA / "klus-active-descriptions-rest1282-live-2026-09-01.json",
    "active_description_verify": DATA / "klus-active-descriptions-positive-final-postverify-2026-09-01.json",
    "prescot_description_queue": DATA / "prescot-active-description-queue-2026-09-01.json",
    "prescot_description_verify": DATA / "prescot-active-descriptions-positive-final-postverify-2026-09-01.json",
    "prescot_eprel_queue": DATA / "prescot-active-eprel-remaining-queue-2026-09-01.json",
    "prescot_eprel_verify": DATA / "prescot-active-eprel-remaining-final-postverify-2026-09-01.json",
    "prescot_exact_ce_queue": DATA / "prescot-active-exact-accessory-ce11-queue-2026-09-01.json",
    "prescot_exact_ce_verify": DATA / "prescot-active-exact-accessory-ce11-final-postverify-2026-09-01.json",
    "prescot_family_ce_queue": DATA / "prescot-active-family-accessory-ce21-queue-2026-09-01.json",
    "prescot_family_ce_verify": DATA / "prescot-active-family-accessory-ce21-final-postverify-2026-09-01.json",
    "prescot_family_datasheet_queue": DATA / "prescot-active-family-datasheet8-queue-2026-09-01.json",
    "prescot_family_datasheet_safe_queue": DATA / "prescot-active-family-datasheet-eprel4-queue-2026-09-01.json",
    "prescot_family_datasheet_verify": DATA / "prescot-active-family-datasheet-eprel4-final-postverify-2026-09-01.json",
    "prescot_exact_datasheet_queue": DATA / "prescot-active-new-exact-datasheets4-queue-2026-09-01.json",
    "prescot_exact_datasheet_pilot": DATA / "prescot-active-new-exact-datasheets4-pilot-live-2026-09-01.json",
    "brand_audit": DATA / "active-brand-offer-live-readonly-post-scharfer-2026-09-01.json",
}

PRESCOT_WRITE_REPORTS = [
    DATA / "prescot-active-descriptions-pilot1-live-2026-09-01.json",
    DATA / "prescot-active-descriptions-batch10-live-2026-09-01.json",
    DATA / "prescot-active-descriptions-rest960-live-2026-09-01.json",
    DATA / "prescot-active-descriptions-rest702-live-2026-09-01.json",
]


def load(name: str):
    path = PATHS[name]
    if not path.exists():
        raise FileNotFoundError(f"Brak wymaganego pliku: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def style(ws, url_columns=()):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="8B174D")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in url_columns:
            cell = row[col - 1]
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"
    for column in ws.columns:
        width = min(65, max(11, max(len(str(cell.value or "")) for cell in list(column)[:500]) + 2))
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def relation_count(item, field):
    value = item.get(field)
    return len(value) if isinstance(value, list) else int(bool(value))


buffer = load("buffer")
activation_queue = load("activation_queue")
activation_verify = load("activation_verify")
document_queue = load("document_queue")
document_verify = load("document_verify")
active_map = load("active_map")
active_document_queue = load("active_document_queue")
active_document_verify = load("active_document_verify")
active_description_queue = load("active_description_queue")
active_description_write = load("active_description_write")
active_description_verify = load("active_description_verify")
prescot_description_queue = load("prescot_description_queue")
prescot_description_verify = load("prescot_description_verify")
prescot_eprel_queue = load("prescot_eprel_queue")
prescot_eprel_verify = load("prescot_eprel_verify")
prescot_exact_ce_queue = load("prescot_exact_ce_queue")
prescot_exact_ce_verify = load("prescot_exact_ce_verify")
prescot_family_ce_queue = load("prescot_family_ce_queue")
prescot_family_ce_verify = load("prescot_family_ce_verify")
prescot_family_datasheet_queue = load("prescot_family_datasheet_queue")
prescot_family_datasheet_safe_queue = load("prescot_family_datasheet_safe_queue")
prescot_family_datasheet_verify = load("prescot_family_datasheet_verify")
prescot_exact_datasheet_queue = load("prescot_exact_datasheet_queue")
prescot_exact_datasheet_pilot = load("prescot_exact_datasheet_pilot")
brand_audit = load("brand_audit")
prescot_write_reports = [json.loads(path.read_text(encoding="utf-8")) for path in PRESCOT_WRITE_REPORTS]

NEW_ACTIVATION_REPORT_PATHS = [
    DATA / "buffer-priority3-activation-final-postverify-2026-09-01.json",
    DATA / "buffer-ip67-priority3-activation-final-postverify-2026-09-01.json",
    DATA / "buffer-ec528-family4-activation-final-postverify-2026-09-01.json",
    DATA / "buffer-ec320-ww27-family1-activation-final-postverify-2026-09-01.json",
    DATA / "buffer-next-derived1-activation-final-postverify-2026-09-01.json",
]
NEW_EPREL_QUEUE_PATHS = [
    DATA / "buffer-eprel-priority-family3-queue-2026-09-01.json",
    DATA / "buffer-eprel-ip67-family3-queue-2026-09-01.json",
    DATA / "buffer-eprel-ec528-family5-queue-2026-09-01.json",
    DATA / "buffer-eprel-ec320-ww27-family1-queue-2026-09-01.json",
    DATA / "buffer-eprel-next-derived3-queue-2026-09-01.json",
]
NEW_DOCUMENT_QUEUE_PATHS = [
    DATA / "buffer-ec528-family4-documents-queue-2026-09-01.json",
    DATA / "buffer-ec320-ww27-family1-documents-queue-2026-09-01.json",
    DATA / "buffer-next-derived2-documents-queue-2026-09-01.json",
    DATA / "prescot-active-family-datasheet-24e009-queue-2026-09-01.json",
]
new_activation_reports = [json.loads(path.read_text(encoding="utf-8")) for path in NEW_ACTIVATION_REPORT_PATHS]
new_eprel_queues = [json.loads(path.read_text(encoding="utf-8")) for path in NEW_EPREL_QUEUE_PATHS]
new_document_queues = [json.loads(path.read_text(encoding="utf-8")) for path in NEW_DOCUMENT_QUEUE_PATHS]
current_buffer_audit = json.loads((DATA / "current-buffer-offer-audit-final-progress-2026-09-01.json").read_text(encoding="utf-8"))
ip67_repair = json.loads((DATA / "buffer-ip67-nw-energy-class-repair-postread-2026-09-01.json").read_text(encoding="utf-8"))
active_24e009_eprel_verify = json.loads((DATA / "prescot-active-eprel-24e009-repair2-final-postverify-2026-09-01.json").read_text(encoding="utf-8"))
active_24e009_datasheet_queue = json.loads((DATA / "prescot-active-family-datasheet-24e009-queue-2026-09-01.json").read_text(encoding="utf-8"))
active_24e009_datasheet_verify = json.loads((DATA / "prescot-active-family-datasheet-24e009-final-postverify-2026-09-01.json").read_text(encoding="utf-8"))

verify_counts = activation_verify["counts"]
if verify_counts.get("verified") != activation_queue["counts"]["ready"]:
    raise RuntimeError("Kontrola aktywacji nie potwierdza całej bezpiecznej kolejki.")
if verify_counts.get("mismatch") or verify_counts.get("locked"):
    raise RuntimeError("Kontrola aktywacji wykryła niezgodność lub blokadę.")
if document_verify["counts"].get("failed") or document_verify.get("blockedWrites"):
    raise RuntimeError("Kontrola dokumentów nie jest czysta.")
if active_document_verify["counts"].get("verified") != 352 or active_document_verify["counts"].get("failed") != 3:
    raise RuntimeError("Końcowa kontrola dokumentów aktywnego KLUŚ ma nieoczekiwany wynik.")
if active_document_verify.get("blockedWrites"):
    raise RuntimeError("Końcowa kontrola aktywnego KLUŚ wykryła blokadę.")
if active_description_verify["counts"].get("total") != active_description_queue["counts"].get("activePositiveNeedsUpdate"):
    raise RuntimeError("Końcowa kontrola opisów nie objęła całej kolejki aktywnych produktów ze stanem dodatnim.")
if active_description_verify["counts"].get("verified") != 1290 or active_description_verify["counts"].get("mismatch") != 3:
    raise RuntimeError("Końcowa kontrola opisów aktywnego KLUŚ ma nieoczekiwany wynik.")
if active_description_verify["counts"].get("failed") or active_description_verify["counts"].get("locked"):
    raise RuntimeError("Końcowa kontrola opisów wykryła błąd odczytu lub blokadę.")
if prescot_description_verify["counts"] != {"total": 971, "verified": 971, "mismatch": 0, "failed": 0, "locked": 0}:
    raise RuntimeError("Końcowa kontrola opisów aktywnego Prescot ma nieoczekiwany wynik.")
if prescot_eprel_verify["counts"].get("verified") != 5 or prescot_eprel_verify["counts"].get("failed"):
    raise RuntimeError("Końcowa kontrola nowej partii EPREL Prescot ma nieoczekiwany wynik.")
if prescot_exact_ce_verify["counts"].get("verified") != 11 or prescot_exact_ce_verify["counts"].get("failed"):
    raise RuntimeError("Końcowa kontrola dokładnych CE akcesoriów Prescot ma nieoczekiwany wynik.")
if prescot_family_ce_verify["counts"].get("verified") != 21 or prescot_family_ce_verify["counts"].get("failed"):
    raise RuntimeError("Końcowa kontrola rodzinnych CE akcesoriów Prescot ma nieoczekiwany wynik.")
if prescot_family_datasheet_verify["counts"].get("verified") != 4 or prescot_family_datasheet_verify["counts"].get("failed"):
    raise RuntimeError("Końcowa kontrola rodzinnych kart katalogowych Prescot ma nieoczekiwany wynik.")
if active_24e009_eprel_verify["counts"].get("verified") != 2 or active_24e009_eprel_verify["counts"].get("failed"):
    raise RuntimeError("Końcowa kontrola naprawy EPREL 24E009 ma nieoczekiwany wynik.")
if active_24e009_datasheet_verify["counts"].get("verified") != 2 or active_24e009_datasheet_verify["counts"].get("failed"):
    raise RuntimeError("Końcowa kontrola kart katalogowych 24E009 ma nieoczekiwany wynik.")

buffer_items = buffer.get("items", [])
klus_buffer = [x for x in buffer_items if str(x.get("manufacturerName") or "").upper() == "KLUŚ"]
excluded_reasons = Counter(x.get("reason") or "nieokreślony" for x in activation_queue["excluded"])
baseline_klus = brand_audit["counts"]["KLUŚ"]
active_desc_counts = active_description_queue["counts"]
prescot_desc_counts = prescot_description_queue["counts"]
prescot_validation_count = sum(report["counts"].get("savedWithValidation", 0) for report in prescot_write_reports)
new_activation_results = [row for report in new_activation_reports for row in report.get("results", [])]
new_activation_verified = sum(report["counts"].get("verified", 0) for report in new_activation_reports)
new_activation_active = sum(report["counts"].get("active", 0) for report in new_activation_reports)
new_activation_waiting = sum(report["counts"].get("awaitingTimApproval", 0) for report in new_activation_reports)
new_activation_mismatch = sum(report["counts"].get("mismatch", 0) for report in new_activation_reports)
new_eprel_items = [row for report in new_eprel_queues for row in report.get("items", [])]
new_document_items = [row for report in new_document_queues for row in report.get("items", [])]
if new_activation_verified != 12 or new_activation_mismatch:
    raise RuntimeError("Końcowa kontrola nowych aktywacji Prescot ma nieoczekiwany wynik.")
if ip67_repair.get("status") != "already_correct" or ip67_repair.get("before", {}).get("energyClass") != "G":
    raise RuntimeError("Końcowa kontrola korekty klasy IP67 nie potwierdza klasy G.")

wb = Workbook()
ws = wb.active
ws.title = "PODSUMOWANIE"
ws.append(["Pozycja", "Wartość", "Znaczenie"])
rows = [
    ("Data raportu", datetime.now().isoformat(timespec="seconds"), "Stan po operacjach i ponownym odczycie TIM/PIMCORE"),
    ("KLUŚ aktywowane z bufora", verify_counts["active"], "Każda karta przeszła kontrolę tożsamości, ceny, dokumentów i publikacji"),
    ("KLUŚ oczekujące na TIM", verify_counts["awaitingTimApproval"], "Workflow zakończony po naszej stronie, decyzja po stronie TIM"),
    ("Niezgodności aktywacji", verify_counts["mismatch"], "Musi wynosić 0"),
    ("Blokady po aktywacji", verify_counts["locked"], "Musi wynosić 0"),
    ("KLUŚ z dokumentami w buforze", document_verify["counts"]["verified"], "Oficjalne dokumenty KLUŚ; 303 relacje PDF potwierdzone"),
    ("Relacje PDF potwierdzone", document_verify["counts"]["downloadableRelations"], "Karty, instrukcje i deklaracje dostępne do pobrania"),
    ("KLUŚ wyłączone z automatycznej aktywacji", activation_queue["counts"]["excluded"], "Powody w arkuszu BLOKADY_I_BRAKI"),
    ("Aktualny bufor ogółem", len(buffer_items), "Pełny odczyt po aktywacjach"),
    ("Aktualny bufor KLUŚ", len(klus_buffer), "Pozostałe rekordy KLUŚ po aktywacjach"),
    ("Aktywna oferta KLUŚ — stan bazowy", baseline_klus["activePublished"], "Przed nowymi aktywacjami z bufora"),
    ("Oficjalne strony KLUŚ — dopasowanie dokładne", active_map["counts"]["exactPageMatch"], "Dowód producenta dla aktywnej oferty"),
    ("Aktywne KLUŚ gotowe do uzupełnienia dokumentów", active_document_queue["counts"]["items"], "Kolejka ma dokładną stronę producenta i zweryfikowany model"),
    ("Aktywne KLUŚ — dokumenty zweryfikowane", active_document_verify["counts"]["verified"], "Pełny ponowny odczyt kart i kontrola pobierania PDF"),
    ("Aktywne KLUŚ — relacje PDF działające", active_document_verify["counts"]["downloadableRelations"], "Karty, instrukcje i deklaracje dostępne z PIMCORE"),
    ("Aktywne KLUŚ — wyjątki", active_document_verify["counts"]["failed"], "2 duplikaty indeksu i 1 karta z ceną 0; bez wymuszania zapisu"),
    ("Aktywne KLUŚ do poprawy opisu — stan > 0", active_desc_counts["activePositiveNeedsUpdate"], "Opis bez EAN/PRE; indeks handlowy i fakty techniczne"),
    ("Aktywne KLUŚ — opisy zgodne po pełnym odczycie", active_description_verify["counts"]["verified"], "Opis obecny i identyczny z przygotowaną treścią; pełna kontrola 1293 kart"),
    ("Aktywne KLUŚ — opisy zablokowane przez TIM", active_description_verify["counts"]["mismatch"], "3 warianty MI-BUBBLE-04; walidator B24 wymaga danych energetycznych mimo wersji bez LED"),
    ("Aktywne KLUŚ — zapisy z ostrzeżeniem walidatora", active_description_write["counts"]["savedWithValidation"], "Opis potwierdzony odczytem mimo komunikatu TIM; bez fałszywych danych EPREL"),
    ("Aktywne KLUŚ do poprawy opisu — stan 0", active_desc_counts["activeZeroNeedsUpdate"], "Niższy priorytet po ofercie dostępnej"),
    ("Aktywne Prescot — opisy zgodne po pełnym odczycie", prescot_description_verify["counts"]["verified"], "971/971 aktywnych kart ze stanem dodatnim; dokładne EAN i indeks handlowy"),
    ("Aktywne Prescot — rozbieżności opisów", prescot_description_verify["counts"]["mismatch"], "Pełny ponowny odczyt TIM/PIMCORE"),
    ("Aktywne Prescot — zapisy z ostrzeżeniem walidatora", prescot_validation_count, "Opis potwierdzony w końcowym odczycie mimo komunikatu TIM"),
    ("Prescot — opisy do poprawy przy stanie 0", prescot_desc_counts["activeZeroNeedsUpdate"], "Niższy priorytet po ofercie dostępnej"),
    ("Prescot — odrzucone z automatu", prescot_desc_counts["rejected"], "Brak jednoznacznej tożsamości, brak EAN w katalogu, różny indeks lub marka wykluczona"),
    ("Aktywne Prescot — nowa partia EPREL", prescot_eprel_verify["counts"]["verified"] + active_24e009_eprel_verify["counts"]["verified"], "2 exact 100% i 5 wariantów długości 90%; 14 relacji pobieralnych"),
    ("Aktywne Prescot — dokładne CE akcesoriów", prescot_exact_ce_verify["counts"]["verified"], "Dokładny model w deklaracji CE/PL/02/AKC/2026 z 11.07.2026"),
    ("Aktywne Prescot — rodzinne CE akcesoriów", prescot_family_ce_verify["counts"]["verified"], "Warianty jawnie objęte wzorcami XX w deklaracji; dopasowanie 90% zatwierdzone do raportowania"),
    ("Aktywne Prescot — rodzinne karty katalogowe zweryfikowane", prescot_family_datasheet_verify["counts"]["verified"] + active_24e009_datasheet_verify["counts"]["verified"], "6/8 produktów; PDF pobieralne po ponownym odczycie"),
    ("Aktywne Prescot — rodzinne karty z blokadą energii", prescot_family_datasheet_queue["counts"]["items"] - prescot_family_datasheet_verify["counts"]["verified"] - active_24e009_datasheet_verify["counts"]["verified"], "Pozostały 2 produkty bez pełnego, bezpiecznego kompletu danych"),
    ("EPREL 2321138 — blokady", len(prescot_eprel_queue.get("rejected", [])), "Oficjalny serwer zwraca HTML zamiast etykiety i PDF; bez fałszywych załączników"),
    ("Nowe exact karty katalogowe — blokada pilota", 4, "OPB-RO5: HTTP 500 przy zapisie relacji; pozostałych 3 nie uruchomiono po nieudanym pilocie"),
    ("Nowe karty Prescot z bufora — zweryfikowane", new_activation_verified, "12/12 po kontroli tożsamości, ceny, opisu i wymaganych dokumentów"),
    ("Nowe karty Prescot — aktywne na TIM", new_activation_active, "Aktywowane automatycznie przez workflow TIM"),
    ("Nowe karty Prescot — oczekują na TIM", new_activation_waiting, "Workflow wysłany poprawnie; karta pozostaje new_for_approval"),
    ("Nowe karty Prescot — rozbieżności", new_activation_mismatch, "Musi wynosić 0"),
    ("Nowe komplety EPREL w buforze", len(new_eprel_items), "Dopasowania 90%, oficjalne etykiety i karty EPREL; warianty długości raportowane jawnie"),
    ("Nowe operacje dokumentowe w buforze", sum(len(row.get("documents", {})) for row in new_document_items), "Karty katalogowe i CE przypięte do bieżącej partii"),
    ("Korekta klasy 24EC320NW50IP67", "F → G", "Oficjalna karta EPREL 2724835 wskazuje G; pozostałe pola chronione bez zmian"),
    ("Bufor po bieżącej partii", current_buffer_audit["counts"]["total"], "Pełny ponowny odczyt PIMCORE, 0 błędów odczytu"),
    ("Bufor — oczekujące na akceptację TIM", current_buffer_audit["counts"]["newForApproval"], "Nie wymagają ponownego wysyłania"),
    ("Bufor — gotowe bez żadnych braków", current_buffer_audit["counts"]["safeActivationCandidates"], "Obecnie 0; pozostałe wymagają danych lub decyzji TIM"),
    ("Ceny", "NIE ZMIENIANO", "Workflow porównywał cenę TIM z XML i blokował niezgodności"),
    ("EAN/nazwy/stany magazynowe", "NIE ZMIENIANO", "Brak migracji EAN i brak zmian nazw w tej partii"),
    ("Ocena TIM widoczna", 3, "Wzrost zależy od cyklu ponownego naliczania TIM; nie jest deklarowany przed aktualizacją panelu"),
]
for row in rows:
    ws.append(row)
style(ws)

ws = wb.create_sheet("AKTYWOWANE_KLUS")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Indeks TIM", "Stan live", "Cena netto", "Wymagane dokumenty", "Weryfikacja"])
for x in activation_verify["results"]:
    ws.append([
        x["id"], x["ean"], x["model"], x.get("timIndex"), x.get("liveState"), x.get("livePrice"),
        ", ".join(x.get("requiredRelations", [])), x.get("status"),
    ])
style(ws)

ws = wb.create_sheet("DOKUMENTY_KLUS")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Stan", "Nazwa TIM", "Karta", "Instrukcja", "CE/deklaracja", "Oficjalna strona KLUŚ"])
for x in document_queue["items"]:
    docs = x.get("documents", {})
    ws.append([
        x["id"], x["ean"], x["model"], x["state"], x["timName"],
        "TAK" if docs.get("dataSheet") else "NIE",
        "TAK" if docs.get("instructions") else "NIE",
        "TAK" if docs.get("certifications") else "NIE",
        x.get("officialProductUrl", ""),
    ])
style(ws, url_columns=(9,))

ws = wb.create_sheet("BLOKADY_I_BRAKI")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Powód", "Cena TIM", "Cena XML", "Stan XML", "Aktywne duplikaty", "Decyzja"])
for x in activation_queue["excluded"]:
    xml = x.get("xml") or {}
    duplicates = x.get("activeDuplicates") or []
    ws.append([
        x.get("id"), x.get("ean"), x.get("model"), x.get("reason"), x.get("timPrice"), xml.get("price"), xml.get("stock"),
        ", ".join(str(d.get("id") if isinstance(d, dict) else d) for d in duplicates),
        "Nie aktywować automatycznie; uzupełnić lub rozstrzygnąć przyczynę",
    ])
style(ws)

ws = wb.create_sheet("POWODY_BLOKAD")
ws.append(["Powód", "Liczba", "Znaczenie"])
explanations = {
    "active_model_duplicate": "Istnieje aktywna karta z tym indeksem; nie tworzyć duplikatu",
    "xml_price_not_positive": "Cena XML wynosi 0; aktywacja zablokowana",
    "xml_stock_not_positive": "Stan XML nie jest dodatni; brak priorytetu sprzedażowego",
    "verified_catalog_card_missing": "Brak dokładnie zweryfikowanej karty katalogowej",
    "ean_missing_in_xml": "EAN bufora nie został znaleziony w źródłowym XML",
    "state_new_for_approval": "Karta już czeka w workflow TIM",
    "state_active": "Niespójny rekord bufora oznaczony jako aktywny",
}
for reason, count in sorted(excluded_reasons.items(), key=lambda x: (-x[1], x[0])):
    ws.append([reason, count, explanations.get(reason, "Wymaga kontroli indywidualnej")])
style(ws)

ws = wb.create_sheet("BUFOR_AKTUALNY")
ws.append(["PIM ID", "Producent", "Stan", "EAN", "Indeks handlowy", "Nazwa TIM", "Cena netto", "Zdjęcie", "Opis", "Karta", "CE", "Instrukcja", "Etykieta EPREL", "Karta EPREL"])
for x in buffer_items:
    ws.append([
        x.get("id"), x.get("manufacturerName"), x.get("state"), x.get("ean"), x.get("model"), x.get("timName"),
        (x.get("listPrice") or {}).get("value") if isinstance(x.get("listPrice"), dict) else x.get("listPrice"),
        "TAK" if x.get("mainPhoto") else "NIE", "TAK" if x.get("descriptionHtml") else "NIE",
        relation_count(x, "dataSheet"), relation_count(x, "certifications"), relation_count(x, "instructions"),
        relation_count(x, "energyClassLabels"), relation_count(x, "energyTechnicalCards"),
    ])
style(ws)

ws = wb.create_sheet("AKTYWNY_KLUS_PLAN")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Nazwa TIM", "Karta do dodania", "Instrukcja do dodania", "CE do dodania", "Oficjalna strona KLUŚ"])
for x in active_document_queue["items"]:
    docs = x.get("documents", {})
    ws.append([
        x["id"], x["ean"], x["model"], x["timName"],
        "TAK" if docs.get("dataSheet") else "NIE",
        "TAK" if docs.get("instructions") else "NIE",
        "TAK" if docs.get("certifications") else "NIE",
        x.get("officialProductUrl", ""),
    ])
style(ws, url_columns=(8,))

ws = wb.create_sheet("AKTYWNY_KLUS_WYNIK")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Indeks TIM", "Stan", "Cena netto", "Karta", "Instrukcja", "CE", "Weryfikacja", "Przyczyna wyjątku"])
failure_reasons = {
    2122722: "Duplikat indeksu A01888N_3; TIM wskazał istniejący produkt 0001-00017-49612",
    2122770: "Aktywna karta ma cenę 0; walidator TIM blokuje zapis dokumentów bez ceny",
    10646118: "Duplikat indeksu C28284C02 według walidatora TIM",
}
for x in active_document_verify["products"]:
    live = x.get("live") or {}
    downloads = live.get("downloads") or {}
    price = live.get("price") or {}
    ws.append([
        x.get("id"), x.get("ean"), x.get("model"), live.get("timIndex"), live.get("state"),
        price.get("value") if isinstance(price, dict) else price,
        len(downloads.get("dataSheet") or []), len(downloads.get("instructions") or []),
        len(downloads.get("certifications") or []), "OK" if x.get("verified") else "WYJĄTEK",
        failure_reasons.get(x.get("id"), ""),
    ])
style(ws)

ws = wb.create_sheet("OPISY_KLUS_WYNIK")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Indeks TIM", "Stan live", "Stan magazynowy", "Cena live", "Długość opisu", "Weryfikacja", "Przyczyna"])
for x in active_description_verify["results"]:
    ws.append([
        x.get("objectId"), x.get("ean"), x.get("manufacturerCode"), x.get("timIndex"), x.get("liveState"),
        x.get("liveStock"), x.get("livePrice"), x.get("descriptionLength"),
        "OK" if x.get("status") == "verified" else "BLOKADA TIM", x.get("reason", ""),
    ])
style(ws)

write_exceptions = [x for x in active_description_write["results"] if x.get("status") != "saved"]
ws = wb.create_sheet("OPISY_WALIDACJE")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Indeks TIM", "Nazwa TIM", "Wynik zapisu", "HTTP", "Wersja przed", "Wersja po", "Powód / komunikat TIM"])
for x in write_exceptions:
    validation = x.get("validationResponse") or x.get("saveResponseBody") or x.get("reason") or ""
    ws.append([
        x.get("objectId"), x.get("ean"), x.get("manufacturerCode"), x.get("timIndex"), x.get("name"),
        x.get("status"), x.get("httpStatus") or x.get("saveResponseStatus"), x.get("beforeVersionCount"),
        x.get("afterVersionCount"), validation,
    ])
style(ws)

ws = wb.create_sheet("OPISY_PRESCOT_WYNIK")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Indeks TIM", "Stan live", "Stan magazynowy", "Cena live", "Długość opisu", "Weryfikacja", "Przyczyna"])
for x in prescot_description_verify["results"]:
    ws.append([
        x.get("objectId"), x.get("ean"), x.get("manufacturerCode"), x.get("timIndex"), x.get("liveState"),
        x.get("liveStock"), x.get("livePrice"), x.get("descriptionLength"),
        "OK" if x.get("status") == "verified" else "WYJĄTEK", x.get("reason", ""),
    ])
style(ws)

prescot_write_exceptions = []
for report in prescot_write_reports:
    prescot_write_exceptions.extend(x for x in report["results"] if x.get("status") != "saved")
ws = wb.create_sheet("OPISY_PRESCOT_WALIDACJE")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Indeks TIM", "Nazwa TIM", "Wynik zapisu", "HTTP", "Wersja przed", "Wersja po", "Wynik końcowy", "Komunikat TIM"])
for x in prescot_write_exceptions:
    validation = x.get("validationResponse") or x.get("saveResponseBody") or x.get("reason") or ""
    ws.append([
        x.get("objectId"), x.get("ean"), x.get("manufacturerCode"), x.get("timIndex"), x.get("name"),
        x.get("status"), x.get("httpStatus") or x.get("saveResponseStatus"), x.get("beforeVersionCount"),
        x.get("afterVersionCount"), "POTWIERDZONY W AUDYCIE 971/971", validation,
    ])
style(ws)

ws = wb.create_sheet("PRESCOT_ODRZUCONE")
ws.append(["PIM ID", "EAN", "Indeks live", "Indeks katalogu", "Nazwa TIM", "Powód", "Decyzja"])
for x in prescot_description_queue["rejected"]:
    ws.append([
        x.get("id"), x.get("ean"), x.get("model"), x.get("catalogTradeIndex"), x.get("timName"), x.get("reason"),
        "Nie zmieniać automatycznie; wymaga jednoznacznego dopasowania danych",
    ])
style(ws)

ws = wb.create_sheet("CENY_NIE_RUSZANE")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Cena live netto", "Kontrola", "Uwagi"])
for x in activation_verify["results"]:
    ws.append([x["id"], x["ean"], x["model"], x.get("livePrice"), "OK" if x.get("checks", {}).get("price") else "BŁĄD", "Cena nie była zmieniana"])
style(ws)

ws = wb.create_sheet("EPREL_PRESCOT_01_09")
ws.append(["PIM ID", "EAN", "Indeks TIM", "Model EPREL", "EPREL ID", "Klasa", "Typ dopasowania", "Pewność %", "Cena netto", "Stan XML", "Weryfikacja"])
verified_eprel_ids = {int(row["expected"]["pimcoreId"]) for row in prescot_eprel_verify["products"] if row.get("verified")}
verified_eprel_ids.update(int(row["expected"]["pimcoreId"]) for row in active_24e009_eprel_verify["products"] if row.get("verified"))
for row in prescot_eprel_queue["items"]:
    ws.append([
        row["pimcoreId"], row["ean"], row["manufacturerCode"], row.get("eprelModel"), row["eprelId"],
        row["energyClass"], row.get("matchType"), row.get("confidence"), row.get("timListPrice"), row.get("xmlStock"),
        "OK" if int(row["pimcoreId"]) in verified_eprel_ids else "BRAK",
    ])
style(ws)

ws = wb.create_sheet("CE_PRESCOT_AKCESORIA")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Cena TIM netto", "Cena XML netto", "Stan XML", "Deklaracja", "Typ dopasowania", "Pewność %", "Weryfikacja"])
verified_ce_ids = {int(row["id"]) for row in prescot_exact_ce_verify["products"] if row.get("verified")}
for row in prescot_exact_ce_queue["items"]:
    ws.append([
        row["id"], row["ean"], row["model"], row["timListPrice"], row["xmlPrice"], row["xmlStock"],
        "CE/PL/02/AKC/2026; 11.07.2026", row.get("matchType"), row.get("confidence"),
        "OK" if int(row["id"]) in verified_ce_ids else "BRAK",
    ])
verified_family_ce_ids = {int(row["id"]) for row in prescot_family_ce_verify["products"] if row.get("verified")}
for row in prescot_family_ce_queue["items"]:
    ws.append([
        row["id"], row["ean"], row["model"], row["timListPrice"], row["xmlPrice"], row["xmlStock"],
        "CE/PL/02/AKC/2026; 11.07.2026", f"{row.get('matchType')} / {row.get('declaredFamily')}", row.get("confidence"),
        "OK" if int(row["id"]) in verified_family_ce_ids else "BRAK",
    ])
style(ws)

ws = wb.create_sheet("PRESCOT_KARTY_RODZINNE8")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Wzorzec karty", "Cena TIM netto", "Cena XML netto", "Stan XML", "Pewność %", "Status"])
verified_family_sheet_ids = {int(row["id"]) for row in prescot_family_datasheet_verify["products"] if row.get("verified")}
verified_family_sheet_ids.update(int(row["id"]) for row in active_24e009_datasheet_verify["products"] if row.get("verified"))
for row in prescot_family_datasheet_queue["items"]:
    ws.append([
        row["id"], row["ean"], row["model"], row.get("cardPattern"), row["timListPrice"], row["xmlPrice"], row["xmlStock"],
        row.get("confidence"), "OK — ZWERYFIKOWANE" if int(row["id"]) in verified_family_sheet_ids else "BLOKADA — BRAK PEŁNEGO EPREL",
    ])
style(ws)

ws = wb.create_sheet("PRESCOT_DOK_NOWE_BLOKADY")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Dokument", "Pewność %", "Wynik", "Powód / decyzja"])
pilot_by_id = {int(row["id"]): row for row in prescot_exact_datasheet_pilot.get("results", [])}
for row in prescot_exact_datasheet_queue["items"]:
    pilot = pilot_by_id.get(int(row["id"]))
    ws.append([
        row["id"], row["ean"], row["model"], row["documents"]["dataSheet"]["filename"], row.get("confidence"),
        pilot.get("status") if pilot else "NIE URUCHOMIONO PO PILOCIE",
        pilot.get("reason") if pilot else "Seria wstrzymana po HTTP 500 pilota OPB-RO5",
    ])
for row in prescot_eprel_queue.get("rejected", []):
    ws.append([
        row.get("id"), row.get("ean"), row.get("model"), f"EPREL {row.get('eprelId', '')}", row.get("confidence"),
        "BLOKADA", f"{row.get('reason', '')}: {row.get('detail', '')}",
    ])
style(ws)

ws = wb.create_sheet("AKTYWACJE_PRESCOT_BUFOR")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Indeks TIM", "Cena netto", "Stan po workflow", "Karta", "CE", "Etykieta EPREL", "Karta EPREL", "Weryfikacja"])
for row in new_activation_results:
    checks = row.get("checks", {})
    ws.append([
        row.get("id"), row.get("ean"), row.get("model"), row.get("timIndex"), row.get("livePrice"), row.get("liveState"),
        "TAK" if checks.get("dataSheet") else "NIE",
        "TAK" if checks.get("certifications") else "NIE",
        "TAK" if checks.get("energyClassLabels") else "NIE",
        "TAK" if checks.get("energyTechnicalCards") else "NIE",
        "OK" if row.get("status") == "verified" else row.get("status"),
    ])
style(ws)

activation_by_id = {int(row["id"]): row for row in new_activation_results}
ws = wb.create_sheet("EPREL_BUFOR_80_100")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Model EPREL", "EPREL ID", "Klasa", "Typ dopasowania", "Pewność %", "Cena netto", "Stan XML", "Stan TIM / decyzja"])
for row in new_eprel_items:
    activation = activation_by_id.get(int(row["pimcoreId"]))
    status = activation.get("liveState") if activation else "EPREL UZUPEŁNIONY — BRAK PEŁNEJ KARTY DO AKTYWACJI"
    ws.append([
        row["pimcoreId"], row["ean"], row["manufacturerCode"], row.get("eprelModel"), row.get("eprelId"),
        row.get("energyClass"), row.get("matchType"), row.get("confidence"), row.get("timListPrice"), row.get("xmlStock"), status,
    ])
style(ws)

ws = wb.create_sheet("DOKUMENTY_BUFOR_NOWE")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Karta katalogowa", "CE / deklaracja", "Pewność %", "Uwagi"])
for row in new_document_items:
    documents = row.get("documents", {})
    note = ""
    if row.get("model") == "EC608-013-8-CCT50":
        note = "CE + EPREL uzupełnione; brak właściwej karty 8 mm, nie użyto błędnej karty 5 mm"
    ws.append([
        row.get("id"), row.get("ean"), row.get("model"),
        documents.get("dataSheet", {}).get("filename", "BRAK"),
        documents.get("certifications", {}).get("filename", "BRAK"),
        row.get("confidence"), note,
    ])
style(ws)

ws = wb.create_sheet("BUFOR_BRAKI_FINAL")
ws.append(["PIM ID", "Marka", "Stan", "EAN", "Indeks handlowy", "Cena TIM", "Cena XML", "Stan XML", "Karta", "CE", "EPREL komplet", "Blokady", "Decyzja"])
for row in current_buffer_audit["items"]:
    xml = row.get("xml") or {}
    blockers = row.get("blockers") or []
    if row.get("safeActivationCandidate"):
        decision = "GOTOWY"
    elif row.get("state") == "new_for_approval":
        decision = "CZEKA NA TIM — NIE WYSYŁAĆ PONOWNIE"
    elif "active_duplicate" in blockers:
        decision = "NIE AKTYWOWAĆ — DUPLIKAT AKTYWNEJ KARTY"
    elif "xml_stock_not_positive" in blockers or "xml_price_not_positive" in blockers:
        decision = "NIE AKTYWOWAĆ — BRAK AKTYWNEJ OFERTY XML"
    else:
        decision = "UZUPEŁNIĆ WSKAZANE BRAKI"
    ws.append([
        row.get("id"), row.get("brand"), row.get("state"), row.get("ean"), row.get("model"), row.get("timPrice"),
        xml.get("price"), xml.get("stock"), row.get("relations", {}).get("dataSheet"), row.get("relations", {}).get("certifications"),
        "TAK" if row.get("energyComplete") else "NIE", ", ".join(blockers), decision,
    ])
style(ws)

ws = wb.create_sheet("POWODY_BUFOR_FINAL")
ws.append(["Powód", "Liczba", "Znaczenie"])
buffer_blocker_explanations = {
    "catalog_card_missing": "Brak bezpiecznie dopasowanej karty katalogowej",
    "required_energy_set_missing": "Dla taśmy brakuje pełnego kompletu: klasa + etykieta + karta EPREL",
    "state_new_for_approval": "Karta została wysłana i czeka na decyzję TIM",
    "active_duplicate": "Istnieje już aktywna karta; nie wolno tworzyć duplikatu",
    "xml_stock_not_positive": "Brak dodatniego stanu w źródłowym XML",
    "xml_price_not_positive": "Brak dodatniej ceny w źródłowym XML",
    "tim_xml_price_mismatch": "Cena TIM różni się od ceny netto XML; bez zapisu ceny",
    "description_identity_guard": "Opis nie przechodzi kontroli indeksu handlowego lub zawiera niedozwolone dane",
}
for reason, count in sorted(current_buffer_audit["blockerCounts"].items(), key=lambda item: (-item[1], item[0])):
    ws.append([reason, count, buffer_blocker_explanations.get(reason, "Wymaga kontroli danych źródłowych")])
style(ws)

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
check = load_workbook(OUTPUT, read_only=True, data_only=True)
sheet_counts = {sheet.title: max(0, sheet.max_row - 1) for sheet in check.worksheets}
check.close()
print(json.dumps({"output": str(OUTPUT), "sheets": sheet_counts}, ensure_ascii=False, indent=2))
