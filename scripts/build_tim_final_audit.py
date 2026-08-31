#!/usr/bin/env python3
"""Build the final, evidence-backed TIM/PRESCOT audit workbook."""

from __future__ import annotations

import argparse
import glob
import html
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def load(path: str | Path, default: Any = None) -> Any:
    candidate = Path(path)
    if not candidate.exists():
        return default
    return json.loads(candidate.read_text(encoding="utf-8"))


def number(value: Any) -> float:
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def price_status(source_price: Any, live_price: Any) -> str:
    if live_price is None or isinstance(live_price, (dict, list, tuple, set)):
        return "brak_odczytu"
    source = number(source_price)
    live = number(live_price)
    if source <= 0:
        return "brak_ceny_w_prescot.xml"
    if live <= 0:
        return "brak_ceny_live_TIM"
    if abs(live - source) < 0.005:
        return "zgodna_1_do_1"
    if abs(live - round(source / 1.23, 2)) < 0.011:
        return "zgodna_po_VAT_23"
    return "różnica"


def plain(value: Any) -> str:
    text = re.sub(r"<[^>]+>", " ", str(value or ""))
    return re.sub(r"\s+", " ", html.unescape(text)).strip()


def parse_buffer(audit: dict[str, Any]) -> list[dict[str, Any]]:
    try:
        return json.loads(audit.get("pimcoreGet", {}).get("body", "{}")).get("nodes", [])
    except (json.JSONDecodeError, AttributeError):
        return []


def add_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append([
            json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list, tuple, set)) else value
            for value in row
        ])
    header_fill = PatternFill("solid", fgColor="8B164B")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 32
    for index, header in enumerate(headers, 1):
        values = [str(header)] + [str(sheet.cell(row=row, column=index).value or "") for row in range(2, min(sheet.max_row, 250) + 1)]
        width = min(58, max(10, max(len(value) for value in values) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--manifest", default="exports/tim/tim-manifest.json")
    parser.add_argument("--wapro-catalog", required=True)
    parser.add_argument("--base-queue", default="exports/tim/remediation/full-description-queue-v4.json")
    parser.add_argument("--supplemental-queue", default="exports/tim/remediation/supplemental-description-queue-v4.json")
    parser.add_argument("--buffer-audit", default="/tmp/tim-live-buffer-prescot-complete.json")
    parser.add_argument("--cloud-audit", default="/tmp/prescot-cloud-audit-2026-08-31.json")
    parser.add_argument(
        "--legacy-document-audit",
        default="",
    )
    parser.add_argument(
        "--ce-asset-map",
        default="",
    )
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    manifest = load(args.manifest, {"products": [], "meta": {}})
    wapro = load(args.wapro_catalog, {"products": [], "meta": {}})
    base = load(args.base_queue, {"stages": {}, "counts": {}})
    supplemental = load(args.supplemental_queue, {"stages": {}, "counts": {}})
    cloud = load(args.cloud_audit, {})
    buffer_nodes = parse_buffer(load(args.buffer_audit, {}))
    ce_asset_map = load(args.ce_asset_map, {}) if args.ce_asset_map else {}
    ce_asset_rows = [
        [filename, details.get("id", ""), details.get("path", ""), details.get("type", ""), "aktywo PIMCORE z mapy 28.08.2026"]
        for filename, details in sorted(ce_asset_map.items())
    ]
    document_audit_rows: list[list[Any]] = []
    document_audit_summary: dict[str, int] = {}
    document_audit_path = Path(args.legacy_document_audit) if args.legacy_document_audit else None
    if document_audit_path and document_audit_path.exists():
        source_workbook = load_workbook(document_audit_path, read_only=True, data_only=True)
        source_sheet = source_workbook["Wszystkie_5980_Produktow"]
        source_headers = [str(cell.value or "") for cell in source_sheet[1]]
        for values in source_sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(source_headers, values))
            document_audit_rows.append([
                row.get("ID", ""), row.get("Indeks TIM", ""), row.get("Indeks Producenta", ""),
                row.get("Nazwa Produktu", ""), row.get("Producent", ""), row.get("Liczba Braków", ""),
                row.get("Lista Braków", ""), row.get("Brak Karty PDF", ""), row.get("Brak Certyfikatu CE", ""),
                row.get("Brak Zdjęcia", ""), row.get("Brak ETIM", ""), row.get("Brak Ceny", ""),
                row.get("Brak Nazwy", ""), row.get("Cena Netto TIM", ""), row.get("Dostępny do sprzedaży", ""),
                row.get("Opublikowany", ""),
            ])
        document_audit_summary = {
            "all": len(document_audit_rows),
            "complete": sum(number(row[5]) == 0 for row in document_audit_rows),
            "missingCard": sum(str(row[7]).upper() == "TAK" for row in document_audit_rows),
            "missingCe": sum(str(row[8]).upper() == "TAK" for row in document_audit_rows),
            "missingImage": sum(str(row[9]).upper() == "TAK" for row in document_audit_rows),
        }
        source_workbook.close()
    eprel_queue = load("exports/tim/remediation/eprel-exact-documents-queue.json", {"items": []})
    eprel_live = load("/tmp/tim-eprel-docs-live-verification2.json", {"results": []})
    missing_sheets = load("/tmp/tim-missing-product-sheet-validation-with-urls.json", [])
    final_name_queue = load("exports/tim/remediation/final-name-queue.json", {"stages": {}})
    name_positive = load("/tmp/tim-final-name-positive-verification.json", {"results": []})
    name_zero = load("/tmp/tim-final-name-zero-verification.json", {"results": []})

    verification_files = sorted(set(
        glob.glob("/tmp/tim-description-v4-*-verification.json")
        + glob.glob("/tmp/tim-description-supp-*-verification.json")
        + glob.glob("/tmp/tim-description-buffer-*-verification.json")
    ))
    description_results: dict[int, dict[str, Any]] = {}
    for path in verification_files:
        for row in load(path, {}).get("results", []):
            row = {**row, "verificationFile": path}
            description_results[int(row.get("objectId") or 0)] = row

    products = manifest.get("products", [])
    product_by_ean = {str(row.get("ean", "")): row for row in products if row.get("ean")}
    wapro_by_key = {str(row.get("key", "")): row for row in wapro.get("products", []) if row.get("key")}
    wapro_by_ean = {str(row.get("ean", "")): row for row in wapro.get("products", []) if row.get("ean")}
    def source_for(row: dict[str, Any]) -> dict[str, Any]:
        source = wapro_by_key.get(str(row.get("productKey") or row.get("key") or ""))
        if not source:
            source = wapro_by_ean.get(str(row.get("ean", "")))
        if source:
            return source
        # Manifest już zawiera pola handlowe przeniesione 1:1 z prescot.xml.
        # Ten fallback zachowuje pozycje bez EAN i nie sięga do źródła Mamezi.
        return {
            "price": row.get("price", ""),
            "stock": row.get("stock", ""),
            "attributes": {"Jednostka": row.get("measureUnit", "")},
        }

    mapped: dict[str, tuple[str, dict[str, Any]]] = {}
    target_object_ids: set[int] = set()
    for prefix, queue in (("bazowa", base), ("uzupełniająca", supplemental)):
        for stage, rows in queue.get("stages", {}).items():
            for row in rows:
                if row.get("ean"):
                    mapped[str(row["ean"])] = (f"{prefix}:{stage}", row)
                if stage in {"activePositiveNeedsUpdate", "activeZeroNeedsUpdate", "bufferNewNeedsUpdate", "bufferApprovalNeedsUpdate"} and row.get("pimcoreId"):
                    target_object_ids.add(int(row["pimcoreId"]))

    verified = sum(row.get("status") == "verified" for row in description_results.values())
    description_mismatch = sum(row.get("status") == "mismatch" for row in description_results.values())
    description_failed = sum(row.get("status") == "failed" for row in description_results.values())
    name_verified = sum(row.get("status") == "verified" for row in name_positive.get("results", []) + name_zero.get("results", []))
    name_mismatch = sum(row.get("status") == "mismatch" for row in name_positive.get("results", []) + name_zero.get("results", []))
    eprel_verified = sum(row.get("status") == "already_current" for row in eprel_live.get("results", []))
    source_positive = [row for row in products if number(source_for(row).get("stock")) > 0]
    price_compared = 0
    price_mismatch = 0
    price_missing_live = 0
    price_zero_live = 0
    price_missing_source = 0
    price_vat_converted = 0
    price_direct = 0
    for row in description_results.values():
        source = wapro_by_ean.get(str(row.get("ean", "")), {})
        if not source:
            continue
        status = price_status(source.get("price"), row.get("livePrice"))
        if status == "brak_odczytu":
            price_missing_live += 1
            continue
        if status == "brak_ceny_live_TIM":
            price_zero_live += 1
            continue
        if status == "brak_ceny_w_prescot.xml":
            price_missing_source += 1
            continue
        price_compared += 1
        if status == "zgodna_1_do_1":
            price_direct += 1
        elif status == "zgodna_po_VAT_23":
            price_vat_converted += 1
        else:
            price_mismatch += 1

    summary_rows = [
        ["Zakres", "Produkty w kontrolowanym zakresie TIM", len(products), "OK", "Kaja i Light Prestige wykluczone"],
        ["WAPRO", "Aktywne oferty w prescot.xml", wapro.get("meta", {}).get("activeProducts", 0), "OK", "Cena, stan i jednostka dla TIM"],
        ["WAPRO", "Zakres TIM ze stanem dodatnim w prescot.xml", len(source_positive), "PRIORYTET", "W pierwszej kolejności"],
        ["Cena", "Karty z dodatnią ceną porównane live z prescot.xml", price_compared, "ODCZYT", f"1:1={price_direct}, po VAT 23%={price_vat_converted}, różnice={price_mismatch}; cena live 0={price_zero_live}, brak odczytu={price_missing_live}, brak ceny źródłowej={price_missing_source}; nie zmieniano ceny ręcznie"],
        ["TIM", "Unikalne karty dokładnie dopasowane i objęte finalną kolejką", len(target_object_ids), "OK", "EAN + indeks handlowy + stan/path; zdublowany rekord źródłowy liczony raz"],
        ["Opisy", "Opisy potwierdzone odczytem live", verified, "OK" if not description_mismatch and not description_failed else "UWAGA", f"mismatch={description_mismatch}, failed={description_failed}"],
        ["Bufor", "Karty w folderze dostawcy", len(buffer_nodes), "ODCZYT", "Bez zmiany statusu/workflow"],
        ["Bufor", "Dokładne karty bufora z opisem w kolejce", len(supplemental.get("stages", {}).get("bufferNewNeedsUpdate", [])) + len(supplemental.get("stages", {}).get("bufferApprovalNeedsUpdate", [])), "OK", "Stan new / new_for_approval zachowany"],
        ["Nowości", "Brak dokładnej karty w wyszukiwaniu TIM", len(supplemental.get("stages", {}).get("missingInTimSearch", [])), "BRAK", "Nie utworzono bez wymaganych mapowań handlowych"],
        ["Nazwy", "Nazwy potwierdzone live", name_verified, "OK" if not name_mismatch else "CZĘŚCIOWO", f"niedopisane={name_mismatch}"],
        ["EPREL", "Pełne zestawy klasa + etykieta + karta produktu", eprel_verified, "OK", "Dokładny model i oficjalne dokumenty"],
        ["EPREL", "Pozostałe karty z blokadą braku karty produktu", max(0, len(missing_sheets) - eprel_verified), "BRAK", "Bez przypisywania podobnego modelu"],
        ["Chmura", "Aktywne produkty Mamezi", cloud.get("freshProducts", 0), "ODCZYT", f"dodane={len(cloud.get('added', []))}, usunięte={len(cloud.get('removed', []))}, zmienione={len(cloud.get('changed', []))}"],
        ["Dokumentacja", "Karty kompletne w audycie 28.08.2026", document_audit_summary.get("complete", 0), "SNAPSHOT", f"z {document_audit_summary.get('all', 0)} kart; raport Gemini zachowany jako materiał porównawczy"],
        ["Dokumentacja", "Brak karty PDF w audycie 28.08.2026", document_audit_summary.get("missingCard", 0), "DO UZUPEŁNIENIA", "Nie przypisywano dokumentu bez dokładnego dopasowania modelu"],
        ["Dokumentacja", "Brak CE w audycie 28.08.2026", document_audit_summary.get("missingCe", 0), "DO UZUPEŁNIENIA", "Aktywa CE są w PIMCORE; wymagają bezpiecznej mapy modeli"],
        ["Dokumentacja", "Brak zdjęcia w audycie 28.08.2026", document_audit_summary.get("missingImage", 0), "DO UZUPEŁNIENIA", "Chmura Mamezi może służyć do treści i zdjęć, ale nie do cen"],
        ["Dokumentacja", "Znane aktywa CE w PIMCORE", len(ce_asset_rows), "OK", "Mapowanie plików zachowane; relacje do kart wynikają z audytu dokumentacji"],
        ["Importy", "Szeroki proces 4765", "98% / IMPORTING", "NIE DOTYKANO", "Nie uruchomiono kolejnego szerokiego importu"],
        ["Importy", "Schemat 648: pilot 10 cen", "Oczekiwanie na akceptację administratora", "NIE ZAAKCEPTOWANO", "Stan potwierdzony odczytem 31.08.2026; nie wysłano do akceptacji"],
    ]

    catalog_rows: list[list[Any]] = []
    positive_rows: list[list[Any]] = []
    for product in products:
        ean = str(product.get("ean", ""))
        source = source_for(product)
        queue_stage, card = mapped.get(ean, ("brak_dopasowania", {}))
        object_id = int(card.get("pimcoreId") or 0)
        verification = description_results.get(object_id, {})
        outcome = verification.get("status") or ("oczekuje_na_weryfikację" if object_id else "brak_karty_lub_konflikt")
        expected_price = number(source.get("price"))
        live_price = number(verification.get("livePrice")) if verification.get("livePrice") is not None and not isinstance(verification.get("livePrice"), (dict, list, tuple, set)) else None
        price_outcome = price_status(source.get("price"), verification.get("livePrice"))
        row = [
            ean, product.get("manufacturerCode", ""), product.get("tradeIndex", ""), product.get("producer", ""),
            product.get("name", ""), product.get("category", ""), source.get("price", ""), round(expected_price / 1.23, 2), live_price, verification.get("liveVatRate", ""), price_outcome,
            source.get("stock", ""), verification.get("liveStock", ""), source.get("attributes", {}).get("Jednostka", ""),
            verification.get("liveMeasureUnit", ""), product.get("status", ""), queue_stage,
            object_id or "", card.get("timIndex", ""), card.get("liveStock", ""), outcome,
            "TAK" if product.get("descriptionHtml") else "NIE", product.get("productUrl", ""),
        ]
        catalog_rows.append(row)
        if number(source.get("stock")) > 0:
            positive_rows.append(row + ["opis/dopasowanie do wyjaśnienia" if not object_id else "opis uzupełniony lub zweryfikowany"])

    description_rows = []
    for object_id, row in sorted(description_results.items(), key=lambda item: (item[1].get("status") != "verified", item[0])):
        description_rows.append([
            object_id, row.get("timIndex", ""), row.get("ean", ""), row.get("manufacturerCode", ""), row.get("name", ""),
            row.get("liveState", ""), row.get("liveStock", ""), row.get("status", ""), row.get("descriptionLength", ""),
            row.get("livePrice", ""), row.get("liveVatRate", ""), row.get("liveMeasureUnit", ""), row.get("productAvailableForSale", ""),
            row.get("reason", ""), Path(row.get("verificationFile", "")).name,
        ])

    supplemental_eans = {
        str(row.get("ean", ""))
        for rows in supplemental.get("stages", {}).values()
        for row in rows if row.get("ean")
    }
    missing_rows = []
    for prefix, queue in (("BAZA", base), ("UZUPEŁNIENIE", supplemental)):
        eligible = {"activePositiveNeedsUpdate", "activeZeroNeedsUpdate", "bufferNewNeedsUpdate", "bufferApprovalNeedsUpdate"}
        for stage, rows in queue.get("stages", {}).items():
            if stage in eligible:
                continue
            for row in rows:
                if prefix == "BAZA" and stage == "missingInMainCatalog" and str(row.get("ean", "")) in supplemental_eans:
                    continue
                missing_rows.append([
                    prefix, stage, row.get("pimcoreId", ""), row.get("ean", ""), row.get("manufacturerCode", ""),
                    row.get("name", ""), row.get("producer", ""), row.get("status", ""), row.get("liveState", ""),
                    row.get("reason", ""), row.get("productUrl", ""),
                ])
    fixed_eprel_ids = {int(row.get("pimcoreId") or 0) for row in eprel_queue.get("items", [])}
    for row in missing_sheets:
        if int(row.get("objectId") or 0) in fixed_eprel_ids:
            continue
        missing_rows.append([
            "WALIDACJA", "brak_karty_informacyjnej_produktu", row.get("objectId", ""), row.get("ean", ""),
            row.get("manufacturerCode", ""), row.get("name", ""), "", "", "active", "Brak dokładnego oficjalnego dokumentu", row.get("productUrl", ""),
        ])
    for row in description_results.values():
        if row.get("status") == "verified":
            continue
        missing_rows.append([
            "WERYFIKACJA LIVE", "opis_niezgodny_lub_nieodczytany", row.get("objectId", ""), row.get("ean", ""),
            row.get("manufacturerCode", ""), row.get("name", ""), "", "", row.get("liveState", ""),
            row.get("reason", "Brak zgodności opisu po bezpiecznym odczycie"), "",
        ])

    supplemental_by_id = {
        int(row.get("pimcoreId") or 0): (stage, row)
        for stage, rows in supplemental.get("stages", {}).items()
        for row in rows if row.get("pimcoreId")
    }
    buffer_rows = []
    for node in buffer_nodes:
        stage, product = supplemental_by_id.get(int(node.get("id") or 0), ("nierozpoznana_po_EAN", {}))
        buffer_rows.append([
            "Karta bufora", node.get("id", ""), product.get("ean", ""), product.get("manufacturerCode", ""),
            node.get("text", ""), product.get("producer", ""), product.get("liveState", ""), stage,
            description_results.get(int(node.get("id") or 0), {}).get("status", ""), node.get("locked", False), node.get("path", ""),
        ])
    for product in supplemental.get("stages", {}).get("missingInTimSearch", []):
        buffer_rows.append([
            "Brak karty w TIM", "", product.get("ean", ""), product.get("manufacturerCode", ""), product.get("name", ""),
            product.get("producer", ""), "", "missingInTimSearch", "nie_utworzono", "", product.get("productUrl", ""),
        ])

    eprel_status = {int(row.get("objectId") or 0): row.get("status", "") for row in eprel_live.get("results", [])}
    eprel_rows = []
    for row in eprel_queue.get("items", []):
        eprel_rows.append([
            row.get("pimcoreId", ""), row.get("timIndex", ""), row.get("ean", ""), row.get("manufacturerCode", ""),
            row.get("eprelId", ""), row.get("energyClass", ""), eprel_status.get(int(row.get("pimcoreId") or 0), ""),
            row.get("labelFile", ""), row.get("productInformationSheet", ""),
        ])

    name_status = {
        int(row.get("objectId") or 0): (row.get("status", ""), row.get("currentName", ""), row.get("reason", ""))
        for row in name_positive.get("results", []) + name_zero.get("results", [])
    }
    name_rows = []
    for stage, rows in final_name_queue.get("stages", {}).items():
        for row in rows:
            status, current_name, reason = name_status.get(int(row.get("pimcoreId") or 0), ("", "", ""))
            name_rows.append([
                stage, row.get("pimcoreId", ""), row.get("timIndex", ""), row.get("ean", ""), row.get("manufacturerCode", ""),
                " + ".join(row.get("cleanupKinds", [])), row.get("beforeName", ""), row.get("afterName", ""), current_name,
                row.get("stock", ""), status, reason,
            ])

    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "PODSUMOWANIE", ["Obszar", "Miara", "Wartość", "Status", "Uwagi"], summary_rows)
    catalog_headers = ["EAN", "Indeks handlowy", "Indeks katalogowy Prescot (raport)", "Producent", "Nazwa źródłowa", "Kategoria", "Cena z prescot.xml", "Cena z prescot.xml / 1,23 (kontrola)", "Cena live TIM", "VAT live", "Zgodność ceny", "Stan z prescot.xml", "Stan live TIM", "Jednostka z prescot.xml", "Jednostka live TIM", "Jakość treści", "Etap dopasowania", "PIMCORE ID", "Indeks TIM", "Stan live z kolejki", "Weryfikacja opisu", "Opis wygenerowany", "URL produktu"]
    add_sheet(workbook, "KATALOG_TIM", catalog_headers, catalog_rows)
    add_sheet(workbook, "AKTYWNE_STAN_DODATNI", catalog_headers + ["Dalsza czynność"], positive_rows)
    add_sheet(workbook, "OPISY_POPRAWIONE", ["PIMCORE ID", "Indeks TIM", "EAN", "Indeks handlowy", "Nazwa", "Stan karty", "Stan live", "Wynik", "Długość opisu", "Cena live TIM", "VAT live", "Jednostka live TIM", "Dostępny do sprzedaży", "Powód", "Raport weryfikacji"], description_rows)
    add_sheet(workbook, "BRAKI_KONFLIKTY", ["Źródło", "Typ braku/konfliktu", "PIMCORE ID", "EAN", "Indeks handlowy", "Nazwa", "Producent", "Jakość", "Stan karty", "Powód", "URL"], missing_rows)
    add_sheet(workbook, "BUFOR_NOWOSCI", ["Rodzaj", "PIMCORE ID", "EAN", "Indeks handlowy", "Nazwa", "Producent", "Stan karty", "Etap", "Weryfikacja opisu", "Blokada", "Ścieżka/URL"], buffer_rows)
    add_sheet(workbook, "EPREL_DOKUMENTY", ["PIMCORE ID", "Indeks TIM", "EAN", "Indeks handlowy", "EPREL ID", "Klasa", "Wynik live", "Etykieta źródłowa", "Karta produktu źródłowa"], eprel_rows)
    if document_audit_rows:
        add_sheet(
            workbook,
            "DOKUMENTY_SNAPSHOT_28_08",
            ["PIMCORE ID", "Indeks TIM", "Indeks producenta", "Nazwa", "Producent", "Liczba braków", "Lista braków", "Brak karty PDF", "Brak CE", "Brak zdjęcia", "Brak ETIM", "Brak ceny", "Brak nazwy", "Cena netto TIM (snapshot)", "Dostępny do sprzedaży", "Opublikowany"],
            document_audit_rows,
        )
    if ce_asset_rows:
        add_sheet(workbook, "AKTYWA_CE", ["Plik", "Asset ID", "Ścieżka PIMCORE", "Typ", "Status źródła"], ce_asset_rows)
    add_sheet(workbook, "POPRAWKI_NAZW", ["Etap", "PIMCORE ID", "Indeks TIM", "EAN", "Indeks handlowy", "Rodzaj poprawki", "Nazwa przed", "Nazwa oczekiwana", "Nazwa live", "Stan", "Wynik", "Powód"], name_rows)
    add_sheet(workbook, "METADANE", ["Pole", "Wartość"], [
        ["Wygenerowano", manifest.get("meta", {}).get("generatedAt", "")],
        ["Źródło ceny/stanu/jednostki", "https://prescot.wapromag.pl/prescot.xml"],
        ["Źródło treści/zdjęć/parametrów", "cloud.appstore.mamezi.pl (prywatny adres feedu niewpisany do raportu)"],
        ["Zasada cenowa", "Cena z Mamezi nie jest używana jako cena TIM"],
        ["Raporty weryfikacji opisów", ", ".join(Path(path).name for path in verification_files)],
        ["Liczba raportów weryfikacji", len(verification_files)],
        ["Statusy opisów", json.dumps(Counter(row.get("status") for row in description_results.values()), ensure_ascii=False)],
    ])

    output = Path(args.output).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    checked = load_workbook(output, read_only=True, data_only=True)
    expected = ["PODSUMOWANIE", "KATALOG_TIM", "AKTYWNE_STAN_DODATNI", "OPISY_POPRAWIONE", "BRAKI_KONFLIKTY", "BUFOR_NOWOSCI", "EPREL_DOKUMENTY"]
    if document_audit_rows:
        expected.append("DOKUMENTY_SNAPSHOT_28_08")
    if ce_asset_rows:
        expected.append("AKTYWA_CE")
    expected.extend(["POPRAWKI_NAZW", "METADANE"])
    if checked.sheetnames != expected:
        raise RuntimeError(f"Nieprawidłowe arkusze: {checked.sheetnames}")
    checked.close()
    print(json.dumps({"output": str(output), "sheets": expected, "verifiedDescriptions": verified, "rows": {"catalog": len(catalog_rows), "positive": len(positive_rows), "buffer": len(buffer_rows), "missing": len(missing_rows)}}, ensure_ascii=False))


if __name__ == "__main__":
    main()
