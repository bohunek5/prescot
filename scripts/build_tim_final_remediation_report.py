#!/usr/bin/env python3
"""Build the final local TIM remediation workbook from JSON/CSV evidence.

The script performs no network requests and does not communicate with TIM.  It
only reads existing files from ``exports/tim/remediation`` and creates one new
XLSX selected with ``--output``.  Existing output files are never overwritten.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "exports" / "tim" / "remediation"

SHEETS = (
    "PODSUMOWANIE",
    "ZROBIONE",
    "BRAKI",
    "DOPASOWANIA_80_99",
    "EPREL_POCHODNE",
    "WYCOFANE",
    "BLEDY_TIM",
    "DO_RECZNEJ_DECYZJI",
)

KLUS_LIVE_PATTERNS = (
    "klus-active-next-documents-exact100-*-live-2026-09-02.json",
    "klus-active-documents-retry-*-live-2026-09-02.json",
)
NONCORE_LIVE_PATTERN = "noncore-withdrawal-*-live-2026-09-02.json"
DELUX_POSTVERIFY_PRIORITY = (
    "prescot-tape-natural-delux52-v8-postverify-2026-09-02.json",
    "prescot-tape-natural-delux52-postverify-2026-09-02.json",
    "prescot-tape-natural-delux10-postverify-2026-09-02.json",
)

PATHS = {
    "klus_mapping": DATA / "klus-active-next-documents-readonly-2026-09-02.json",
    "klus_postverify": DATA / "klus-active-next-documents-exact100-final-postverify-2026-09-02.json",
    "klus_unverified": DATA / "klus-active-documents-355-unverified-2026-09-02.csv",
    "prescot_mapping": DATA / "prescot-active-positive-local-document-mapping-audit-2026-09-02.json",
    "prescot_live": DATA / "prescot-active-positive-exact100-documents-live-2026-09-02.json",
    "prescot_write_queue": DATA / "prescot-active-positive-exact100-documents-live-write-queue-2026-09-02.json",
    "http500": DATA / "prescot-active-positive-exact100-documents-http500-readonly-analysis-2026-09-02.json",
    "noncore_queue": DATA / "noncore-withdrawal-queue-2026-09-02.json",
}

HEADER_FILL = PatternFill("solid", fgColor="8B174D")
SUBTLE_FILL = PatternFill("solid", fgColor="F4EAF0")
ERROR_FILL = PatternFill("solid", fgColor="FCE8E6")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a local final TIM remediation workbook from existing evidence."
    )
    parser.add_argument(
        "--output",
        required=True,
        type=Path,
        help="Path of a new .xlsx file. Existing files are not overwritten.",
    )
    return parser.parse_args()


def load_json(path: Path, warnings: list[str]) -> Any:
    if not path.exists():
        warnings.append(f"Brak pliku: {path.relative_to(ROOT)}")
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        warnings.append(f"Nie można odczytać {path.relative_to(ROOT)}: {exc}")
        return {}


def load_csv(path: Path, warnings: list[str]) -> list[dict[str, str]]:
    if not path.exists():
        warnings.append(f"Brak pliku: {path.relative_to(ROOT)}")
        return []
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    except (OSError, csv.Error) as exc:
        warnings.append(f"Nie można odczytać {path.relative_to(ROOT)}: {exc}")
        return []


def discover(patterns: Iterable[str]) -> list[Path]:
    found: set[Path] = set()
    for pattern in patterns:
        found.update(DATA.glob(pattern))
    return sorted(found)


def source_name(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def first(item: dict[str, Any], *names: str, default: Any = "") -> Any:
    for name in names:
        value = item.get(name)
        if value is not None and value != "":
            return value
    return default


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TAK" if value else "NIE"
    if isinstance(value, (dict, list, tuple, set)):
        if isinstance(value, dict):
            parts = [f"{key}: {as_text(subvalue)}" for key, subvalue in value.items()]
        else:
            parts = [as_text(part) for part in value]
        return " | ".join(part for part in parts if part)
    return str(value).strip()


def excel_safe(value: Any) -> Any:
    if isinstance(value, Path):
        value = str(value)
    if isinstance(value, (dict, list, tuple, set)):
        value = as_text(value)
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def stable_id(item: dict[str, Any]) -> str:
    return str(first(item, "pimcoreId", "objectId", "id", default=""))


def documents_text(documents: Any) -> str:
    if not isinstance(documents, dict):
        return as_text(documents)
    entries: list[str] = []
    for field, value in documents.items():
        if isinstance(value, dict):
            url = first(value, "officialUrl", "localFile", "filename")
            entries.append(f"{field}: {url}")
        else:
            entries.append(f"{field}: {value}")
    return " | ".join(entries)


def item_identity(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "PIMCORE ID": stable_id(item),
        "EAN": first(item, "ean"),
        "MODEL / INDEKS HANDLOWY": first(
            item, "model", "modelHandlowy", "manufacturerCode"
        ),
        "NAZWA": first(item, "name", "timName"),
        "STAN": first(item, "stock", "liveStock", "xmlStock", "queueXmlStock"),
    }


def append_unique(
    target: list[dict[str, Any]],
    seen: set[tuple[str, ...]],
    row: dict[str, Any],
    key_fields: tuple[str, ...],
) -> None:
    key = tuple(as_text(row.get(field)) for field in key_fields)
    if key in seen:
        return
    seen.add(key)
    target.append(row)


def collect_live_reports(
    paths: Iterable[Path], warnings: list[str]
) -> list[tuple[Path, dict[str, Any]]]:
    reports: list[tuple[Path, dict[str, Any]]] = []
    for path in paths:
        data = load_json(path, warnings)
        if isinstance(data, dict):
            reports.append((path, data))
    return reports


def newest_delux_path() -> Path | None:
    for filename in DELUX_POSTVERIFY_PRIORITY:
        path = DATA / filename
        if path.exists():
            return path
    return None


def collect_eprel_derived(warnings: list[str]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    rejected_rows: list[dict[str, Any]] = []
    seen: set[tuple[str, ...]] = set()
    rejected_seen: set[tuple[str, ...]] = set()
    paths = discover(("*eprel*queue*2026-09-01.json", "*eprel*queue*2026-09-02.json"))

    for path in paths:
        data = load_json(path, warnings)
        if not isinstance(data, dict):
            continue
        for item in data.get("items", []):
            if not isinstance(item, dict):
                continue
            match_type = as_text(first(item, "matchType", "eprelMatchType")).lower()
            confidence = first(item, "confidence", "eprelConfidence")
            reason = as_text(first(item, "matchReason", "reason", "eprelReason"))
            family_signal = any(
                token in f"{path.name} {match_type} {reason}".lower()
                for token in ("length", "family", "derived", "pochod", "variant", "suffix")
            )
            try:
                below_exact = float(confidence) < 100
            except (TypeError, ValueError):
                below_exact = False
            if not family_signal and not below_exact:
                continue
            row = {
                **item_identity(item),
                "MODEL EPREL": first(item, "eprelModel"),
                "NR EPREL": first(item, "eprelId", "registrationNumber"),
                "KLASA": first(item, "energyClass"),
                "TYP DOPASOWANIA": first(item, "matchType", "eprelMatchType"),
                "PEWNOŚĆ %": confidence,
                "UZASADNIENIE": reason,
                "EPREL URL": first(item, "eprelPublicUrl"),
                "ETYKIETA / PLIK": first(item, "labelFile", "labelAddress"),
                "KARTA PRODUKTU": first(item, "productInformationSheet", "eprelFicheUrl"),
                "QR POTWIERDZONY": first(item, "qrVerified"),
                "KARTA POTWIERDZONA": first(item, "ficheVerified"),
                "ŹRÓDŁO": source_name(path),
            }
            append_unique(
                rows,
                seen,
                row,
                ("PIMCORE ID", "MODEL / INDEKS HANDLOWY", "NR EPREL", "TYP DOPASOWANIA"),
            )

        for item in data.get("rejected", []):
            if not isinstance(item, dict):
                continue
            row = {
                **item_identity(item),
                "MODEL EPREL": first(item, "eprelModel"),
                "NR EPREL": first(item, "eprelId", "registrationNumber"),
                "TYP DOPASOWANIA": first(item, "matchType", "eprelMatchType"),
                "PEWNOŚĆ %": first(item, "confidence", "eprelConfidence"),
                "POWÓD ODRZUCENIA": first(item, "reason"),
                "SZCZEGÓŁ": first(item, "detail"),
                "ŹRÓDŁO": source_name(path),
            }
            append_unique(
                rejected_rows,
                rejected_seen,
                row,
                ("PIMCORE ID", "MODEL / INDEKS HANDLOWY", "NR EPREL", "POWÓD ODRZUCENIA"),
            )
    return rows, rejected_rows


def collect_rows() -> tuple[dict[str, list[dict[str, Any]]], list[str], dict[str, Any]]:
    warnings: list[str] = []
    rows: dict[str, list[dict[str, Any]]] = {name: [] for name in SHEETS}

    klus_mapping = load_json(PATHS["klus_mapping"], warnings)
    klus_postverify = load_json(PATHS["klus_postverify"], warnings)
    prescot_mapping = load_json(PATHS["prescot_mapping"], warnings)
    prescot_live = load_json(PATHS["prescot_live"], warnings)
    prescot_write_queue = load_json(PATHS["prescot_write_queue"], warnings)
    http500 = load_json(PATHS["http500"], warnings)
    noncore_queue = load_json(PATHS["noncore_queue"], warnings)
    klus_unverified = load_csv(PATHS["klus_unverified"], warnings)

    klus_live_paths = discover(KLUS_LIVE_PATTERNS)
    klus_live = collect_live_reports(klus_live_paths, warnings)
    noncore_live_paths = discover((NONCORE_LIVE_PATTERN,))
    noncore_live = collect_live_reports(noncore_live_paths, warnings)

    klus_exact_by_id = {
        stable_id(item): item
        for item in klus_mapping.get("exact100", [])
        if isinstance(item, dict)
    }
    prescot_exact_by_id = {
        stable_id(item): item
        for item in prescot_mapping.get("exact100", [])
        if isinstance(item, dict)
    }

    saved_klus_ids: set[str] = set()
    error_seen: set[tuple[str, ...]] = set()
    done_seen: set[tuple[str, ...]] = set()

    for path, report in klus_live:
        for result in report.get("results", []):
            if not isinstance(result, dict):
                continue
            object_id = stable_id(result)
            mapped = klus_exact_by_id.get(object_id, {})
            status = as_text(result.get("status")).lower()
            if status in {"saved", "already_current", "verified"}:
                saved_klus_ids.add(object_id)
                row = {
                    "DATA RAPORTU": report.get("generatedAt", ""),
                    "OBSZAR": "KLUŚ — oficjalne dokumenty exact100",
                    "PRODUCENT": "KLUŚ",
                    **item_identity({**mapped, **result}),
                    "OPERACJA": "dołączenie dokumentów",
                    "POLA": as_text(first(result, "dirtyFields", default=mapped.get("targetFields", []))),
                    "DOKUMENTY": documents_text(first(result, "documents", default=mapped.get("documents", {}))),
                    "WYNIK": result.get("status", ""),
                    "HTTP": first(result, "saveResponseStatus", "httpStatus"),
                    "ŹRÓDŁO": source_name(path),
                }
                append_unique(
                    rows["ZROBIONE"], done_seen, row, ("OBSZAR", "PIMCORE ID", "POLA", "WYNIK")
                )
            elif status:
                row = {
                    "DATA RAPORTU": report.get("generatedAt", ""),
                    "OBSZAR": "KLUŚ — zapis dokumentów",
                    "PRODUCENT": "KLUŚ",
                    **item_identity({**mapped, **result}),
                    "STATUS": result.get("status", ""),
                    "HTTP": first(result, "saveResponseStatus", "httpStatus"),
                    "KOD / POWÓD": first(result, "reason", default=report.get("fatalError", "")),
                    "ZALECENIE": "Nie ponawiać automatycznie bez rozpoznania walidacji lub duplikatu.",
                    "ŹRÓDŁO": source_name(path),
                }
                append_unique(
                    rows["BLEDY_TIM"], error_seen, row, ("PIMCORE ID", "HTTP", "KOD / POWÓD")
                )

    # Exact KLUŚ mapping not evidenced as saved remains work to do, not "done".
    missing_seen: set[tuple[str, ...]] = set()
    for item in klus_mapping.get("exact100", []):
        if not isinstance(item, dict) or stable_id(item) in saved_klus_ids:
            continue
        row = {
            "PRIORYTET": "1 — exact100",
            "PRODUCENT": "KLUŚ",
            **item_identity(item),
            "BRAKUJĄCE / DOCELOWE POLA": as_text(item.get("targetFields")),
            "DOKUMENTY GOTOWE": documents_text(item.get("documents")),
            "PEWNOŚĆ %": item.get("confidence", 100),
            "BLOKER / UWAGA": as_text(item.get("conflicts")),
            "ŹRÓDŁO": source_name(PATHS["klus_mapping"]),
        }
        append_unique(rows["BRAKI"], missing_seen, row, ("PRODUCENT", "PIMCORE ID", "BRAKUJĄCE / DOCELOWE POLA"))

    # Prescot exact live results: saved items go to done, failures to TIM errors.
    saved_prescot_ids: set[str] = set()
    if isinstance(prescot_live, dict):
        for result in prescot_live.get("results", []):
            if not isinstance(result, dict):
                continue
            object_id = stable_id(result)
            mapped = prescot_exact_by_id.get(object_id, {})
            status = as_text(result.get("status")).lower()
            if status in {"saved", "already_current", "verified"}:
                saved_prescot_ids.add(object_id)
                row = {
                    "DATA RAPORTU": prescot_live.get("generatedAt", ""),
                    "OBSZAR": "Prescot — dokumenty exact100",
                    "PRODUCENT": "PRESCOT",
                    **item_identity({**mapped, **result}),
                    "OPERACJA": "dołączenie dokumentów",
                    "POLA": as_text(first(result, "dirtyFields", default=mapped.get("field", ""))),
                    "DOKUMENTY": documents_text(first(result, "documents", default=mapped.get("file", ""))),
                    "WYNIK": result.get("status", ""),
                    "HTTP": first(result, "saveResponseStatus", "httpStatus"),
                    "ŹRÓDŁO": source_name(PATHS["prescot_live"]),
                }
                append_unique(rows["ZROBIONE"], done_seen, row, ("OBSZAR", "PIMCORE ID", "POLA", "WYNIK"))
            else:
                row = {
                    "DATA RAPORTU": prescot_live.get("generatedAt", ""),
                    "OBSZAR": "Prescot — zapis dokumentów",
                    "PRODUCENT": "PRESCOT",
                    **item_identity({**mapped, **result}),
                    "STATUS": result.get("status", ""),
                    "HTTP": first(result, "saveResponseStatus", "httpStatus"),
                    "KOD / POWÓD": first(result, "reason", default=prescot_live.get("fatalError", "")),
                    "ZALECENIE": "Najpierw rozwiązać wymagania energii B24; nie ponawiać zapisu samej karty.",
                    "ŹRÓDŁO": source_name(PATHS["prescot_live"]),
                }
                append_unique(rows["BLEDY_TIM"], error_seen, row, ("PIMCORE ID", "HTTP", "KOD / POWÓD"))

    queued_prescot_ids: set[str] = set()
    for container in (prescot_write_queue, prescot_live):
        if not isinstance(container, dict):
            continue
        for item in first(container, "items", "queue", "records", default=[]):
            if isinstance(item, dict):
                queued_prescot_ids.add(stable_id(item))

    for object_id in sorted(queued_prescot_ids):
        if not object_id or object_id in saved_prescot_ids:
            continue
        matches = [
            item
            for item in prescot_mapping.get("exact100", [])
            if isinstance(item, dict) and stable_id(item) == object_id
        ]
        base = matches[0] if matches else {"pimcoreId": object_id}
        fields = sorted({as_text(item.get("field")) for item in matches if item.get("field")})
        files = sorted({as_text(item.get("file")) for item in matches if item.get("file")})
        row = {
            "PRIORYTET": "1 — exact100",
            "PRODUCENT": "PRESCOT",
            **item_identity(base),
            "BRAKUJĄCE / DOCELOWE POLA": ", ".join(fields),
            "DOKUMENTY GOTOWE": " | ".join(files),
            "PEWNOŚĆ %": 100,
            "BLOKER / UWAGA": "Nie ma potwierdzonego skutecznego zapisu live.",
            "ŹRÓDŁO": source_name(PATHS["prescot_write_queue"]),
        }
        append_unique(rows["BRAKI"], missing_seen, row, ("PRODUCENT", "PIMCORE ID", "BRAKUJĄCE / DOCELOWE POLA"))

    # Prescot products with no mapping are explicit gaps.
    for item in prescot_mapping.get("products", []):
        if not isinstance(item, dict):
            continue
        mapping_class = as_text(item.get("mappingClass")).lower()
        if mapping_class not in {"", "unmatched", "none", "no_match"} and item.get("matches"):
            continue
        missing_fields = []
        if not item.get("currentDataSheetCount"):
            missing_fields.append("dataSheet")
        if not item.get("currentCertificationsCount"):
            missing_fields.append("certifications")
        if not missing_fields:
            continue
        row = {
            "PRIORYTET": "3 — brak bezpiecznego mapowania",
            "PRODUCENT": "PRESCOT",
            **item_identity(item),
            "BRAKUJĄCE / DOCELOWE POLA": ", ".join(missing_fields),
            "DOKUMENTY GOTOWE": "",
            "PEWNOŚĆ %": first(item, "bestConfidence"),
            "BLOKER / UWAGA": as_text(item.get("conflicts")) or "Brak dopasowania lokalnego dokumentu.",
            "ŹRÓDŁO": source_name(PATHS["prescot_mapping"]),
        }
        append_unique(rows["BRAKI"], missing_seen, row, ("PRODUCENT", "PIMCORE ID", "BRAKUJĄCE / DOCELOWE POLA"))

    # KLUŚ records rejected for lack of official evidence are also unresolved gaps.
    for item in klus_mapping.get("rejected", []):
        if not isinstance(item, dict):
            continue
        row = {
            "PRIORYTET": "3 — brak oficjalnego dokumentu",
            "PRODUCENT": "KLUŚ",
            **item_identity(item),
            "BRAKUJĄCE / DOCELOWE POLA": "dataSheet / certifications — do ustalenia",
            "DOKUMENTY GOTOWE": "",
            "PEWNOŚĆ %": "",
            "BLOKER / UWAGA": first(item, "reason"),
            "ŹRÓDŁO": source_name(PATHS["klus_mapping"]),
        }
        append_unique(rows["BRAKI"], missing_seen, row, ("PRODUCENT", "PIMCORE ID", "BRAKUJĄCE / DOCELOWE POLA"))

    # 80–99 mappings remain explicitly separated from exact matches.
    fuzzy_seen: set[tuple[str, ...]] = set()
    for producer, mapping, model_key, source in (
        ("KLUŚ", klus_mapping, "model", PATHS["klus_mapping"]),
        ("PRESCOT", prescot_mapping, "modelHandlowy", PATHS["prescot_mapping"]),
    ):
        for item in mapping.get("fuzzy80to99", []):
            if not isinstance(item, dict):
                continue
            row = {
                "PRODUCENT": producer,
                **item_identity(item),
                "POLE": first(item, "field", default=as_text(item.get("targetFields"))),
                "PEWNOŚĆ %": item.get("confidence", ""),
                "TYP DOPASOWANIA": first(item, "matchType"),
                "UZASADNIENIE": first(item, "reason", "confidenceReason"),
                "DOKUMENT / URL": first(
                    item,
                    "file",
                    default=documents_text(item.get("documents")) or item.get("officialProductUrl", ""),
                ),
                "KONFLIKTY": as_text(first(item, "conflicts", "productConflicts", default=[])),
                "DECYZJA": "Wymaga potwierdzenia ręcznego przed zapisem.",
                "ŹRÓDŁO": source_name(source),
            }
            append_unique(
                rows["DOPASOWANIA_80_99"],
                fuzzy_seen,
                row,
                ("PRODUCENT", "PIMCORE ID", "POLE", "DOKUMENT / URL"),
            )

    eprel_rows, eprel_rejected = collect_eprel_derived(warnings)
    rows["EPREL_POCHODNE"].extend(eprel_rows)

    # Withdrawal: one row per queue item, enriched with the newest live attempt.
    latest_withdrawal: dict[str, tuple[str, Path, dict[str, Any]]] = {}
    for path, report in noncore_live:
        stamp = as_text(report.get("generatedAt"))
        for result in report.get("results", []):
            if not isinstance(result, dict):
                continue
            object_id = stable_id(result)
            previous = latest_withdrawal.get(object_id)
            if previous is None or stamp >= previous[0]:
                latest_withdrawal[object_id] = (stamp, path, result)

    withdrawal_seen: set[tuple[str, ...]] = set()
    queue_items = noncore_queue.get("items", []) if isinstance(noncore_queue, dict) else []
    queued_ids = {stable_id(item) for item in queue_items if isinstance(item, dict)}
    all_withdrawal_items = [item for item in queue_items if isinstance(item, dict)]
    all_withdrawal_items.extend(
        result
        for object_id, (_, _, result) in latest_withdrawal.items()
        if object_id not in queued_ids
    )
    for item in all_withdrawal_items:
        object_id = stable_id(item)
        live_tuple = latest_withdrawal.get(object_id)
        live = live_tuple[2] if live_tuple else {}
        path = live_tuple[1] if live_tuple else PATHS["noncore_queue"]
        combined = {**item, **live}
        row = {
            "PRODUCENT": first(combined, "brand"),
            **item_identity(combined),
            "REGUŁA": first(combined, "rule"),
            "STATUS PRZED": first(combined, "beforeState", "beforeStatus"),
            "WYNIK / STATUS": first(combined, "status", default="queued_not_attempted"),
            "POWÓD / WALIDACJA": first(combined, "reason"),
            "KARTA": first(combined, "dataSheetCount"),
            "CE": first(combined, "certificationsCount"),
            "DANE PRODUKTU NIENARUSZONE": first(combined, "protectedProductDataUnchanged"),
            "ŹRÓDŁO": source_name(path),
        }
        append_unique(rows["WYCOFANE"], withdrawal_seen, row, ("PIMCORE ID",))
        status = as_text(row["WYNIK / STATUS"]).lower()
        if any(token in status for token in ("failed", "rejected")):
            error = {
                "DATA RAPORTU": live_tuple[0] if live_tuple else "",
                "OBSZAR": "Wycofanie oferty non-core",
                "PRODUCENT": row["PRODUCENT"],
                **item_identity(combined),
                "STATUS": row["WYNIK / STATUS"],
                "HTTP": first(live, "transitionResponse", default={}).get("status", "")
                if isinstance(live.get("transitionResponse"), dict)
                else "",
                "KOD / POWÓD": row["POWÓD / WALIDACJA"],
                "ZALECENIE": "Decyzja ręczna; nie obchodź walidacji wymaganych pól.",
                "ŹRÓDŁO": source_name(path),
            }
            append_unique(rows["BLEDY_TIM"], error_seen, error, ("PIMCORE ID", "STATUS", "KOD / POWÓD"))

    # Latest Delux post-verification is treated as evidence, not as a new write.
    delux_path = newest_delux_path()
    delux = load_json(delux_path, warnings) if delux_path else {}
    if delux_path is None:
        warnings.append("Brak raportu Delux postverify z 2026-09-02.")
    if isinstance(delux, dict):
        for result in delux.get("results", []):
            if not isinstance(result, dict):
                continue
            status = as_text(result.get("status")).lower()
            if status in {"already_current", "saved", "saved_with_validation", "verified"}:
                row = {
                    "DATA RAPORTU": delux.get("generatedAt", ""),
                    "OBSZAR": "Prescot — opisy Delux postverify",
                    "PRODUCENT": "PRESCOT",
                    **item_identity(result),
                    "OPERACJA": "weryfikacja opisu",
                    "POLA": "description",
                    "DOKUMENTY": "",
                    "WYNIK": result.get("status", ""),
                    "HTTP": "",
                    "ŹRÓDŁO": source_name(delux_path),
                }
                append_unique(rows["ZROBIONE"], done_seen, row, ("OBSZAR", "PIMCORE ID", "WYNIK"))
            elif status:
                row = {
                    "PRIORYTET": "2 — opis Delux",
                    "PRODUCENT": "PRESCOT",
                    **item_identity(result),
                    "BRAKUJĄCE / DOCELOWE POLA": "description",
                    "DOKUMENTY GOTOWE": "",
                    "PEWNOŚĆ %": "",
                    "BLOKER / UWAGA": first(result, "reason", default=result.get("status", "")),
                    "ŹRÓDŁO": source_name(delux_path),
                }
                append_unique(rows["BRAKI"], missing_seen, row, ("PRODUCENT", "PIMCORE ID", "BRAKUJĄCE / DOCELOWE POLA"))

    # Read-only HTTP 500 diagnosis supplies the actionable error explanation.
    if isinstance(http500, dict):
        failed_attempt = http500.get("failedAttempt", {})
        for product in http500.get("products", []):
            if not isinstance(product, dict):
                continue
            is_failed_object = stable_id(product) == str(failed_attempt.get("objectId", ""))
            row = {
                "DATA RAPORTU": http500.get("generatedAt", ""),
                "OBSZAR": "Prescot — diagnoza HTTP 500 / walidacja B24",
                "PRODUCENT": "PRESCOT",
                **item_identity(product),
                "STATUS": "HTTP 500 przy próbie" if is_failed_object else "nie próbowano po zatrzymaniu kolejki",
                "HTTP": failed_attempt.get("saveHttpStatus", "") if is_failed_object else product.get("priorValidationHttpStatus", ""),
                "KOD / POWÓD": "Brak wymaganych pól: " + as_text(product.get("missingRequiredForCurrentB24Validation")),
                "ZALECENIE": as_text(http500.get("safeRecommendation")),
                "ŹRÓDŁO": source_name(PATHS["http500"]),
            }
            append_unique(rows["BLEDY_TIM"], error_seen, row, ("PIMCORE ID", "STATUS", "KOD / POWÓD"))

    # Three unresolved records from the original KLUŚ 355 batch.
    # Earlier batch failures that were later saved and verified are historical,
    # not current TIM blockers.  Keep only unresolved KLUŚ errors.
    rows["BLEDY_TIM"] = [
        row
        for row in rows["BLEDY_TIM"]
        if not (
            row.get("PRODUCENT") == "KLUŚ"
            and as_text(row.get("PIMCORE ID")) in saved_klus_ids
        )
    ]

    # Three unresolved records from the original KLUŚ 355 batch.
    manual_seen: set[tuple[str, ...]] = set()
    for item in klus_unverified:
        row = {
            "PRIORYTET": "1 — rozliczenie KLUŚ 355",
            "OBSZAR": "KLUŚ — nieweryfikowane po 355",
            "PRODUCENT": "KLUŚ",
            **item_identity(item),
            "PEWNOŚĆ %": "",
            "PROBLEM / KONFLIKT": item.get("cause", ""),
            "BEZPIECZNY NASTĘPNY KROK": item.get("safeNextStep", ""),
            "DOKUMENT / URL": " | ".join(
                filter(None, (item.get("officialProductUrl", ""), item.get("localFiles", "")))
            ),
            "ŹRÓDŁO": source_name(PATHS["klus_unverified"]),
        }
        append_unique(rows["DO_RECZNEJ_DECYZJI"], manual_seen, row, ("OBSZAR", "PIMCORE ID"))

    # Every fuzzy mapping is also a manual-decision item.
    for item in rows["DOPASOWANIA_80_99"]:
        row = {
            "PRIORYTET": "2 — potwierdzenie 80–99%",
            "OBSZAR": "Dopasowanie dokumentu",
            "PRODUCENT": item.get("PRODUCENT", ""),
            **{key: item.get(key, "") for key in item_identity({}).keys()},
            "PEWNOŚĆ %": item.get("PEWNOŚĆ %", ""),
            "PROBLEM / KONFLIKT": item.get("KONFLIKTY", "") or item.get("UZASADNIENIE", ""),
            "BEZPIECZNY NASTĘPNY KROK": "Potwierdzić zakres modelu/rodziny w oficjalnym dokumencie przed zapisem.",
            "DOKUMENT / URL": item.get("DOKUMENT / URL", ""),
            "ŹRÓDŁO": item.get("ŹRÓDŁO", ""),
        }
        for key in ("PIMCORE ID", "EAN", "MODEL / INDEKS HANDLOWY", "NAZWA", "STAN"):
            row[key] = item.get(key, "")
        append_unique(rows["DO_RECZNEJ_DECYZJI"], manual_seen, row, ("OBSZAR", "PRODUCENT", "PIMCORE ID", "DOKUMENT / URL"))

    for item in eprel_rejected:
        row = {
            "PRIORYTET": "1 — EPREL odrzucone",
            "OBSZAR": "EPREL — odrzucone lub niespójne źródło",
            "PRODUCENT": "PRESCOT",
            **{key: item.get(key, "") for key in ("PIMCORE ID", "EAN", "MODEL / INDEKS HANDLOWY", "NAZWA", "STAN")},
            "PEWNOŚĆ %": item.get("PEWNOŚĆ %", ""),
            "PROBLEM / KONFLIKT": " | ".join(
                filter(None, (as_text(item.get("POWÓD ODRZUCENIA")), as_text(item.get("SZCZEGÓŁ"))))
            ),
            "BEZPIECZNY NASTĘPNY KROK": "Nie przypinać; zweryfikować dokładny model i oficjalne pliki EPREL.",
            "DOKUMENT / URL": "",
            "ŹRÓDŁO": item.get("ŹRÓDŁO", ""),
        }
        append_unique(rows["DO_RECZNEJ_DECYZJI"], manual_seen, row, ("OBSZAR", "PIMCORE ID", "PROBLEM / KONFLIKT"))

    for item in rows["BLEDY_TIM"]:
        row = {
            "PRIORYTET": "1 — błąd TIM / walidacja",
            "OBSZAR": item.get("OBSZAR", ""),
            "PRODUCENT": item.get("PRODUCENT", ""),
            **{key: item.get(key, "") for key in ("PIMCORE ID", "EAN", "MODEL / INDEKS HANDLOWY", "NAZWA", "STAN")},
            "PEWNOŚĆ %": "",
            "PROBLEM / KONFLIKT": " | ".join(
                filter(None, (as_text(item.get("STATUS")), as_text(item.get("KOD / POWÓD"))))
            ),
            "BEZPIECZNY NASTĘPNY KROK": item.get("ZALECENIE", ""),
            "DOKUMENT / URL": "",
            "ŹRÓDŁO": item.get("ŹRÓDŁO", ""),
        }
        append_unique(rows["DO_RECZNEJ_DECYZJI"], manual_seen, row, ("OBSZAR", "PIMCORE ID", "PROBLEM / KONFLIKT"))

    # Summary is calculated from evidence rows after deduplication.
    klus_saved = sum(
        1 for row in rows["ZROBIONE"] if row.get("OBSZAR") == "KLUŚ — oficjalne dokumenty exact100"
    )
    prescot_saved = sum(
        1 for row in rows["ZROBIONE"] if row.get("OBSZAR") == "Prescot — dokumenty exact100"
    )
    delux_verified = sum(
        1 for row in rows["ZROBIONE"] if row.get("OBSZAR") == "Prescot — opisy Delux postverify"
    )
    withdrawal_statuses = Counter(as_text(row.get("WYNIK / STATUS")) for row in rows["WYCOFANE"])
    metadata = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceDirectory": source_name(DATA),
        "klusLiveReports": len(klus_live_paths),
        "noncoreLiveReports": len(noncore_live_paths),
        "deluxReport": source_name(delux_path) if delux_path else "brak",
    }
    klus_postverify_counts = (
        klus_postverify.get("counts", {}) if isinstance(klus_postverify, dict) else {}
    )
    rows["PODSUMOWANIE"] = [
        {"OBSZAR": "Raport", "MIARA": "Wygenerowano", "WARTOŚĆ": metadata["generatedAt"], "STATUS / INTERPRETACJA": "Lokalna agregacja istniejących dowodów; bez połączenia z TIM.", "ŹRÓDŁO / UWAGI": metadata["sourceDirectory"]},
        {"OBSZAR": "KLUŚ", "MIARA": "Exact100 w mapowaniu", "WARTOŚĆ": len(klus_mapping.get("exact100", [])), "STATUS / INTERPRETACJA": "Bezpieczne mapowanie dokumentów, nie jest równoznaczne z zapisem.", "ŹRÓDŁO / UWAGI": source_name(PATHS["klus_mapping"])},
        {"OBSZAR": "KLUŚ", "MIARA": "Potwierdzone zapisy exact100", "WARTOŚĆ": klus_saved, "STATUS / INTERPRETACJA": "Tylko wyniki saved/already_current/verified z raportów live.", "ŹRÓDŁO / UWAGI": f"{len(klus_live_paths)} raportów pilot/batch/retry"},
        {"OBSZAR": "KLUŚ", "MIARA": "Końcowa kontrola live", "WARTOŚĆ": f"{klus_postverify_counts.get('verified', 0)}/{klus_postverify_counts.get('total', 0)}", "STATUS / INTERPRETACJA": f"Relacje PDF dostępne: {klus_postverify_counts.get('downloadableRelations', 0)}; błędy: {klus_postverify_counts.get('failed', 0)}.", "ŹRÓDŁO / UWAGI": source_name(PATHS["klus_postverify"])},
        {"OBSZAR": "KLUŚ", "MIARA": "Exact100 pozostałe", "WARTOŚĆ": max(0, len(klus_mapping.get("exact100", [])) - klus_saved), "STATUS / INTERPRETACJA": "Brak dowodu skutecznego zapisu w uwzględnionych raportach live.", "ŹRÓDŁO / UWAGI": "Arkusz BRAKI"},
        {"OBSZAR": "KLUŚ", "MIARA": "Dopasowania 80–99%", "WARTOŚĆ": len(klus_mapping.get("fuzzy80to99", [])), "STATUS / INTERPRETACJA": "Do ręcznego potwierdzenia.", "ŹRÓDŁO / UWAGI": "Arkusz DOPASOWANIA_80_99"},
        {"OBSZAR": "KLUŚ", "MIARA": "Nieweryfikowane z partii 355", "WARTOŚĆ": len(klus_unverified), "STATUS / INTERPRETACJA": "Duplikaty/zerowa cena; bez automatycznego ponawiania.", "ŹRÓDŁO / UWAGI": source_name(PATHS["klus_unverified"])},
        {"OBSZAR": "PRESCOT", "MIARA": "Exact100 w mapowaniu", "WARTOŚĆ": len(prescot_mapping.get("exact100", [])), "STATUS / INTERPRETACJA": "Mapowanie lokalnych dokumentów.", "ŹRÓDŁO / UWAGI": source_name(PATHS["prescot_mapping"])},
        {"OBSZAR": "PRESCOT", "MIARA": "Potwierdzone zapisy exact100", "WARTOŚĆ": prescot_saved, "STATUS / INTERPRETACJA": "Nie zalicza nieudanych prób HTTP 500.", "ŹRÓDŁO / UWAGI": source_name(PATHS["prescot_live"])},
        {"OBSZAR": "PRESCOT", "MIARA": "Dopasowania 80–99%", "WARTOŚĆ": len(prescot_mapping.get("fuzzy80to99", [])), "STATUS / INTERPRETACJA": "Do ręcznego potwierdzenia.", "ŹRÓDŁO / UWAGI": "Arkusz DOPASOWANIA_80_99"},
        {"OBSZAR": "PRESCOT", "MIARA": "Opisy Delux potwierdzone", "WARTOŚĆ": delux_verified, "STATUS / INTERPRETACJA": "Najnowszy dostępny postverify.", "ŹRÓDŁO / UWAGI": metadata["deluxReport"]},
        {"OBSZAR": "EPREL", "MIARA": "Dopasowania pochodne", "WARTOŚĆ": len(rows["EPREL_POCHODNE"]), "STATUS / INTERPRETACJA": "Warianty/rodziny poniżej exact100; osobny arkusz kontrolny.", "ŹRÓDŁO / UWAGI": "Lokalne kolejki EPREL 2026-09-01/02"},
        {"OBSZAR": "Wycofania non-core", "MIARA": "Pozycje", "WARTOŚĆ": len(rows["WYCOFANE"]), "STATUS / INTERPRETACJA": as_text(dict(withdrawal_statuses)), "ŹRÓDŁO / UWAGI": f"{len(noncore_live_paths)} raporty live + kolejka"},
        {"OBSZAR": "TIM", "MIARA": "Błędy i blokery", "WARTOŚĆ": len(rows["BLEDY_TIM"]), "STATUS / INTERPRETACJA": "HTTP 500/422 i odrzucone workflow pozostają nierozwiązane.", "ŹRÓDŁO / UWAGI": "Arkusz BLEDY_TIM"},
        {"OBSZAR": "Kontrola", "MIARA": "Decyzje ręczne", "WARTOŚĆ": len(rows["DO_RECZNEJ_DECYZJI"]), "STATUS / INTERPRETACJA": "Nie zapisywać automatycznie.", "ŹRÓDŁO / UWAGI": "Arkusz DO_RECZNEJ_DECYZJI"},
        {"OBSZAR": "Kontrola", "MIARA": "Ostrzeżenia źródeł", "WARTOŚĆ": len(warnings), "STATUS / INTERPRETACJA": " | ".join(warnings) if warnings else "Brak", "ŹRÓDŁO / UWAGI": "Generator"},
    ]

    return rows, warnings, metadata


def ordered_headers(sheet_name: str, rows: list[dict[str, Any]]) -> list[str]:
    preferred = {
        "PODSUMOWANIE": ["OBSZAR", "MIARA", "WARTOŚĆ", "STATUS / INTERPRETACJA", "ŹRÓDŁO / UWAGI"],
        "ZROBIONE": ["DATA RAPORTU", "OBSZAR", "PRODUCENT", "PIMCORE ID", "EAN", "MODEL / INDEKS HANDLOWY", "NAZWA", "STAN", "OPERACJA", "POLA", "DOKUMENTY", "WYNIK", "HTTP", "ŹRÓDŁO"],
        "BRAKI": ["PRIORYTET", "PRODUCENT", "PIMCORE ID", "EAN", "MODEL / INDEKS HANDLOWY", "NAZWA", "STAN", "BRAKUJĄCE / DOCELOWE POLA", "DOKUMENTY GOTOWE", "PEWNOŚĆ %", "BLOKER / UWAGA", "ŹRÓDŁO"],
        "DOPASOWANIA_80_99": ["PRODUCENT", "PIMCORE ID", "EAN", "MODEL / INDEKS HANDLOWY", "NAZWA", "STAN", "POLE", "PEWNOŚĆ %", "TYP DOPASOWANIA", "UZASADNIENIE", "DOKUMENT / URL", "KONFLIKTY", "DECYZJA", "ŹRÓDŁO"],
        "EPREL_POCHODNE": ["PIMCORE ID", "EAN", "MODEL / INDEKS HANDLOWY", "NAZWA", "STAN", "MODEL EPREL", "NR EPREL", "KLASA", "TYP DOPASOWANIA", "PEWNOŚĆ %", "UZASADNIENIE", "EPREL URL", "ETYKIETA / PLIK", "KARTA PRODUKTU", "QR POTWIERDZONY", "KARTA POTWIERDZONA", "ŹRÓDŁO"],
        "WYCOFANE": ["PRODUCENT", "PIMCORE ID", "EAN", "MODEL / INDEKS HANDLOWY", "NAZWA", "STAN", "REGUŁA", "STATUS PRZED", "WYNIK / STATUS", "POWÓD / WALIDACJA", "KARTA", "CE", "DANE PRODUKTU NIENARUSZONE", "ŹRÓDŁO"],
        "BLEDY_TIM": ["DATA RAPORTU", "OBSZAR", "PRODUCENT", "PIMCORE ID", "EAN", "MODEL / INDEKS HANDLOWY", "NAZWA", "STAN", "STATUS", "HTTP", "KOD / POWÓD", "ZALECENIE", "ŹRÓDŁO"],
        "DO_RECZNEJ_DECYZJI": ["PRIORYTET", "OBSZAR", "PRODUCENT", "PIMCORE ID", "EAN", "MODEL / INDEKS HANDLOWY", "NAZWA", "STAN", "PEWNOŚĆ %", "PROBLEM / KONFLIKT", "BEZPIECZNY NASTĘPNY KROK", "DOKUMENT / URL", "ŹRÓDŁO"],
    }[sheet_name]
    extras = sorted({key for row in rows for key in row if key not in preferred})
    return preferred + extras


def style_sheet(ws, *, error_sheet: bool = False) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if error_sheet:
            fill = ERROR_FILL
        elif row_number % 2 == 0:
            fill = SUBTLE_FILL
        else:
            fill = None
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill
            if isinstance(cell.value, str) and re.match(r"^https?://", cell.value):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        sample = list(column)[:750]
        maximum = max((len(str(cell.value or "")) for cell in sample), default=10)
        header = str(column[0].value or "")
        cap = 70 if any(token in header for token in ("POWÓD", "UWAGA", "DOKUMENT", "URL", "KROK", "ŹRÓDŁO")) else 42
        ws.column_dimensions[letter].width = min(cap, max(11, maximum + 2))


def write_workbook(output: Path) -> dict[str, int]:
    if output.suffix.lower() != ".xlsx":
        raise ValueError("--output musi wskazywać plik z rozszerzeniem .xlsx")
    output = output.expanduser().resolve()
    if output.exists():
        raise FileExistsError(f"Plik już istnieje; generator go nie nadpisze: {output}")

    rows, _, metadata = collect_rows()
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "PRESCOT local remediation report generator"
    workbook.properties.title = "TIM — końcowy raport działań, braków i decyzji"
    workbook.properties.description = (
        "Lokalna agregacja raportów JSON/CSV; generator nie łączy się z TIM."
    )
    workbook.properties.created = datetime.now(timezone.utc).replace(tzinfo=None)

    counts: dict[str, int] = {}
    for sheet_name in SHEETS:
        sheet_rows = rows[sheet_name]
        headers = ordered_headers(sheet_name, sheet_rows)
        ws = workbook.create_sheet(sheet_name)
        ws.append(headers)
        for row in sheet_rows:
            ws.append([excel_safe(row.get(header, "")) for header in headers])
        style_sheet(ws, error_sheet=sheet_name == "BLEDY_TIM")
        counts[sheet_name] = len(sheet_rows)

    workbook.save(output)
    print(json.dumps({"output": str(output), "sheets": counts, **metadata}, ensure_ascii=False, indent=2))
    return counts


def main() -> None:
    args = parse_args()
    write_workbook(args.output)


if __name__ == "__main__":
    main()
