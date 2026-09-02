#!/usr/bin/env python3
"""Build the guarded priority EPREL queue for three new Prescot length variants."""

from __future__ import annotations

import json
import re
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

import cv2
from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path("/Users/karolbohdanowicz/my-ai-agents/prescot")
BUFFER_PATH = ROOT / "exports/tim/remediation/buffer-current-live-readonly-after-klus-activations-2026-09-01.json"
AUDIT_PATH = ROOT / "exports/tim/remediation/current-buffer-offer-audit-2026-09-01.json"
XML_PATH = ROOT / "tmp/sources/prescot-live-2026-08-31.xml"
EPREL_PATH = Path("/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/TIM.PL/EPREL_KOPIA CAŁOŚĆ.xlsx")
ASSET_ROOT = ROOT / "tmp/pdfs/eprel-buffer-priority-family"
QUEUE_PATH = ROOT / "exports/tim/remediation/buffer-eprel-priority-family6-all-queue-2026-09-01.json"
PRIORITY3_QUEUE_PATH = ROOT / "exports/tim/remediation/buffer-eprel-priority-family3-queue-2026-09-01.json"
IP67_QUEUE_PATH = ROOT / "exports/tim/remediation/buffer-eprel-ip67-family3-queue-2026-09-01.json"
EC528_QUEUE_PATH = ROOT / "exports/tim/remediation/buffer-eprel-ec528-family5-queue-2026-09-01.json"
EC320_WW27_QUEUE_PATH = ROOT / "exports/tim/remediation/buffer-eprel-ec320-ww27-family1-queue-2026-09-01.json"
NEXT_DERIVED_QUEUE_PATH = ROOT / "exports/tim/remediation/buffer-eprel-next-derived3-queue-2026-09-01.json"

TARGETS = {
    15907493: ("12EC480WW2750", "12EC480WW27", "terminal length suffix 50 removed"),
    10047256: ("E003-025-8-W100", "E003-025-8-W", "terminal length suffix 100 removed"),
    10648939: ("24EC384-042-8-NWL1", "24EC384-042-8-NWL", "terminal sale length suffix 1 removed"),
    15907502: ("24EC320WW50IP67", "24EC320WW1IP67", "50 m sale-length variant mapped to registered 1 m light source"),
    15907505: ("24EC320NW50IP67", "24EC320NW1IP67", "50 m sale-length variant mapped to registered 1 m light source"),
    15907508: ("24EC320W50IP67", "24EC320W1IP67", "50 m sale-length variant mapped to registered 1 m light source"),
    10647886: ("24EC528-045-10-NW1", "24EC528-045-10-NW", "terminal sale length suffix 1 removed"),
    10650923: ("EC528-045-10-WW1", "EC528-045-10-WW", "terminal sale length suffix 1 removed"),
    10648789: ("EC528-045-10-NW1", "EC528-045-10-NW", "terminal sale length suffix 1 removed"),
    10648891: ("EC528-045-10-W1", "EC528-045-10-W", "terminal sale length suffix 1 removed"),
    10648960: ("24EC528-045-10-WW1", "24EC528-045-10-WW", "terminal sale length suffix 1 removed"),
    15907472: ("EC320-025-8-WW2750", "EC320-025-8-WW27", "50 m sale-length variant mapped to registered light source"),
    15907481: ("EC608-013-8-CCT50", "EC608-013-8-CCT", "50 m sale-length variant mapped to registered light source"),
    10648972: ("EC608-026-5-CCT1", "EC608-026-5-CCT", "terminal sale length suffix 1 removed"),
    10649251: ("24EC384-042-8-WWL1", "24EC384-042-8-WWL", "terminal sale length suffix 1 removed"),
}
PRIORITY3_IDS = {15907493, 10047256, 10648939}
IP67_IDS = {15907502, 15907505, 15907508}
EC528_IDS = {10647886, 10650923, 10648789, 10648891, 10648960}
EC320_WW27_IDS = {15907472}
NEXT_DERIVED_IDS = {15907481, 10648972, 10649251}

# The local EPREL workbook contains a stale class for this registration. The
# official product fiche Fiche_2724835_PL.pdf explicitly states class G.
OFFICIAL_CLASS_OVERRIDES = {
    "24EC320NW1IP67": "G",
}


def xml_by_ean():
    result = {}
    for product in ET.parse(XML_PATH).getroot().findall("o"):
        attrs_node = product.find("attrs")
        attrs = {
            str(node.attrib.get("name") or "").strip(): str(node.text or "").strip()
            for node in list(attrs_node) if attrs_node is not None
        }
        ean = attrs.get("EAN", "")
        if ean:
            result[ean] = {
                "model": attrs.get("Kod producenta", "") or attrs.get("Kod_produktu", ""),
                "price": float(product.attrib.get("price", "0") or 0),
                "stock": float(product.attrib.get("stock", "0") or 0),
            }
    return result


def eprel_by_model():
    sheet = load_workbook(EPREL_PATH, read_only=True, data_only=True)["EPREL wzbogacone"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    indexes = {header: index for index, header in enumerate(headers)}
    result = {}
    for values in rows:
        model = str(values[indexes["modelIdentifier"]] or "").strip()
        if model:
            result[model] = {header: values[index] for header, index in indexes.items()}
    return result


def download(url, target):
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 TIM EPREL audit"})
    with urllib.request.urlopen(request, timeout=45) as response:
        payload = response.read()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(payload)


def prepare_assets(live_model, registered_model, registration):
    safe_live = re.sub(r"[^A-Za-z0-9._+-]+", "_", live_model)
    label_png = ASSET_ROOT / f"{safe_live}_EPREL_{registration}.png"
    label_jpg = ASSET_ROOT / f"{safe_live}_EPREL_{registration}_tim.jpg"
    fiche_pdf = ASSET_ROOT / f"{safe_live}_Fiche_{registration}_PL.pdf"
    label_api = f"https://eprel.ec.europa.eu/api/products/lightsources/{registration}/labels?noRedirect=true&format=PNG"
    request = urllib.request.Request(label_api, headers={"User-Agent": "Mozilla/5.0 TIM EPREL audit"})
    with urllib.request.urlopen(request, timeout=45) as response:
        address = json.loads(response.read().decode("utf-8"))["address"]
    if not label_png.exists():
        download(f"https://eprel.ec.europa.eu{address}", label_png)
    if not fiche_pdf.exists():
        download(f"https://eprel.ec.europa.eu/fiches/lightsources/Fiche_{registration}_PL.pdf", fiche_pdf)
    if label_png.read_bytes()[:8] != b"\x89PNG\r\n\x1a\n":
        raise ValueError(f"invalid EPREL label PNG: {label_png}")
    image = cv2.imread(str(label_png))
    decoded, _, _ = cv2.QRCodeDetector().detectAndDecode(image)
    if not decoded.rstrip("/").endswith(f"/{registration}"):
        raise ValueError(f"EPREL QR mismatch for {live_model}: {decoded}")
    if not cv2.imwrite(str(label_jpg), image, [cv2.IMWRITE_JPEG_QUALITY, 95]):
        raise ValueError(f"cannot create TIM JPEG label: {label_jpg}")
    text = "\n".join(page.extract_text() or "" for page in PdfReader(fiche_pdf).pages)
    if registered_model not in text or registration not in text:
        raise ValueError(f"EPREL fiche identity mismatch: {registered_model}/{registration}")
    return {
        "labelFile": str(label_png),
        "productInformationSheet": str(fiche_pdf),
        "labelAddress": f"https://eprel.ec.europa.eu{address}",
        "qrVerified": True,
        "ficheVerified": True,
    }


buffer = json.loads(BUFFER_PATH.read_text(encoding="utf-8"))
buffer_by_id = {int(row["id"]): row for row in buffer["items"]}
audit = json.loads(AUDIT_PATH.read_text(encoding="utf-8"))
audit_by_id = {int(row["id"]): row for row in audit["items"]}
xml = xml_by_ean()
eprel = eprel_by_model()
items = []
rejected = []

for object_id, (model, registered_model, reason) in TARGETS.items():
    product = buffer_by_id.get(object_id)
    audit_row = audit_by_id.get(object_id)
    ean = str(product.get("ean") or "") if product else ""
    xml_product = xml.get(ean)
    entry = eprel.get(registered_model)
    tim_price_value = product.get("listPrice") if product else None
    tim_price = tim_price_value.get("value") if isinstance(tim_price_value, dict) else tim_price_value
    guards = {
        "liveProduct": bool(product),
        "identity": bool(product) and product.get("model") == model,
        "state": bool(product) and product.get("state") == "new" and product.get("status") == "new" and product.get("published") is True,
        "emptyEnergy": bool(product) and not product.get("energyClassLabels") and not product.get("energyTechnicalCards") and not str(product.get("energyClass") or "").strip(),
        "description": bool(product) and model in str(product.get("descriptionHtml") or "") and not re.search(r"\b\d{13}\b", str(product.get("descriptionHtml") or "")),
        "noActiveDuplicate": bool(audit_row) and not audit_row.get("activeDuplicates"),
        "xmlIdentity": bool(xml_product) and xml_product["model"] == model,
        "xmlStock": bool(xml_product) and xml_product["stock"] > 0,
        "xmlPrice": bool(xml_product) and xml_product["price"] > 0 and abs(float(tim_price) - xml_product["price"]) < 0.0001,
        "eprelPublished": bool(entry) and str(entry.get("status") or "").upper() == "PUBLISHED" and str(entry.get("blocked") or "").lower() != "true",
    }
    if not all(guards.values()):
        rejected.append({"id": object_id, "model": model, "reason": "guard_failed", "guards": guards})
        continue
    registration = str(entry.get("registrationNumber") or "")
    energy_class = OFFICIAL_CLASS_OVERRIDES.get(registered_model, str(entry.get("energy_class") or ""))
    if not registration or energy_class not in set("ABCDEFG"):
        rejected.append({"id": object_id, "model": model, "reason": "invalid_eprel_metadata"})
        continue
    try:
        assets = prepare_assets(model, registered_model, registration)
    except Exception as error:
        rejected.append({
            "id": object_id,
            "model": model,
            "eprelModel": registered_model,
            "eprelId": registration,
            "reason": "official_eprel_assets_unavailable_or_invalid",
            "detail": str(error),
        })
        continue
    items.append({
        "pimcoreId": object_id,
        "ean": ean,
        "manufacturerCode": model,
        "state": "new",
        "timName": str(product.get("timName") or ""),
        "timListPrice": float(tim_price),
        "xmlPrice": xml_product["price"],
        "xmlStock": xml_product["stock"],
        "eprelId": registration,
        "energyClass": energy_class,
        "eprelModel": registered_model,
        "matchType": "length_variant",
        "confidence": 90,
        "matchReason": reason,
        "eprelPublicUrl": f"https://eprel.ec.europa.eu/screen/product/lightsources/{registration}",
        "eprelFicheUrl": f"https://eprel.ec.europa.eu/fiches/lightsources/Fiche_{registration}_PL.pdf",
        **assets,
    })

report = {
    "generatedAt": datetime.now(timezone.utc).isoformat(),
    "sources": {"buffer": str(BUFFER_PATH), "xml": str(XML_PATH), "eprel": str(EPREL_PATH)},
    "rules": [
        "new, published, positive-stock Prescot buffer cards only",
        "exact EAN and trade model in Prescot XML",
        "TIM price equals XML net price; no price writes",
        "no active exact model/EAN duplicate in evidence",
        "supplier-approved terminal sale-length mapping",
        "official EPREL registration published and unblocked",
        "QR, registered model and registration verified locally",
    ],
    "counts": {"items": len(items), "rejected": len(rejected)},
    "items": items,
    "rejected": rejected,
}
QUEUE_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
priority3_report = {**report, "scope": "first completed priority family", "items": [row for row in items if row["pimcoreId"] in PRIORITY3_IDS]}
priority3_report["counts"] = {"items": len(priority3_report["items"]), "rejected": 0}
ip67_report = {**report, "scope": "24EC320 IP67 50 m family", "items": [row for row in items if row["pimcoreId"] in IP67_IDS]}
ip67_report["counts"] = {"items": len(ip67_report["items"]), "rejected": 0}
ec528_report = {**report, "scope": "EC528 and 24EC528 1 m family", "items": [row for row in items if row["pimcoreId"] in EC528_IDS]}
ec528_report["counts"] = {"items": len(ec528_report["items"]), "rejected": 0}
ec320_ww27_report = {**report, "scope": "EC320 WW27 50 m family", "items": [row for row in items if row["pimcoreId"] in EC320_WW27_IDS]}
ec320_ww27_report["counts"] = {"items": len(ec320_ww27_report["items"]), "rejected": 0}
next_derived_report = {**report, "scope": "next approved Prescot length-derived families", "items": [row for row in items if row["pimcoreId"] in NEXT_DERIVED_IDS]}
next_derived_report["counts"] = {"items": len(next_derived_report["items"]), "rejected": len([row for row in rejected if row["id"] in NEXT_DERIVED_IDS])}
PRIORITY3_QUEUE_PATH.write_text(json.dumps(priority3_report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
IP67_QUEUE_PATH.write_text(json.dumps(ip67_report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
EC528_QUEUE_PATH.write_text(json.dumps(ec528_report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
EC320_WW27_QUEUE_PATH.write_text(json.dumps(ec320_ww27_report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
NEXT_DERIVED_QUEUE_PATH.write_text(json.dumps(next_derived_report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print(json.dumps({
    "output": str(QUEUE_PATH),
    "priority3Output": str(PRIORITY3_QUEUE_PATH),
    "ip67Output": str(IP67_QUEUE_PATH),
    "ec528Output": str(EC528_QUEUE_PATH),
    "ec320Ww27Output": str(EC320_WW27_QUEUE_PATH),
    "nextDerivedOutput": str(NEXT_DERIVED_QUEUE_PATH),
    "counts": report["counts"],
    "priority3Counts": priority3_report["counts"],
    "ip67Counts": ip67_report["counts"],
    "ec528Counts": ec528_report["counts"],
    "ec320Ww27Counts": ec320_ww27_report["counts"],
    "nextDerivedCounts": next_derived_report["counts"],
}, ensure_ascii=False, indent=2))
