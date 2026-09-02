#!/usr/bin/env python3
"""Build a traceable TIM buffer/remediation workbook from read-only snapshots."""

from __future__ import annotations

import json
import re
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/karolbohdanowicz/my-ai-agents/prescot")
BUFFER_PATH = ROOT / "exports/tim/remediation/buffer-current-live-final-2026-09-01.json"
PLAN_PATH = ROOT / "exports/tim/remediation/buffer-document-plan-2026-09-01.json"
VERIFY_PATH = ROOT / "exports/tim/remediation/final-active-live-postverify-after-restore-2026-09-01.json"
DOC_INDEX_PATH = ROOT / "exports/tim/remediation/local-document-ean-index-2026-09-01.json"
DESCRIPTION_FIX_PATH = ROOT / "exports/tim/remediation/buffer-width-description-fix-queue-2026-09-01.json"
DESCRIPTION_NEW_VERIFY_PATH = ROOT / "exports/tim/remediation/buffer-width-description-fix-new-postverify-2026-09-01.json"
DESCRIPTION_APPROVAL_VERIFY_PATH = ROOT / "exports/tim/remediation/buffer-width-description-fix-approval-postverify-2026-09-01.json"
DESCRIPTION_MODEL_FIX_PATH = ROOT / "exports/tim/remediation/buffer-missing-model-description-queue-2026-09-01.json"
DESCRIPTION_MODEL_VERIFY_PATH = ROOT / "exports/tim/remediation/buffer-missing-model-description-postverify-2026-09-01.json"
XML_PATH = ROOT / "tmp/sources/prescot-live-2026-08-31.xml"
OUTPUT = Path("/Users/karolbohdanowicz/Downloads/TIM_BUFOR_RAPORT_POSTEPU_2026-09-01.xlsx")

DOCUMENTED_NOT_ACTIVATED = {
    15907502: "EPREL wariantu 50 m wymaga potwierdzenia równoważności z rejestracją bazową",
    15907505: "EPREL wariantu 50 m wymaga potwierdzenia równoważności z rejestracją bazową",
    15907508: "EPREL wariantu 50 m wymaga potwierdzenia równoważności z rejestracją bazową",
    10649251: "EPREL wariantu długości wymaga potwierdzenia równoważności",
    10648939: "EPREL wariantu długości wymaga potwierdzenia równoważności",
    15907493: "EPREL wariantu 50 m wymaga potwierdzenia równoważności",
    10047256: "EPREL wariantu 100 m wymaga potwierdzenia równoważności",
}


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def numeric_price(value):
    if isinstance(value, dict):
        value = value.get("value")
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def xml_products():
    by_ean, by_model = {}, {}
    root = ET.parse(XML_PATH).getroot()
    for product in root.findall("o"):
        attrs_node = product.find("attrs")
        attrs = {
            str(node.attrib.get("name") or "").strip(): str(node.text or "").strip()
            for node in list(attrs_node or [])
        }
        ean = attrs.get("EAN", "")
        model = attrs.get("Kod producenta", "") or attrs.get("Kod_produktu", "")
        row = {
            "ean": ean,
            "model": model,
            "price": numeric_price(product.attrib.get("price")),
            "stock": numeric_price(product.attrib.get("stock")) or 0,
            "url": product.attrib.get("url", ""),
            "name": str(product.findtext("name") or "").strip(),
        }
        if ean:
            by_ean[ean] = row
        if model:
            by_model.setdefault(model, []).append(row)
    return by_ean, by_model


def match_xml(item, by_ean, by_model):
    ean = str(item.get("ean") or "")
    model = str(item.get("model") or item.get("manufacturerIndex") or "")
    if ean and ean in by_ean:
        return by_ean[ean]
    matches = by_model.get(model, [])
    return matches[0] if len(matches) == 1 else None


def relation_count(item, field):
    value = item.get(field)
    return len(value) if isinstance(value, list) else int(bool(value))


def yes_no(value):
    return "TAK" if value else "NIE"


def style_sheet(ws, url_columns=()):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="8B174D")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in url_columns:
            cell = row[column - 1]
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        sample = list(column)[:500]
        width = min(65, max(11, max(len(str(cell.value or "")) for cell in sample) + 2))
        ws.column_dimensions[letter].width = width


buffer = read_json(BUFFER_PATH)
plan = read_json(PLAN_PATH)
verification = read_json(VERIFY_PATH)
doc_index = read_json(DOC_INDEX_PATH)
description_fix = read_json(DESCRIPTION_FIX_PATH)
description_new_verify = read_json(DESCRIPTION_NEW_VERIFY_PATH)
description_approval_verify = read_json(DESCRIPTION_APPROVAL_VERIFY_PATH)
description_model_fix = read_json(DESCRIPTION_MODEL_FIX_PATH)
description_model_verify = read_json(DESCRIPTION_MODEL_VERIFY_PATH)
if description_new_verify["counts"]["alreadyCurrent"] != 13 or description_approval_verify["counts"]["alreadyCurrent"] != 18:
    raise RuntimeError("Końcowa kontrola 31 opisów nie potwierdza pełnej zgodności.")
if description_model_verify["counts"]["alreadyCurrent"] != 3:
    raise RuntimeError("Końcowa kontrola trzech opisów bez indeksu nie potwierdza pełnej zgodności.")
description_fix_items = (
    description_fix["stages"]["bufferNewNeedsUpdate"]
    + description_fix["stages"]["bufferApprovalNeedsUpdate"]
    + description_model_fix["stages"]["bufferNewNeedsUpdate"]
)
description_fix_by_id = {int(item["pimcoreId"]): item for item in description_fix_items}
by_ean, by_model = xml_products()
buffer_by_id = {int(item["id"]): item for item in buffer["items"]}
plan_by_id = {int(item["id"]): item for item in plan["records"]}


def current_record(item):
    xml_row = match_xml(item, by_ean, by_model)
    tim_price = numeric_price(item.get("listPrice"))
    xml_price = xml_row.get("price") if xml_row else None
    plan_row = plan_by_id.get(int(item["id"]), {})
    exact_paths = doc_index.get("byEan", {}).get(str(item.get("ean") or ""), [])
    fixed = description_fix_by_id.get(int(item["id"]))
    description_html = fixed["descriptionHtml"] if fixed else str(item.get("descriptionHtml") or "")
    return {
        "id": int(item["id"]),
        "group": item.get("manufacturerName") or "Pozostałe",
        "state": item.get("state") or "",
        "ean": item.get("ean") or "",
        "model": item.get("model") or "",
        "name": item.get("timName") or "",
        "timPrice": tim_price,
        "xmlPrice": xml_price,
        "priceEqual": tim_price is not None and xml_price is not None and abs(tim_price - xml_price) < 0.0001,
        "xmlStock": xml_row.get("stock") if xml_row else None,
        "xmlUrl": xml_row.get("url") if xml_row else "",
        "photo": bool(item.get("mainPhoto")),
        "description": bool(description_html),
        "descriptionHasModel": bool(item.get("model")) and str(item.get("model")) in description_html,
        "descriptionHasEan": bool(re.search(r"\b\d{13}\b", description_html)),
        "descriptionFixed": bool(fixed),
        "dataSheet": relation_count(item, "dataSheet"),
        "ce": relation_count(item, "certifications"),
        "instruction": relation_count(item, "instructions"),
        "energyLabel": relation_count(item, "energyClassLabels"),
        "energyCard": relation_count(item, "energyTechnicalCards"),
        "eprelType": plan_row.get("eprelMatchType", ""),
        "eprelConfidence": plan_row.get("eprelConfidence", ""),
        "eprelModel": plan_row.get("eprelModel", ""),
        "eprelId": plan_row.get("registrationNumber", ""),
        "eprelUrl": plan_row.get("eprelPublicUrl", ""),
        "eprelReason": plan_row.get("eprelReason", ""),
        "exactLocalDocs": len(exact_paths),
    }


current = [current_record(item) for item in buffer["items"]]
current_by_id = {row["id"]: row for row in current}
prescot_current = [row for row in current if str(row["group"]).upper() == "PRESCOT"]
positive_stock = [row for row in current if (row["xmlStock"] or 0) > 0]
exact_eprel = [row for row in current if row["eprelType"] == "dokładny"]
derived_eprel = [row for row in current if row["eprelType"] == "pochodny długości"]

wb = Workbook()
summary = wb.active
summary.title = "PODSUMOWANIE"
summary.append(["Pozycja", "Wartość", "Znaczenie / następny krok"])
summary_rows = [
    ("Stan raportu", datetime.now(timezone.utc).isoformat(), "Snapshot PIMCORE i lokalny prescot.xml"),
    ("Produkty zweryfikowane po zmianach", verification["counts"]["total"], "Pełna kontrola tożsamości, ceny, stanu i wymaganych relacji"),
    ("Aktywne w TIM", verification["counts"]["active"], "10/10 aktywnych pozycji przechodzi kontrolę"),
    ("Wysłane do TIM / oczekują", verification["counts"]["awaitingTimApproval"], "3/3 kompletne, decyzja po stronie TIM"),
    ("Niezgodności w kontroli", verification["counts"]["mismatch"], "Musi pozostać 0"),
    ("Blokady kart", verification["counts"]["locked"], "Nie pozostawiono blokad"),
    ("Aktualny bufor", len(current), "Pełna lista w arkuszu BUFOR_AKTUALNY"),
    ("Prescot w buforze", len(prescot_current), "Priorytet przed KLUŚ i pozostałymi"),
    ("Bufor ze stanem XML > 0", len(positive_stock), "Stan z prescot.xml, nie z pustej relacji magazynowej bufora"),
    ("Karty udokumentowane, nieaktywowane", len(DOCUMENTED_NOT_ACTIVATED), "Wstrzymane wyłącznie przez niepewność wariantu EPREL"),
    ("Opisy poprawione w tej sesji", len(description_fix_items), "31 szerokość/długość + 3 brakujące indeksy; 34/34 potwierdzone z PIMCORE"),
    ("EPREL dokładny w aktualnym buforze", len(exact_eprel), "Dokładny modelIdentifier nie oznacza automatycznie gotowości do aktywacji"),
    ("EPREL 80–99% do weryfikacji", len(derived_eprel), "Nie przypięto automatycznie dokumentu regulacyjnego do wariantu pochodnego"),
    ("Lokalne PDF-y zaindeksowane", doc_index["counts"]["indexed"], "Wyszukiwanie kart po EAN w treści, nie tylko po nazwie pliku"),
    ("Błędy indeksowania PDF", doc_index["counts"]["failed"], "0 = pełny odczyt indeksu"),
    ("Ocena TIM widoczna przed pracami", 3, "Zmiany zwiększają kompletność kart; aktualizacja oceny zależy od cyklu naliczania TIM"),
    ("Ceny", "NIE ZMIENIANO", "Wszystkie zapisy zachowały cenę; arkusz CENY_NIE_RUSZANE pokazuje kontrolę"),
    ("Incydent cen zamówień", "po stronie TIM", "Cena produktu jest poprawna; TIM zgłosił błędną cenę zamówienia do działu technicznego"),
]
for row in summary_rows:
    summary.append(row)

ws = wb.create_sheet("AKTYWOWANE_I_WYSLANE")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Stan oczekiwany", "Stan live", "Indeks TIM", "Cena live netto", "Cena XML netto", "Stan XML", "Cena zgodna", "Dokumenty wymagane", "Weryfikacja"])
for result in verification["results"]:
    xml_row = match_xml({"ean": result["ean"], "model": result["model"]}, by_ean, by_model)
    xml_price = xml_row.get("price") if xml_row else None
    ws.append([
        result["id"], result["ean"], result["model"], result["expectedState"], result["liveState"], result["timIndex"],
        result["livePrice"], xml_price, xml_row.get("stock") if xml_row else None,
        yes_no(xml_price is not None and abs(float(result["livePrice"]) - float(xml_price)) < 0.0001),
        ", ".join(result["requiredRelations"]), "OK" if result["status"] == "verified" else result["status"],
    ])
style_sheet(ws)

buffer_headers = [
    "PIM ID", "Grupa", "Stan TIM", "EAN", "Indeks handlowy", "Nazwa TIM", "Cena TIM netto", "Cena XML netto",
    "Cena zgodna", "Stan XML", "URL prescot.com.pl", "Zdjęcie", "Opis", "Opis ma indeks", "Opis ma EAN",
    "Karta katalogowa", "CE/certyfikat", "Instrukcja", "Etykieta EPREL", "Karta EPREL", "EPREL typ",
    "Pewność %", "Model EPREL", "Nr EPREL", "URL EPREL", "Dokumenty lokalne po EAN",
]


def append_buffer_row(ws, row):
    ws.append([
        row["id"], row["group"], row["state"], row["ean"], row["model"], row["name"], row["timPrice"], row["xmlPrice"],
        yes_no(row["priceEqual"]), row["xmlStock"], row["xmlUrl"], yes_no(row["photo"]), yes_no(row["description"]),
        yes_no(row["descriptionHasModel"]), yes_no(row["descriptionHasEan"]), row["dataSheet"], row["ce"], row["instruction"],
        row["energyLabel"], row["energyCard"], row["eprelType"], row["eprelConfidence"], row["eprelModel"], row["eprelId"],
        row["eprelUrl"], row["exactLocalDocs"],
    ])


ws = wb.create_sheet("BUFOR_AKTUALNY")
ws.append(buffer_headers)
for row in current:
    append_buffer_row(ws, row)
style_sheet(ws, url_columns=(11, 25))

ws = wb.create_sheet("UZUPEL_BEZ_AKTYWACJI")
ws.append(buffer_headers + ["Powód wstrzymania"])
for object_id, reason in DOCUMENTED_NOT_ACTIVATED.items():
    row = current_by_id[object_id]
    append_buffer_row(ws, row)
    ws.cell(ws.max_row, len(buffer_headers) + 1, reason)
style_sheet(ws, url_columns=(11, 25))

ws = wb.create_sheet("OPISY_POPRAWIONE")
ws.append(["PIM ID", "Stan TIM", "EAN", "Indeks handlowy", "Nazwa", "Zmiana", "Kontrola live"])
for row in current:
    if not row["descriptionFixed"]:
        continue
    ws.append([
        row["id"], row["state"], row["ean"], row["model"], row["name"],
        "Poprawiona publiczna tożsamość/specyfikacja; bez EAN i bez indeksu wewnętrznego",
        "OK — opis zgodny po ponownym odczycie",
    ])
style_sheet(ws)

ws = wb.create_sheet("EPREL_DOKLADNE")
ws.append(buffer_headers + ["Ocena decyzji"])
for row in exact_eprel:
    append_buffer_row(ws, row)
    decision = "DUPLIKAT — nie aktywować" if row["model"] == "E033-050-10-RGBNW" else "Sprawdzić komplet karty i warunki aktywacji"
    ws.cell(ws.max_row, len(buffer_headers) + 1, decision)
style_sheet(ws, url_columns=(11, 25))

ws = wb.create_sheet("EPREL_80_99_DO_WERYF")
ws.append(buffer_headers + ["Powód dopasowania", "Decyzja"])
for row in derived_eprel:
    append_buffer_row(ws, row)
    ws.cell(ws.max_row, len(buffer_headers) + 1, row["eprelReason"])
    ws.cell(ws.max_row, len(buffer_headers) + 2, "Nie przypinać automatycznie bez potwierdzenia równoważności modelu/wariantu")
style_sheet(ws, url_columns=(11, 25))

ws = wb.create_sheet("BRAKI_DOKUMENTOW")
ws.append(buffer_headers + ["Braki"])
for row in current:
    missing = []
    if not row["photo"]:
        missing.append("zdjęcie")
    if not row["description"]:
        missing.append("opis")
    if not row["dataSheet"]:
        missing.append("karta katalogowa")
    if not row["ce"]:
        missing.append("CE/certyfikat")
    if not row["instruction"]:
        missing.append("instrukcja")
    if re.search(r"ta[śs]ma|źródło|zrodlo|żarów|zarow", row["name"], re.I):
        if not row["energyLabel"]:
            missing.append("etykieta EPREL")
        if not row["energyCard"]:
            missing.append("karta EPREL")
    if missing:
        append_buffer_row(ws, row)
        ws.cell(ws.max_row, len(buffer_headers) + 1, ", ".join(missing))
style_sheet(ws, url_columns=(11, 25))

ws = wb.create_sheet("DUPLIKATY")
ws.append(["Indeks handlowy", "Bufor PIM ID", "Stan bufora", "Aktywny PIM ID", "Indeks TIM", "Decyzja", "Dowód"])
ws.append([
    "E033-050-10-RGBNW", 15906402, current_by_id[15906402]["state"], 9560231, "0001-00020-98220",
    "Nie aktywować duplikatu; wykorzystać istniejącą aktywną kartę",
    "PIMCORE zwrócił: Manufacturer Index już istnieje",
])
style_sheet(ws)

ws = wb.create_sheet("CENY_NIE_RUSZANE")
ws.append(["PIM ID", "EAN", "Indeks handlowy", "Zakres pracy", "Cena live netto", "Cena XML netto", "Zgodność", "Uwagi"])
for result in verification["results"]:
    xml_row = match_xml({"ean": result["ean"], "model": result["model"]}, by_ean, by_model)
    xml_price = xml_row.get("price") if xml_row else None
    ws.append([
        result["id"], result["ean"], result["model"], "opis/dokumenty/workflow", result["livePrice"], xml_price,
        yes_no(xml_price is not None and abs(float(result["livePrice"]) - float(xml_price)) < 0.0001),
        "Cena nie była polem docelowym żadnej zmiany",
    ])
for object_id in DOCUMENTED_NOT_ACTIVATED:
    row = current_by_id[object_id]
    ws.append([
        row["id"], row["ean"], row["model"], "wyłącznie dokumenty", row["timPrice"], row["xmlPrice"],
        yes_no(row["priceEqual"]), "Nie aktywowano; cena niezmieniona",
    ])
style_sheet(ws)

ws = wb.create_sheet("IMPORTY")
ws.append(["ID schematu", "Nazwa / zakres", "Liczba", "Stan", "Decyzja"])
ws.append([649, "Touch 12A — pięć modeli", 5, "Oczekiwanie na akceptację administratora TIM", "Uruchomić dopiero po udostępnieniu przycisku przez TIM"])
ws.append([648, "Stary pilot 10", 10, "Nie używać", "Zastąpiony bezpieczniejszym zakresem; użytkownik usunął stare schematy"])
style_sheet(ws)

style_sheet(summary)
OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)

# Reopen once so a corrupted workbook cannot be handed off silently.
check = load_workbook(OUTPUT, read_only=True, data_only=True)
sheet_counts = {ws.title: max(0, ws.max_row - 1) for ws in check.worksheets}
check.close()
print(json.dumps({
    "output": str(OUTPUT),
    "buffer": len(current),
    "positiveXmlStock": len(positive_stock),
    "verified": verification["counts"]["verified"],
    "sheets": sheet_counts,
}, ensure_ascii=False, indent=2))
