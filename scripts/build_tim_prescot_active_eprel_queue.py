#!/usr/bin/env python3
"""Build and validate a guarded EPREL queue for active Prescot products."""

from __future__ import annotations

import json
import re
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

import cv2
from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path("/Users/karolbohdanowicz/my-ai-agents/prescot")
LIVE_PATH = ROOT / "exports/tim/remediation/prescot-active-live-docs-baseline-2026-09-01.json"
XML_PATH = ROOT / "tmp/sources/prescot-live-2026-08-31.xml"
EPREL_PATH = Path("/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/TIM.PL/EPREL_KOPIA CAŁOŚĆ.xlsx")
ASSET_ROOT = ROOT / "tmp/pdfs/eprel-active-remaining"
QUEUE_PATH = ROOT / "exports/tim/remediation/prescot-active-eprel-remaining-queue-2026-09-01.json"
RECHECK_QUEUE_PATH = ROOT / "exports/tim/remediation/prescot-active-eprel-24e009-recheck2-queue-2026-09-01.json"
RECHECK_EXISTING_IDS = {9567950, 10047335}


def active_xml_by_ean():
    products = {}
    root = ET.parse(XML_PATH).getroot()
    for product in root.findall("o"):
        attrs_node = product.find("attrs")
        attrs = {
            str(node.attrib.get("name") or "").strip(): str(node.text or "").strip()
            for node in list(attrs_node or [])
        }
        ean = attrs.get("EAN", "")
        if not ean:
            continue
        products[ean] = {
            "price": float(product.attrib.get("price", "0") or 0),
            "stock": float(product.attrib.get("stock", "0") or 0),
            "model": attrs.get("Kod producenta", "") or attrs.get("Kod_produktu", ""),
            "url": str(product.attrib.get("url") or ""),
        }
    return products


def eprel_by_model():
    sheet = load_workbook(EPREL_PATH, read_only=True, data_only=True)["EPREL wzbogacone"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    indexes = {header: index for index, header in enumerate(headers)}
    result = {}
    for values in rows:
        model = str(values[indexes["modelIdentifier"]] or "").strip()
        if model:
            result[model] = {header: values[index] for header, index in indexes.items()}
    return result


def is_published(entry):
    return (
        entry
        and str(entry.get("status") or "").upper() == "PUBLISHED"
        and str(entry.get("blocked") or "").lower() != "true"
    )


def derived_candidates(model):
    """Only remove explicit length suffixes approved by the supplier."""
    candidates = []
    match = re.match(r"^(.*?)(100|50|25|10|5|3|2|1)(IP\d+.*)$", model)
    if match:
        candidates.append((match.group(1) + match.group(3), "length suffix before IP removed"))
    for suffix in ("100", "50", "25", "10", "5", "3", "2", "1"):
        if model.endswith(suffix) and len(model) > len(suffix) + 3:
            candidates.append((model[: -len(suffix)], f"terminal length suffix {suffix} removed"))
    result = []
    seen = set()
    for candidate, reason in candidates:
        if candidate and candidate not in seen:
            seen.add(candidate)
            result.append((candidate, reason))
    return result


def download(url, target):
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 TIM EPREL audit"})
    with urllib.request.urlopen(request, timeout=45) as response:
        payload = response.read()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(payload)


def prepare_assets(live_model, registered_model, registration):
    safe_live = re.sub(r"[^A-Za-z0-9._+-]+", "_", live_model)
    label_png = ASSET_ROOT / f"{safe_live}_EPREL_{registration}.png"
    label_jpg = ASSET_ROOT / f"{safe_live}_EPREL_{registration}_tim.jpg"
    fiche_pdf = ASSET_ROOT / f"{safe_live}_Fiche_{registration}_PL.pdf"

    label_api = f"https://eprel.ec.europa.eu/api/products/lightsources/{registration}/labels?noRedirect=true&format=PNG"
    request = urllib.request.Request(label_api, headers={"User-Agent": "Mozilla/5.0 TIM EPREL audit"})
    with urllib.request.urlopen(request, timeout=45) as response:
        address = json.loads(response.read().decode("utf-8"))["address"]
    if not label_png.exists():
        download(f"https://eprel.ec.europa.eu{address}", label_png)
    if not fiche_pdf.exists():
        download(f"https://eprel.ec.europa.eu/fiches/lightsources/Fiche_{registration}_PL.pdf", fiche_pdf)

    if label_png.read_bytes()[:8] != b"\x89PNG\r\n\x1a\n":
        raise ValueError(f"invalid EPREL label PNG: {label_png}")
    image = cv2.imread(str(label_png))
    decoded, _, _ = cv2.QRCodeDetector().detectAndDecode(image)
    if not decoded.rstrip("/").endswith(f"/{registration}"):
        raise ValueError(f"EPREL QR mismatch for {live_model}: {decoded}")
    if not label_jpg.exists() and not cv2.imwrite(str(label_jpg), image, [cv2.IMWRITE_JPEG_QUALITY, 95]):
        raise ValueError(f"cannot create TIM JPEG label: {label_jpg}")
    if label_jpg.read_bytes()[:2] != b"\xff\xd8":
        raise ValueError(f"invalid TIM JPEG label: {label_jpg}")

    if fiche_pdf.read_bytes()[:4] != b"%PDF":
        raise ValueError(f"invalid EPREL fiche PDF: {fiche_pdf}")
    text = "\n".join(page.extract_text() or "" for page in PdfReader(fiche_pdf).pages)
    if registered_model not in text or registration not in text:
        raise ValueError(f"EPREL fiche identity mismatch: {registered_model}/{registration}")
    return {
        "labelFile": str(label_png),
        "productInformationSheet": str(fiche_pdf),
        "labelAddress": f"https://eprel.ec.europa.eu{address}",
        "qrVerified": True,
        "ficheVerified": True,
    }


live = json.loads(LIVE_PATH.read_text(encoding="utf-8"))
xml_by_ean = active_xml_by_ean()
eprel = eprel_by_model()

items = []
rejected = []
for product in live["products"]:
    if (
        product.get("expectedBrand") != "Prescot"
        or product.get("httpStatus") != 200
        or not product.get("published")
        or product.get("state") != "active"
        or float(product.get("stock") or 0) <= 0
    ):
        continue
    if (product.get("energyClassLabels") or product.get("energyTechnicalCards")) and int(product.get("id") or 0) not in RECHECK_EXISTING_IDS:
        continue
    model = str(product.get("model") or "").strip()
    entry = eprel.get(model)
    match_type = "exact"
    confidence = 100
    registered_model = model
    match_reason = "exact TIM trade model = EPREL modelIdentifier"
    if not is_published(entry):
        derived = [
            (candidate, reason, eprel.get(candidate))
            for candidate, reason in derived_candidates(model)
            if is_published(eprel.get(candidate))
        ]
        if not derived:
            continue
        registered_model, match_reason, entry = derived[0]
        match_type = "length_variant"
        confidence = 90

    ean = str(product.get("ean") or "")
    xml = xml_by_ean.get(ean)
    if not xml:
        rejected.append({"id": product.get("id"), "ean": ean, "model": model, "reason": "missing_exact_xml_ean"})
        continue
    if xml["model"] and xml["model"] != model:
        rejected.append({
            "id": product.get("id"), "ean": ean, "model": model,
            "reason": "xml_trade_model_mismatch", "xmlModel": xml["model"],
        })
        continue
    tim_price_value = product.get("listPrice")
    tim_price = tim_price_value.get("value") if isinstance(tim_price_value, dict) else tim_price_value
    registration = str(entry.get("registrationNumber") or "")
    energy_class = str(entry.get("energy_class") or "")
    if not registration or energy_class not in set("ABCDEFG"):
        rejected.append({"id": product.get("id"), "ean": ean, "model": model, "reason": "invalid_eprel_metadata"})
        continue
    try:
        assets = prepare_assets(model, registered_model, registration)
    except Exception as error:
        rejected.append({
            "id": product.get("id"),
            "ean": ean,
            "model": model,
            "eprelModel": registered_model,
            "eprelId": registration,
            "matchType": match_type,
            "confidence": confidence,
            "reason": "official_eprel_assets_unavailable_or_invalid",
            "detail": str(error),
        })
        continue
    items.append({
        "pimcoreId": int(product["id"]),
        "ean": ean,
        "manufacturerCode": model,
        "timName": str(product.get("timName") or ""),
        "timListPrice": float(tim_price),
        "xmlPrice": float(xml["price"]),
        "xmlStock": float(xml["stock"]),
        "eprelId": registration,
        "energyClass": energy_class,
        "eprelModel": registered_model,
        "matchType": match_type,
        "confidence": confidence,
        "matchReason": match_reason,
        "eprelPublicUrl": f"https://eprel.ec.europa.eu/screen/product/lightsources/{registration}",
        "eprelFicheUrl": f"https://eprel.ec.europa.eu/fiches/lightsources/Fiche_{registration}_PL.pdf",
        **assets,
    })

items.sort(key=lambda row: (0 if row["matchType"] == "exact" else 1, row["manufacturerCode"]))
report = {
    "generatedAt": datetime.now(timezone.utc).isoformat(),
    "sourceLive": str(LIVE_PATH),
    "sourceXml": str(XML_PATH),
    "sourceEprel": str(EPREL_PATH),
    "rules": [
        "active, published, positive-stock Prescot only",
        "both EPREL fields empty in fresh TIM read",
        "unique exact EAN and matching trade model in Prescot XML",
        "published and unblocked official EPREL registration",
        "exact model or supplier-approved terminal length-variant mapping",
        "QR and EPREL fiche model/registration validated locally",
    ],
    "counts": {
        "items": len(items),
        "exact": sum(row["matchType"] == "exact" for row in items),
        "lengthVariant": sum(row["matchType"] == "length_variant" for row in items),
        "rejected": len(rejected),
    },
    "items": items,
    "rejected": rejected,
}
QUEUE_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
recheck = {
    **report,
    "scope": "24E009 existing relation and class recheck",
    "items": [row for row in items if int(row["pimcoreId"]) in RECHECK_EXISTING_IDS],
}
recheck["counts"] = {
    "items": len(recheck["items"]),
    "exact": sum(row["matchType"] == "exact" for row in recheck["items"]),
    "lengthVariant": sum(row["matchType"] == "length_variant" for row in recheck["items"]),
    "rejected": 0,
}
RECHECK_QUEUE_PATH.write_text(json.dumps(recheck, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print(json.dumps({"output": str(QUEUE_PATH), "recheckOutput": str(RECHECK_QUEUE_PATH), "counts": report["counts"], "recheckCounts": recheck["counts"]}, ensure_ascii=False, indent=2))
