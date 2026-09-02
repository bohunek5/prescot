#!/usr/bin/env python3
import difflib
import json
import re
import unicodedata
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

import cv2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

ROOT = Path("/Users/karolbohdanowicz/my-ai-agents/prescot")
BUFFER_PATH = ROOT / "exports/tim/remediation/buffer-current-live-readonly-after-klus-activations-2026-09-01.json"
XML_PATH = ROOT / "tmp/sources/prescot-live-2026-08-31.xml"
EPREL_PATH = Path("/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/TIM.PL/EPREL_KOPIA CAŁOŚĆ.xlsx")
DOCS_ROOT = Path("/Users/karolbohdanowicz/Desktop/_PRESCOT/_2026/TIM.PL/TIM - karty ce")
CARD_ROOT = DOCS_ROOT / "Karty katalogowe" / "Taśmy LED"
EPREL_OUTPUT = ROOT / "tmp/pdfs/eprel-buffer-exact"
QUEUE_OUTPUT = ROOT / "exports/tim/remediation/buffer-eprel-exact-queue-2026-09-01.json"
JSON_OUTPUT = ROOT / "exports/tim/remediation/buffer-document-plan-2026-09-01.json"
XLSX_OUTPUT = Path("/Users/karolbohdanowicz/Downloads/TIM_BUFOR_DOKUMENTY_I_EPREL_2026-09-01.xlsx")


def normalize(value):
    value = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^A-Z0-9]+", "", value.upper())


def is_light_source(item):
    return bool(re.search(r"ta[śs]ma|żar[oó]w|swietl[oó]w|świetl[oó]w", item.get("timName", ""), re.I))


def derived_candidates(model):
    candidates = []
    match = re.match(r"^(.*?)(100|50|25|10|5|3|2|1)(IP\d+.*)$", model)
    if match:
        candidates.append((match.group(1) + match.group(3), "usunięto końcówkę długości przed oznaczeniem IP"))
    for suffix in ("100", "50", "25", "10", "5", "3", "2", "1"):
        if model.endswith(suffix) and len(model) > len(suffix) + 3:
            candidates.append((model[:-len(suffix)], f"usunięto końcówkę długości {suffix}"))
    seen = set()
    return [(candidate, reason) for candidate, reason in candidates if candidate and not (candidate in seen or seen.add(candidate))]


def xml_products():
    by_ean, by_model = {}, {}
    root = ET.parse(XML_PATH).getroot()
    for product in root.findall("o"):
        attrs_node = product.find("attrs")
        attrs = {str(node.attrib.get("name") or "").strip(): str(node.text or "").strip() for node in list(attrs_node or [])}
        ean = attrs.get("EAN", "")
        model = attrs.get("Kod producenta", "") or attrs.get("Kod_produktu", "")
        row = {
            "shopId": product.attrib.get("id", ""),
            "url": product.attrib.get("url", ""),
            "price": float(product.attrib.get("price", "0") or 0),
            "stock": float(product.attrib.get("stock", "0") or 0),
            "available": product.attrib.get("avail", ""),
            "name": str(product.findtext("name") or "").strip(),
            "ean": ean,
            "model": model,
            "category": str(product.findtext("cat") or "").strip(),
            "mainImage": (product.find("imgs/main").attrib.get("url", "") if product.find("imgs/main") is not None else ""),
        }
        if ean:
            by_ean[ean] = row
        if model:
            by_model.setdefault(model, []).append(row)
    return by_ean, by_model


def eprel_products():
    ws = load_workbook(EPREL_PATH, read_only=True, data_only=True)["EPREL wzbogacone"]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    indexes = {header: index for index, header in enumerate(headers)}
    products = {}
    for values in rows:
        model = str(values[indexes["modelIdentifier"]] or "").strip()
        if not model:
            continue
        products[model] = {header: values[index] for header, index in indexes.items()}
    return products


def card_candidates(model, paths):
    exact, scored = [], []
    model_normalized = normalize(model)
    for path in paths:
        stem = path.stem
        stem_normalized = normalize(stem)
        if model_normalized == stem_normalized or model_normalized in stem_normalized:
            exact.append((100, path))
            continue
        generic = normalize(re.sub(r"X{2,}", "", stem, flags=re.I))
        score = int(round(100 * difflib.SequenceMatcher(None, model_normalized, generic).ratio()))
        if score >= 55:
            scored.append((score, path))
    matches = exact + sorted(scored, key=lambda row: (-row[0], str(row[1])))[:3]
    return [{"score": score, "path": str(path)} for score, path in matches[:3]]


def ce_path(item):
    model = item.get("model", "").upper()
    name = item.get("timName", "").lower()
    if model.startswith("24D") or "delux" in name:
        return DOCS_ROOT / "Taśmy LED" / "Prescot Taśmy led Delux CE 2026.pdf"
    if model.startswith(("PR", "EH")) and "economic" in name:
        return DOCS_ROOT / "Taśmy LED" / "Prescot Taśmy led Economic CE.pdf"
    if is_light_source(item):
        return DOCS_ROOT / "Taśmy LED" / "Prescot Taśmy led Premium CE 2026.pdf"
    return None


def download(url, target):
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 TIM catalogue document audit"})
    with urllib.request.urlopen(request, timeout=45) as response:
        data = response.read()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(data)
    return data


def prepare_exact_assets(entry):
    model = entry["model"]
    registration = str(entry["registrationNumber"])
    safe_model = re.sub(r"[^A-Za-z0-9._+-]+", "_", model)
    label_target = EPREL_OUTPUT / f"{safe_model}_EPREL_{registration}.png"
    fiche_target = EPREL_OUTPUT / f"{safe_model}_Fiche_{registration}_PL.pdf"

    label_api = f"https://eprel.ec.europa.eu/api/products/lightsources/{registration}/labels?noRedirect=true&format=PNG"
    request = urllib.request.Request(label_api, headers={"User-Agent": "Mozilla/5.0 TIM catalogue document audit"})
    with urllib.request.urlopen(request, timeout=45) as response:
        address = json.loads(response.read().decode("utf-8"))["address"]
    if not label_target.exists():
        download(f"https://eprel.ec.europa.eu{address}", label_target)
    if not fiche_target.exists():
        download(f"https://eprel.ec.europa.eu/fiches/lightsources/Fiche_{registration}_PL.pdf", fiche_target)

    label_bytes = label_target.read_bytes()
    if label_bytes[:8] != b"\x89PNG\r\n\x1a\n":
        raise ValueError(f"Nieprawidłowy PNG etykiety: {label_target}")
    image = cv2.imread(str(label_target))
    decoded, _, _ = cv2.QRCodeDetector().detectAndDecode(image)
    if not decoded.rstrip("/").endswith(f"/{registration}"):
        raise ValueError(f"Kod QR nie prowadzi do EPREL {registration}: {decoded}")

    fiche_bytes = fiche_target.read_bytes()
    if fiche_bytes[:4] != b"%PDF":
        raise ValueError(f"Nieprawidłowy PDF karty EPREL: {fiche_target}")
    reader = PdfReader(str(fiche_target))
    text = "\n".join((page.extract_text() or "") for page in reader.pages)
    if model not in text or registration not in text:
        raise ValueError(f"Karta EPREL nie potwierdza modelu/ID: {model}/{registration}")
    return {
        "labelFile": str(label_target),
        "productInformationSheet": str(fiche_target),
        "labelAddress": f"https://eprel.ec.europa.eu{address}",
        "qrVerified": True,
        "ficheVerified": True,
    }


def style_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="8B174D")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = min(70, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        ws.column_dimensions[letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


buffer = json.loads(BUFFER_PATH.read_text(encoding="utf-8"))
by_ean, by_model = xml_products()
eprel = eprel_products()
card_paths = sorted(CARD_ROOT.rglob("*.pdf"))

records = []
exact_queue = []
for item in buffer["items"]:
    xml_row = by_ean.get(item.get("ean", ""))
    if not xml_row:
        matches = by_model.get(item.get("model", ""), [])
        xml_row = matches[0] if len(matches) == 1 else None
    model = item.get("model", "")
    exact = eprel.get(model)
    match_type, confidence, matched_model, match_reason = "brak", 0, "", "brak dokładnego wpisu EPREL"
    if exact and str(exact.get("status") or "").upper() == "PUBLISHED" and str(exact.get("blocked") or "").lower() != "true":
        match_type, confidence, matched_model, match_reason = "dokładny", 100, model, "identyczny indeks handlowy i modelIdentifier EPREL"
    elif is_light_source(item):
        derived = [(candidate, reason, eprel[candidate]) for candidate, reason in derived_candidates(model)
                   if candidate in eprel and str(eprel[candidate].get("status") or "").upper() == "PUBLISHED"
                   and str(eprel[candidate].get("blocked") or "").lower() != "true"]
        if derived:
            matched_model, match_reason, exact = derived[0]
            match_type, confidence = "pochodny długości", 90
        else:
            suggestions = difflib.get_close_matches(model, list(eprel), n=3, cutoff=0.72)
            if suggestions:
                match_reason = "podobne wpisy: " + ", ".join(suggestions)

    ce = ce_path(item)
    cards = card_candidates(model, card_paths) if item.get("manufacturerName") == "PRESCOT" else []
    record = {
        "id": item.get("id"),
        "group": item.get("manufacturerName") or "",
        "state": item.get("state"),
        "ean": item.get("ean"),
        "model": model,
        "timName": item.get("timName"),
        "timListPrice": (item.get("listPrice") or {}).get("value") if isinstance(item.get("listPrice"), dict) else item.get("listPrice"),
        "xmlPrice": xml_row.get("price") if xml_row else None,
        "xmlStock": xml_row.get("stock") if xml_row else None,
        "xmlUrl": xml_row.get("url") if xml_row else "",
        "xmlMatched": bool(xml_row),
        "dataSheetInTim": len(item.get("dataSheet") or []),
        "ceInTim": len(item.get("certifications") or []),
        "instructionInTim": len(item.get("instructions") or []),
        "energyLabelInTim": len(item.get("energyClassLabels") or []),
        "energyCardInTim": len(item.get("energyTechnicalCards") or []),
        "cardCandidates": cards,
        "ceCandidate": str(ce) if ce and ce.exists() else "",
        "eprelMatchType": match_type,
        "eprelConfidence": confidence,
        "eprelModel": matched_model,
        "eprelReason": match_reason,
        "registrationNumber": str(exact.get("registrationNumber") or "") if exact else "",
        "energyClass": str(exact.get("energy_class") or "") if exact else "",
        "eprelPublicUrl": str(exact.get("eprel_public_url") or "") if exact else "",
        "eprelFicheUrl": str(exact.get("eprel_fiche_pl_url") or "") if exact else "",
        "eprelLengthNote": str(exact.get("public_test_sample_or_length_note") or "") if exact else "",
        "action": "",
    }
    if item.get("state") == "new" and item.get("manufacturerName") == "PRESCOT" and is_light_source(item):
        tim_price = record["timListPrice"]
        exact_offer_guard = bool(xml_row) and xml_row["stock"] > 0 and xml_row["price"] > 0 \
            and tim_price is not None and abs(float(tim_price) - float(xml_row["price"])) < 0.0001 \
            and xml_row["model"] == model and bool(item.get("ean"))
        if match_type == "dokładny" and exact_offer_guard and not item.get("energyClassLabels") and not item.get("energyTechnicalCards"):
            assets = prepare_exact_assets(record)
            record.update(assets)
            record["action"] = "gotowe do przypięcia dokładnej etykiety i karty EPREL"
            exact_queue.append({
                "pimcoreId": item["id"], "ean": item.get("ean", ""), "manufacturerCode": model,
                "timName": item.get("timName", ""), "timListPrice": record["timListPrice"],
                "xmlPrice": record["xmlPrice"], "xmlStock": record["xmlStock"],
                "eprelId": record["registrationNumber"], "energyClass": record["energyClass"],
                "eprelPublicUrl": record["eprelPublicUrl"], "eprelFicheUrl": record["eprelFicheUrl"],
                **assets,
            })
        elif match_type == "pochodny długości":
            record["action"] = "do weryfikacji równoważności wariantu długości przed przypięciem"
        elif match_type == "brak":
            record["action"] = "brak pewnego wpisu EPREL - nie przypinać"
    records.append(record)

generated_at = datetime.now(timezone.utc).isoformat()
plan = {
    "generatedAt": generated_at,
    "bufferSource": str(BUFFER_PATH),
    "xmlSource": "https://prescot.wapromag.pl/prescot.xml",
    "eprelSource": str(EPREL_PATH),
    "counts": {
        "buffer": len(records),
        "prescot": sum(row["group"] == "PRESCOT" for row in records),
        "positiveXmlStock": sum((row["xmlStock"] or 0) > 0 for row in records),
        "exactEprelReady": len(exact_queue),
        "derivedEprelReview": sum(row["eprelMatchType"] == "pochodny długości" for row in records),
    },
    "records": records,
}
JSON_OUTPUT.write_text(json.dumps(plan, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
QUEUE_OUTPUT.write_text(json.dumps({"generatedAt": generated_at, "items": exact_queue}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

wb = Workbook()
summary = wb.active
summary.title = "PODSUMOWANIE"
summary.append(["Pozycja", "Wartość", "Uwagi"])
summary_rows = [
    ("Aktualny bufor TIM", len(records), "odczyt na żywo, bez zmian"),
    ("Prescot w buforze", plan["counts"]["prescot"], "wg producenta w PIM"),
    ("Produkty ze stanem XML > 0", plan["counts"]["positiveXmlStock"], "aktualny plik prescot.xml"),
    ("EPREL dokładne - gotowe", len(exact_queue), "identyczny modelIdentifier, pobrana etykieta i karta z EPREL"),
    ("EPREL pochodne długości", plan["counts"]["derivedEprelReview"], "90%; wymagają potwierdzenia równoważności"),
    ("Aktywowane PR-MAD", 5, "36/60/100/150/200 W"),
    ("Import Touch", 5, "schemat 649 czeka na akceptację administratora TIM"),
]
for row in summary_rows:
    summary.append(row)

headers = [
    "PIM ID", "Grupa", "Stan TIM", "EAN", "Indeks handlowy", "Nazwa TIM", "Cena TIM netto", "Cena XML netto",
    "Stan XML", "URL prescot.com.pl", "Karta w TIM", "CE w TIM", "Instrukcja w TIM", "Etykieta EPREL w TIM",
    "Karta EPREL w TIM", "Kandydat karty katalogowej", "Ocena karty %", "Kandydat CE", "Typ dopasowania EPREL",
    "Pewność EPREL %", "Model EPREL", "Numer EPREL", "Klasa", "Publiczny link EPREL", "Karta EPREL PL",
    "Uwagi o długości", "Działanie",
]


def row_values(record):
    best_card = record["cardCandidates"][0] if record["cardCandidates"] else {}
    return [
        record["id"], record["group"], record["state"], record["ean"], record["model"], record["timName"],
        record["timListPrice"], record["xmlPrice"], record["xmlStock"], record["xmlUrl"], record["dataSheetInTim"],
        record["ceInTim"], record["instructionInTim"], record["energyLabelInTim"], record["energyCardInTim"],
        best_card.get("path", ""), best_card.get("score", ""), record["ceCandidate"], record["eprelMatchType"],
        record["eprelConfidence"], record["eprelModel"], record["registrationNumber"], record["energyClass"],
        record["eprelPublicUrl"], record["eprelFicheUrl"], record["eprelLengthNote"], record["action"],
    ]


for title, subset in [
    ("BUFOR_AKTUALNY", records),
    ("EPREL_DOKLADNE", [row for row in records if row["eprelMatchType"] == "dokładny"]),
    ("EPREL_DO_WERYF", [row for row in records if row["eprelMatchType"] == "pochodny długości"]),
    ("BRAKI_I_DOKUMENTY", [row for row in records if row["state"] == "new"]),
]:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for record in subset:
        ws.append(row_values(record))
    style_sheet(ws)

style_sheet(summary)
XLSX_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(XLSX_OUTPUT)
print(json.dumps({"xlsx": str(XLSX_OUTPUT), "json": str(JSON_OUTPUT), "queue": str(QUEUE_OUTPUT), "counts": plan["counts"]}, ensure_ascii=False, indent=2))
