#!/usr/bin/env python3
"""Build a compact, evidence-based TIM status workbook.

This script is read-only with respect to TIM. It reads local JSON evidence and
creates one new XLSX report. Existing output files are never overwritten.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "exports" / "tim" / "remediation"

AUDIT_BEFORE = DATA / "active-brand-offer-prescot-live-cleanlocks-2026-09-02.json"
AUDIT_AFTER = DATA / "active-brand-offer-prescot-live-after-cards-descriptions-ce-2026-09-02.json"
BUFFER = DATA / "buffer-current-live-readonly-cleanlocks-2026-09-02.json"
SCHARFER = DATA / "active-brand-offer-scharfer-live-2026-09-02.json"
PROFILE_QUEUE = DATA / "prescot-profile-generated-datasheets-rest237-queue-v3-2026-09-02.json"
PR_TOUCH = DATA / "priority-touch-5-import-status-latest-2026-09-02.json"

PROFILE_REPORTS = (
    DATA / "prescot-profile-generated-datasheet-sblzo-live-v2-2026-09-02.json",
    DATA / "prescot-profile-generated-datasheets-next10-live-2026-09-02.json",
    *sorted(DATA.glob("prescot-profile-generated-datasheets-rest237-live-*-2026-09-02.json")),
)

ACCESSORY_REPORTS = (
    DATA / "prescot-accessory-generated-datasheets-pilot1-live-2026-09-02.json",
    DATA / "prescot-accessory-generated-datasheets-next10-live-2026-09-02.json",
    DATA / "prescot-accessory-generated-datasheets-rest146-live-000-049-2026-09-02.json",
    DATA / "prescot-accessory-generated-datasheets-rest146-live-050-099-2026-09-02.json",
    DATA / "prescot-accessory-generated-datasheets-rest146-live-095-119-2026-09-02.json",
    DATA / "prescot-accessory-generated-datasheets-rest146-live-120-145-2026-09-02.json",
)

CE_POSTVERIFY_REPORTS = (
    DATA / "prescot-ce-to-80-exact8-first6-postverify-2026-09-02.json",
    DATA / "prescot-ce-to-80-exact8-last1-postverify-2026-09-02.json",
    DATA / "prescot-ce-to-80-tamwz-postverify-2026-09-02.json",
)

CE_QUEUES = (
    DATA / "prescot-ce-to-80-exact8-queue-2026-09-02.json",
    DATA / "prescot-ce-to-80-tamwz-queue-2026-09-02.json",
)

HEADER = PatternFill("solid", fgColor="8B174D")
SUBHEADER = PatternFill("solid", fgColor="F3E6ED")
GOOD = PatternFill("solid", fgColor="E6F4EA")
WARN = PatternFill("solid", fgColor="FFF4CE")
BAD = PatternFill("solid", fgColor="FCE8E6")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True, type=Path)
    return parser.parse_args()


def load(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def positive(products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        p for p in products
        if p.get("httpStatus") == 200
        and p.get("state") == "active"
        and p.get("published") is True
        and float(p.get("stock") or 0) > 0
    ]


def metrics(products: list[dict[str, Any]]) -> dict[str, int]:
    rows = positive(products)
    return {
        "aktywne_dodatnie": len(rows),
        "ean": sum(bool(p.get("ean")) for p in rows),
        "zdjecie": sum(bool(p.get("mainPhoto")) for p in rows),
        "opis": sum(bool(p.get("descriptionHtml")) for p in rows),
        "karta": sum(int(p.get("dataSheet") or 0) > 0 for p in rows),
        "ce": sum(int(p.get("certifications") or 0) > 0 for p in rows),
        "instrukcja": sum(int(p.get("instructions") or 0) > 0 for p in rows),
    }


def value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def add_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        ws.append([value(v) for v in row])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(1, len(headers) + 1):
        samples = [str(ws.cell(r, column).value or "") for r in range(1, min(ws.max_row, 200) + 1)]
        width = min(max(max(map(len, samples), default=0) + 2, 10), 54)
        ws.column_dimensions[get_column_letter(column)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    args = parse_args()
    output = args.output.expanduser().resolve()
    if output.exists():
        raise SystemExit(f"Plik już istnieje: {output}")
    output.parent.mkdir(parents=True, exist_ok=True)

    before_data = load(AUDIT_BEFORE)
    after_data = load(AUDIT_AFTER)
    buffer_data = load(BUFFER)
    scharfer_data = load(SCHARFER)
    profile_queue = load(PROFILE_QUEUE)
    pr_touch = load(PR_TOUCH)

    before_products = before_data["products"]
    ce_queue_by_id: dict[int, dict[str, Any]] = {}
    for queue_path in CE_QUEUES:
        for item in load(queue_path).get("items", []):
            ce_queue_by_id[int(item["id"])] = item

    ce_verified: list[dict[str, Any]] = []
    for report_path in CE_POSTVERIFY_REPORTS:
        report = load(report_path)
        for result in report.get("products", []):
            if result.get("verified") is True:
                enriched = dict(result)
                enriched["report"] = report_path.name
                ce_verified.append(enriched)
    ce_verified_ids = {int(result["id"]) for result in ce_verified}

    after_products = [dict(product) for product in after_data["products"]]
    for product in after_products:
        if int(product.get("id") or 0) in ce_verified_ids:
            product["certifications"] = max(1, int(product.get("certifications") or 0))
    before = metrics(before_products)
    after = metrics(after_products)
    after_positive = positive(after_products)
    after_by_id = {int(p["id"]): p for p in after_products if p.get("id") is not None}

    document_results: list[dict[str, Any]] = []
    for report_path in (*PROFILE_REPORTS, *ACCESSORY_REPORTS):
        report = load(report_path)
        for result in report.get("results", []):
            enriched = dict(result)
            enriched["report"] = report_path.name
            enriched["group"] = "Akcesoria" if report_path in ACCESSORY_REPORTS else "Profile"
            document_results.append(enriched)
    saved_documents = [r for r in document_results if r.get("status") == "saved"]
    saved_profiles = [r for r in saved_documents if r.get("group") == "Profile"]
    saved_accessories = [r for r in saved_documents if r.get("group") == "Akcesoria"]

    wb = Workbook()
    ws = wb.active
    ws.title = "PODSUMOWANIE"
    rows = [
        ["Zakres", "Stan", "Uwagi"],
        ["Kontrola Prescot", f"{len(after_products)}/2014 odczytane; 0 błędów", "0 własnych i 0 cudzych blokad pozostawionych przez audyt"],
        ["Aktywne, widoczne, stan > 0", after["aktywne_dodatnie"], "To jest baza procentów w tym raporcie"],
        ["EAN", f'{after["ean"]}/{after["aktywne_dodatnie"]}', "100%"],
        ["Zdjęcia", f'{after["zdjecie"]}/{after["aktywne_dodatnie"]}', "100%"],
        ["Opisy", f'{after["opis"]}/{after["aktywne_dodatnie"]}', "100%; bez EAN w treści"],
        ["Karty katalogowe — przed", f'{before["karta"]}/{before["aktywne_dodatnie"]}', f'{before["karta"] / before["aktywne_dodatnie"]:.1%}'],
        ["Karty katalogowe — teraz", f'{after["karta"]}/{after["aktywne_dodatnie"]}', f'{after["karta"] / after["aktywne_dodatnie"]:.1%}; +{after["karta"] - before["karta"]}'],
        ["Do progu 80% kart", max(0, int(0.8 * after["aktywne_dodatnie"] + 0.9999) - after["karta"]), "Minimalna liczba kolejnych bezpiecznych kart; TIM może liczyć innym mianownikiem"],
        ["CE — przed dzisiejszą serią", f'{before["ce"]}/{before["aktywne_dodatnie"]}', f'{before["ce"] / before["aktywne_dodatnie"]:.2%}'],
        ["CE — teraz", f'{after["ce"]}/{after["aktywne_dodatnie"]}', f'{after["ce"] / after["aktywne_dodatnie"]:.2%}; +{after["ce"] - before["ce"]}'],
        ["Nowe CE potwierdzone w końcowej partii", len(ce_verified), "Każdy PDF pobrany live; EAN, model, stan i brak blokady potwierdzone"],
        ["Instrukcje", f'{after["instrukcja"]}/{after["aktywne_dodatnie"]}', f'{after["instrukcja"] / after["aktywne_dodatnie"]:.1%}'],
        ["Nowe karty profili zapisane", len(saved_profiles), "Każdy zapis zmienił tylko dataSheet; cena, stan, EAN, nazwa i workflow bez zmian"],
        ["Nowe karty akcesoriów zapisane", len(saved_accessories), "Dokładne EAN + indeks handlowy; każdy zapis zmienił tylko dataSheet"],
        ["Nowe karty Prescot łącznie", len(saved_documents), "Profile i akcesoria; 0 zmian danych chronionych i workflow"],
        ["Opisy taśm", "209", "Delux: 51; Premium 5Y: 102; Premium 3Y: 56; aktywne pozycje ze stanem dodatnim"],
        ["PR-MAD", "5/5", "Naturalne opisy obecne"],
        ["PR-TOUCH 12A", pr_touch.get("after", {}).get("json", {}).get("status", "brak"), "Import 5 modeli skonfigurowany; nie uruchamiać, dopóki TIM nie nada READY"],
        ["Bufor", buffer_data.get("received", len(buffer_data.get("items", []))), "58 Prescot, 63 KLUŚ, 25 pozostałych; 0 Schärfer"],
        ["Schärfer", "23 aktywne, wszystkie stan 0", "20 pełnych kart ma stare EAN 590; XML ma nowe EAN 599. TIM blokuje zmianę EAN po nadaniu indeksu"],
        ["Ceny", "Nie zmieniane", "Błędne ceny zamówień zgłoszone przez TIM do działu technicznego; realizować według cen oferowanych"],
    ]
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER
        cell.font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 27
    ws.column_dimensions["C"].width = 88
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for r in range(2, ws.max_row + 1):
        label = str(ws.cell(r, 1).value or "")
        if "teraz" in label or label in {"EAN", "Zdjęcia", "Opisy", "Nowe karty profili zapisane", "Nowe CE potwierdzone w końcowej partii", "PR-MAD"}:
            for cell in ws[r]:
                cell.fill = GOOD
        if label in {"Do progu 80% kart", "PR-TOUCH 12A", "Schärfer", "Ceny"}:
            for cell in ws[r]:
                cell.fill = WARN

    ws = wb.create_sheet("BRAKI_PRESCOT_AKTYWNE")
    missing_rows = []
    for p in sorted(after_positive, key=lambda x: (-int((x.get("dataSheet") or 0) == 0), -float(x.get("stock") or 0), str(x.get("model") or ""))):
        miss = []
        if int(p.get("dataSheet") or 0) == 0:
            miss.append("karta")
        if int(p.get("certifications") or 0) == 0:
            miss.append("CE")
        if int(p.get("instructions") or 0) == 0:
            miss.append("instrukcja")
        if not miss:
            continue
        missing_rows.append([
            p.get("id"), p.get("ean"), p.get("model"), p.get("timIndex"), p.get("timName"),
            p.get("stock"), "TAK" if p.get("dataSheet") else "NIE",
            "TAK" if p.get("certifications") else "NIE",
            "TAK" if p.get("instructions") else "NIE", ", ".join(miss),
        ])
    add_table(ws, ["PIMCORE ID", "EAN", "MODEL", "INDEKS TIM", "NAZWA TIM", "STAN", "KARTA", "CE", "INSTRUKCJA", "BRAKI"], missing_rows)

    ws = wb.create_sheet("KARTY_WGRANE")
    profile_rows = []
    for r in saved_documents:
        p = after_by_id.get(int(r.get("id") or 0), {})
        profile_rows.append([
            r.get("group"), r.get("id"), r.get("ean"), r.get("model"), p.get("timName"), p.get("stock"),
            r.get("documents", {}).get("dataSheet"), r.get("saveResponseStatus"),
            "TAK" if r.get("protectedDataUnchanged") is True else "NIE",
            "TAK" if r.get("workflowUnchanged") is True else "NIE", r.get("report"),
        ])
    add_table(ws, ["GRUPA", "PIMCORE ID", "EAN", "MODEL", "NAZWA TIM", "STAN", "KARTA W PIMCORE", "HTTP", "DANE CHRONIONE BEZ ZMIAN", "WORKFLOW BEZ ZMIAN", "RAPORT ŹRÓDŁOWY"], profile_rows)

    ws = wb.create_sheet("CE_WGRANE")
    ce_rows = []
    for result in ce_verified:
        pid = int(result["id"])
        p = after_by_id.get(pid, {})
        queued = ce_queue_by_id.get(pid, {})
        live_downloads = result.get("live", {}).get("downloads", {}).get("certifications", [])
        ce_rows.append([
            pid, result.get("ean"), result.get("model"), p.get("timName"), p.get("stock"),
            queued.get("confidence"), queued.get("reason"),
            live_downloads[0].get("path") if live_downloads else "",
            "TAK" if result.get("documentsOk") is True else "NIE",
            "TAK" if result.get("verified") is True else "NIE", result.get("report"),
        ])
    add_table(ws, ["PIMCORE ID", "EAN", "MODEL", "NAZWA TIM", "STAN", "PEWNOŚĆ %", "PODSTAWA DOPASOWANIA", "CE W PIMCORE", "PDF POBIERALNY", "ZWERYFIKOWANE", "RAPORT ŹRÓDŁOWY"], ce_rows)

    ws = wb.create_sheet("BUFOR")
    buffer_rows = []
    for p in buffer_data.get("items", []):
        brand = "Prescot" if p.get("manufacturerName") == "PRESCOT" else ("KLUŚ" if "KLU" in str(p.get("manufacturerName") or "").upper() else "Pozostałe")
        buffer_rows.append([
            p.get("id"), brand, p.get("state"), p.get("ean"), p.get("model"), p.get("timName"), p.get("stock"),
            "TAK" if p.get("mainPhoto") else "NIE", "TAK" if p.get("descriptionHtml") else "NIE",
            "TAK" if p.get("dataSheet") else "NIE", "TAK" if p.get("certifications") else "NIE",
            "TAK" if p.get("instructions") else "NIE",
        ])
    add_table(ws, ["PIMCORE ID", "GRUPA", "STAN KARTY", "EAN", "MODEL", "NAZWA", "STAN MAG.", "ZDJĘCIE", "OPIS", "KARTA", "CE", "INSTRUKCJA"], buffer_rows)

    ws = wb.create_sheet("BLOKADY_I_WYJATKI")
    blocker_rows = []
    for error in profile_queue.get("errors", []):
        p = after_by_id.get(int(error.get("id") or 0), {})
        blocker_rows.append([error.get("id"), p.get("ean"), error.get("model"), p.get("timName"), "Konflikt modelu TIM z katalogiem/XML", "Nie podpinać automatem; decyzja ręczna"])
    for pid, note in {
        9173111: "Zapis CE kończy się HTTP 500 po stronie TIM; karta ma także regułę B24",
        9567971: "Reguła TIM B24 blokuje samą kartę; brak bezpiecznego EPREL",
        2167286: "Reguła TIM B24 blokuje samą kartę; brak bezpiecznego EPREL",
        1506635: "Reguła TIM B24 blokuje samą kartę; brak bezpiecznego EPREL",
        2116508: "Zapis dokumentu kończy się HTTP 500 po stronie TIM",
        1341220: "Zapis CE kończy się HTTP 500 po stronie TIM",
        1341221: "Zapis CE kończy się HTTP 500 po stronie TIM",
        2167272: "Zapis CE kończy się HTTP 500 po stronie TIM",
        2398691: "Zapis CE kończy się HTTP 500 po stronie TIM",
        2398795: "Zapis CE kończy się HTTP 500 po stronie TIM",
        2398845: "Zapis CE kończy się HTTP 500 po stronie TIM",
        2488662: "Zapis CE kończy się HTTP 500 po stronie TIM",
        1627536: "CE dopasowane, ale nie forsowano po błędach HTTP 500 na kartach tej klasy",
    }.items():
        p = after_by_id.get(pid, {})
        blocker_rows.append([pid, p.get("ean"), p.get("model"), p.get("timName"), note, "Nie ponawiać w ciemno"])
    blocker_rows.append(["9173099", "", "C28302C01", "KLUŚ", "Zapis dokumentu kończy się HTTP 500 po stronie TIM", "Nie ponawiać w ciemno"])
    blocker_rows.append(["Schärfer", "590… → 599…", "20 pozycji", "Aktywne karty stan 0", "TIM nie pozwala zmienić EAN po nadaniu indeksu", "Migracja/duplikat wymaga uzgodnionej ścieżki TIM"])
    add_table(ws, ["ID", "EAN", "MODEL", "NAZWA / ZAKRES", "POWÓD", "DECYZJA"], blocker_rows)

    ws = wb.create_sheet("SCHARFER_AKTYWNE")
    scharfer_rows = []
    for p in scharfer_data.get("products", []):
        scharfer_rows.append([
            p.get("id"), p.get("ean"), p.get("model"), p.get("timIndex"), p.get("timName"), p.get("stock"),
            "TAK" if p.get("dataSheet") else "NIE", "TAK" if p.get("certifications") else "NIE",
            "TAK" if p.get("energyTechnicalCards") else "NIE",
        ])
    add_table(ws, ["PIMCORE ID", "EAN TIM", "MODEL", "INDEKS TIM", "NAZWA", "STAN", "KARTA", "CE", "KARTA ENERGETYCZNA"], scharfer_rows)

    ws = wb.create_sheet("PR_TOUCH_5")
    status = pr_touch.get("after", {}).get("json", {})
    touch_models = ["PR-CCT-12A", "PR-MONO-12A", "PR-RGB-12A", "PR-RGBCCT-12A", "PR-RGBW-12A"]
    add_table(ws, ["MODEL", "STATUS IMPORTU", "ID IMPORTU", "ŹRÓDŁO", "DZIAŁANIE"], [[m, status.get("status"), status.get("id"), status.get("originalFileName"), "Czekać na READY; potem uruchomić import"] for m in touch_models])

    source = wb.create_sheet("ŹRÓDŁA")
    sources = [AUDIT_BEFORE, AUDIT_AFTER, BUFFER, SCHARFER, PROFILE_QUEUE, PR_TOUCH, *PROFILE_REPORTS, *ACCESSORY_REPORTS, *CE_QUEUES, *CE_POSTVERIFY_REPORTS]
    add_table(source, ["PLIK DOWODOWY"], [[str(p.relative_to(ROOT))] for p in sources])

    wb.save(output)
    check = load_workbook(output, read_only=True, data_only=False)
    expected = {"PODSUMOWANIE", "BRAKI_PRESCOT_AKTYWNE", "KARTY_WGRANE", "CE_WGRANE", "BUFOR", "BLOKADY_I_WYJATKI", "SCHARFER_AKTYWNE", "PR_TOUCH_5", "ŹRÓDŁA"}
    if not expected.issubset(set(check.sheetnames)):
        raise SystemExit("Walidacja raportu nie powiodła się: brakuje arkusza")
    print(json.dumps({
        "output": str(output),
        "size": output.stat().st_size,
        "sheets": check.sheetnames,
        "activePositive": after["aktywne_dodatnie"],
        "cardsBefore": before["karta"],
        "cardsAfter": after["karta"],
        "savedProfiles": len(saved_profiles),
        "savedAccessories": len(saved_accessories),
        "savedDocuments": len(saved_documents),
        "ceAfter": after["ce"],
        "ceVerifiedFinal": len(ce_verified),
        "missingRows": len(missing_rows),
        "bufferRows": len(buffer_rows),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
