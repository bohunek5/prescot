#!/usr/bin/env python3
"""
Usuwa produkty ze stanem 0 poza taśmami LED z plików EL-PLUS / MegaCennik:
- Pozostawia taśmy LED (chyba że w nazwie mają 'wyc').
- Usuwa ze stanem 0: profile, osłony, zaślepki, akcesoria, sterowniki, zestawy, żarówki.
- Usuwa także wszelkie pozycje z 'wyc' w nazwie.
- Przenumerowuje Lp od 1 do N.
"""

import openpyxl
import re
import os

FILES = [
    "/Users/karolbohdanowicz/Downloads/1294 EL-PLUS.xlsx",
    "/Users/karolbohdanowicz/Downloads/PRESCOT_EL-PLUS_ZESTAWIENIE_KODOW_2026.xlsx"
]

def clean_zero_stock(fpath):
    if not os.path.exists(fpath):
        print(f"Brak pliku: {fpath}")
        return

    print(f"\n--- Przetwarzanie: {fpath} ---")
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active

    rows_to_delete = []

    for r in range(2, ws.max_row + 1):
        lp = ws.cell(r, 1).value
        mcode = str(ws.cell(r, 2).value or "").strip()
        code = str(ws.cell(r, 3).value or "").strip()
        ean = str(ws.cell(r, 4).value or "").strip()
        name = str(ws.cell(r, 5).value or "").strip()
        cat = str(ws.cell(r, 6).value or "").strip()
        
        raw_stock = ws.cell(r, 11).value
        try:
            stock = float(raw_stock or 0)
        except:
            stock = 0.0

        uname = name.lower()
        has_wyc = bool(re.search(r'\bwyc\.?\b', uname, re.I) or 'wycofan' in uname)

        # Czy to rzeczywista taśma LED?
        is_tape = False
        if uname.startswith('taśma') or uname.startswith('tasma'):
            if not any(w in uname for w in ['zaślepk', 'uchwyt', 'złączk', 'klips', 'profil', 'zestaw zaślepek']):
                is_tape = True

        # Warunki usunięcia:
        # 1. Wszystko co ma 'wyc'
        # 2. Wszystko co ma stan 0 i NIE jest taśmą LED
        should_delete = False
        reason = ""

        if has_wyc:
            should_delete = True
            reason = "WYC_IN_NAME"
        elif stock <= 0 and not is_tape:
            should_delete = True
            reason = "NON_TAPE_STOCK_0"

        if should_delete:
            rows_to_delete.append((r, lp, mcode, name, cat, stock, reason))

    print(f"Znaleziono {len(rows_to_delete)} pozycji do usunięcia (stan 0 poza taśmami LED).")
    
    # Usuwanie od dołu do góry
    for r, lp, mcode, name, cat, stock, reason in sorted(rows_to_delete, key=lambda x: x[0], reverse=True):
        ws.delete_rows(r, 1)

    # Przenumerowanie Lp
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).value = r - 1

    remaining = ws.max_row - 1
    wb.save(fpath)
    print(f"✅ Zapisano {fpath}:")
    print(f"   Usunięto: {len(rows_to_delete)} pozycji")
    print(f"   Pozostało produktów: {remaining}")
    print(f"   Przenumerowano Lp: 1 do {remaining}")

for f in FILES:
    clean_zero_stock(f)
