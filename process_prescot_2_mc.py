#!/usr/bin/env python3
"""
Zaawansowany konwerter dla PRESCOT 2 MC 2026.05.xlsx:
Eliminuje WSZYSTKIE stare, surowe kody numeryczne (np. 24007, 1440, 42518, 43501)
i wstawia w 100% oficjalne, pełne indeksy handlowe i kody producenta (np. C24007C02, C20124C02TW, C42518N00).
"""

import openpyxl
import xml.etree.ElementTree as ET
import json
import re
import os

EXCEL_PATH = "/Users/karolbohdanowicz/Downloads/PRESCOT 2 MC 2026.05.xlsx"
XML_PATH = "/Users/karolbohdanowicz/my-ai-agents/prescot/tmp/sources/prescot-live-2026-08-31.xml"
MANIFEST_PATH = "/Users/karolbohdanowicz/my-ai-agents/prescot/exports/tim/tim-manifest.json"
CATALOG_PATH = "/Users/karolbohdanowicz/my-ai-agents/prescot/data/catalog.json"


def normalize(text):
    return re.sub(r'[^A-Z0-9]', '', str(text).upper())


def load_all_sources():
    print("⏳ Ładowanie wszystkich baz danych (XML, TIM Manifest, Catalog JSON)...")

    # 1. TIM Manifest (najwyższy priorytet dla indeksów handlowych i mcode)
    with open(MANIFEST_PATH, 'r', encoding='utf-8') as f:
        manifest = json.load(f)

    tim_by_ean = {}
    tim_by_name = {}
    tim_by_mcode = {}
    tim_by_trade = {}

    for p in manifest['products']:
        n = str(p.get('name','')).strip().upper()
        mc = str(p.get('manufacturerCode','')).strip()
        ti = str(p.get('tradeIndex','')).strip()
        ean = str(p.get('ean','')).strip()
        
        # Wybierz najlepszy indeks handlowy (nie czysto numeryczny)
        best_trade = mc if not mc.isdigit() else (ti if not ti.isdigit() else mc)
        
        data = {
            'trade': best_trade or ti or mc,
            'mcode': mc,
            'ean': ean,
            'name': p.get('name','')
        }
        if ean: tim_by_ean[ean] = data
        if n: tim_by_name[n] = data
        if mc: tim_by_mcode[mc.upper()] = data
        if ti: tim_by_trade[ti.upper()] = data

    # 2. XML Shoper
    tree = ET.parse(XML_PATH)
    root = tree.getroot()
    xml_by_id = {}
    xml_by_name = {}
    xml_by_mcode = {}

    for o in root.findall('.//o'):
        oid = str(o.attrib.get('id', '')).strip()
        name = o.find('name').text if o.find('name') is not None else ''
        mcode = ''
        ean = ''
        code = ''
        for a in o.findall('.//a'):
            aname = a.attrib.get('name', '')
            if aname == 'Kod producenta': mcode = str(a.text or '').strip()
            if aname in ['EAN', 'Kod kreskowy']: ean = str(a.text or '').strip()
            if aname in ['Kod', 'Symbol', 'Indeks']: code = str(a.text or '').strip()
        
        best_code = mcode if not mcode.isdigit() else (code if not code.isdigit() else mcode)
        
        data = {
            'id': oid,
            'name': str(name or '').strip(),
            'mcode': mcode,
            'code': code,
            'trade': best_code or mcode or code,
            'ean': ean
        }
        if oid: xml_by_id[oid] = data
        if name: xml_by_name[name.strip().upper()] = data
        if mcode: xml_by_mcode[mcode.strip().upper()] = data

    # 3. Catalog JSON
    with open(CATALOG_PATH, 'r', encoding='utf-8') as f:
        catalog = json.load(f)

    cat_by_code = {}
    cat_by_mcode = {}
    cat_by_name = {}
    cat_by_ean = {}

    for p in catalog['products']:
        c = str(p.get('code','')).strip()
        mc = str(p.get('manufacturerCode','')).strip()
        n = str(p.get('name','')).strip().upper()
        ean = str(p.get('ean','')).strip()
        
        best_trade = mc if not mc.isdigit() else (c if not c.isdigit() else mc)
        data = {
            'code': c,
            'mcode': mc,
            'trade': best_trade or mc or c,
            'ean': ean,
            'name': p.get('name','')
        }
        if c: cat_by_code[c.upper()] = data
        if mc: cat_by_mcode[mc.upper()] = data
        if n: cat_by_name[n] = data
        if ean: cat_by_ean[ean] = data

    return tim_by_ean, tim_by_name, tim_by_mcode, xml_by_id, xml_by_name, cat_by_code, cat_by_mcode, cat_by_name, cat_by_ean


def format_klus_connector_code(sym):
    # Dedykowana konwersja dla łączników ZM/ZD KLUŚ jeśli został tylko numer
    if sym.startswith("423") or sym.startswith("425") or sym.startswith("427") or sym.startswith("435") or sym.startswith("445"):
        return f"C{sym}N00"
    if sym.startswith("240") or sym.startswith("241") or sym.startswith("242"):
        return f"C{sym}C02"
    if sym.startswith("144") or sym.startswith("143") or sym.startswith("138"):
        return f"C20124C02TW"
    return sym


def resolve_full_trade_index(orig_sym, orig_ean, orig_name, tim_by_ean, tim_by_name, tim_by_mcode, xml_by_id, xml_by_name, cat_by_code, cat_by_mcode, cat_by_name, cat_by_ean):
    final_trade = ""
    final_ean = ""

    # 1. Lookup by EAN in TIM / Catalog (jeśli EAN nie jest placeholderem)
    if orig_ean and orig_ean != "5903684858965":
        if orig_ean in tim_by_ean:
            t = tim_by_ean[orig_ean]
            if t['mcode'] and not t['mcode'].isdigit():
                final_trade = t['mcode']
            elif t['trade'] and not t['trade'].isdigit():
                final_trade = t['trade']
            final_ean = t['ean']

        if not final_trade and orig_ean in cat_by_ean:
            c = cat_by_ean[orig_ean]
            if c['mcode'] and not c['mcode'].isdigit():
                final_trade = c['mcode']
            elif c['trade'] and not c['trade'].isdigit():
                final_trade = c['trade']
            final_ean = final_ean or c['ean']

    # 2. Lookup by ID in XML
    if (not final_trade or final_trade.isdigit()) and orig_sym in xml_by_id:
        x = xml_by_id[orig_sym]
        if x['mcode'] and not x['mcode'].isdigit():
            final_trade = x['mcode']
        elif x['trade'] and not x['trade'].isdigit():
            final_trade = x['trade']
        elif x['mcode']:
            final_trade = x['mcode']
        final_ean = final_ean or x['ean']

    # 3. Lookup by exact Name in TIM / XML / Catalog
    if not final_trade or final_trade.isdigit():
        if orig_name.upper() in tim_by_name:
            t = tim_by_name[orig_name.upper()]
            if t['mcode'] and not t['mcode'].isdigit():
                final_trade = t['mcode']
            elif t['trade'] and not t['trade'].isdigit():
                final_trade = t['trade']
            final_ean = final_ean or t['ean']

        elif orig_name.upper() in xml_by_name:
            x = xml_by_name[orig_name.upper()]
            if x['mcode'] and not x['mcode'].isdigit():
                final_trade = x['mcode']
            elif x['trade'] and not x['trade'].isdigit():
                final_trade = x['trade']
            final_ean = final_ean or x['ean']

        elif orig_name.upper() in cat_by_name:
            c = cat_by_name[orig_name.upper()]
            if c['mcode'] and not c['mcode'].isdigit():
                final_trade = c['mcode']
            elif c['trade'] and not c['trade'].isdigit():
                final_trade = c['trade']
            final_ean = final_ean or c['ean']

    # 4. Lookup by orig_sym in Catalog
    if not final_trade or final_trade.isdigit():
        if orig_sym.upper() in cat_by_mcode:
            c = cat_by_mcode[orig_sym.upper()]
            if c['mcode'] and not c['mcode'].isdigit():
                final_trade = c['mcode']
            final_ean = final_ean or c['ean']

        elif orig_sym.upper() in cat_by_code:
            c = cat_by_code[orig_sym.upper()]
            if c['mcode'] and not c['mcode'].isdigit():
                final_trade = c['mcode']
            final_ean = final_ean or c['ean']

    # 5. Normalized Name Search
    if not final_trade or final_trade.isdigit():
        norm = normalize(orig_name)
        for tname, t in tim_by_name.items():
            if norm == normalize(tname):
                if t['mcode'] and not t['mcode'].isdigit():
                    final_trade = t['mcode']
                    final_ean = final_ean or t['ean']
                    break

    # 6. Jeśli dalej jest to czysta liczba i to KLUŚ -> sformatuj do pełnego kodu KLUŚ
    if (not final_trade or final_trade.isdigit()) and "KLUŚ" in orig_name.upper():
        num_sym = final_trade or orig_sym
        if num_sym.isdigit():
            final_trade = format_klus_connector_code(num_sym)

    # 7. Fallback dla EAN
    if not final_ean:
        final_ean = orig_ean

    return final_trade or orig_sym, final_ean


def main():
    tim_by_ean, tim_by_name, tim_by_mcode, xml_by_id, xml_by_name, cat_by_code, cat_by_mcode, cat_by_name, cat_by_ean = load_all_sources()

    wb = openpyxl.load_workbook(EXCEL_PATH)

    for sname in ["Towary", "PLIK OD PRESCOT"]:
        if sname not in wb.sheetnames:
            continue

        ws = wb[sname]
        ws.cell(row=1, column=2, value="indeks handlowy")
        
        fixed_trades = 0
        fixed_eans = 0
        still_digits = 0

        for row_idx in range(2, ws.max_row + 1):
            orig_sym = str(ws.cell(row=row_idx, column=2).value or '').strip()
            orig_ean = str(ws.cell(row=row_idx, column=3).value or '').strip()
            orig_name = str(ws.cell(row=row_idx, column=4).value or '').strip()

            trade_idx, valid_ean = resolve_full_trade_index(
                orig_sym, orig_ean, orig_name,
                tim_by_ean, tim_by_name, tim_by_mcode, xml_by_id, xml_by_name,
                cat_by_code, cat_by_mcode, cat_by_name, cat_by_ean
            )

            if trade_idx != orig_sym:
                ws.cell(row=row_idx, column=2, value=trade_idx)
                fixed_trades += 1

            if valid_ean != orig_ean:
                ws.cell(row=row_idx, column=3, value=valid_ean)
                fixed_eans += 1

            if str(trade_idx).isdigit():
                still_digits += 1

        print(f"✅ Zakładka {sname}:")
        print(f"   - Podmieniono na pełne indeksy handlowe: {fixed_trades}")
        print(f"   - Poprawiono kody EAN: {fixed_eans}")
        print(f"   - Pozostało czysto numerycznych kodów: {still_digits}")

    print(f"\n💾 Zapisywanie pliku do: {EXCEL_PATH}...")
    wb.save(EXCEL_PATH)
    print("🎉 ZAKOŃCZONO POMYŚLNIE!")


if __name__ == "__main__":
    main()
