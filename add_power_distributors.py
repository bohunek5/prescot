#!/usr/bin/env python3
"""
Dodaje 5 rozdzielaczy mocy Prescot (RM-MONO-19, RM-CCT-16, RM-RGB-14, RM-RGBW-13, RM-RGBCW-13)
do plików EL-PLUS / MegaCennik:
- Kody producenta oznaczone na zielono.
- Czyste opisy WAPRO HTML bez powielania nazw w <h2> i bez bloków parametrów w punktach.
- Aktualne ceny i stany magazynowe z chmury WAPRO.
- Prawidłowe linki do zdjęć na serwerze Prescot.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

FILES = [
    "/Users/karolbohdanowicz/Downloads/1294 EL-PLUS.xlsx",
    "/Users/karolbohdanowicz/Downloads/PRESCOT_EL-PLUS_ZESTAWIENIE_KODOW_2026.xlsx"
]

PRODUCTS = [
    {
        "mcode": "RM-MONO-19",
        "code": "Taś000748",
        "ean": "5905475368301",
        "name": "Rozdzielacz mocy LED Prescot Mono 1-9 DC 5-48V 25A",
        "cat": "Akcesoria do zasilaczy i taśm LED/Rozdzielacze mocy",
        "price": 18.00,
        "stock": 93,
        "unit": "szt",
        "cn": "8536 90 10",
        "img": "https://prescot.com.pl/userdata/gfx/69349.jpg",
        "desc": """<section>
<h2>Bezpieczny podział i dystrybucja zasilania w instalacjach jednokolorowych LED</h2>
<p>Kompaktowy moduł dystrybucji zasilania Prescot umożliwia wygodne, centralne rozdzielenie jednej linii zasilającej na 9 niezależnych obwodów taśm jednokolorowych LED. Pozwala na zachowanie idealnego porządku w okablowaniu oraz eliminuje konieczność stosowania plątaniny tradycyjnych złączek w puszkach i szafach instalacyjnych.</p>
<p>Solidna płytka PCB ze wzmocnionymi ścieżkami prądowymi obsługuje bezpiecznie prąd sumaryczny do 25 A w szerokim zakresie napięć od 5V do 48V DC, zapewniając stabilne napięcie na każdym wyjściu i chroniąc instalację przed spadkami jasności na dalszych odcinkach.</p>
<h3>Dlaczego warto:</h3>
<p>- Szybkie i przejrzyste podłączenie do 9 odcinków taśmy LED z jednego punktu zasilania</p>
<p>- Wysoka obciążalność prądowa do 25 A gwarantująca stabilną pracę rozległych systemów oświetleniowych</p>
<p>- Zabezpieczenie przed przegrzewaniem styków i spadkami napięć w instalacji</p>
<p>- Kompaktowe wymiary ułatwiające estetyczny montaż w korytkach kablowych, puszkach i profilach</p>
<h3>Zastosowanie i miejsce montażu:</h3>
<p>- Rozdzielanie zasilania taśm LED mono w oświetleniu meblowym, wnękowym, sufitowym oraz architektonicznym</p>
<p>- Dystrybucja zasilania w rozdzielnicach, szafach sterowniczych i centralnych punktach zasilania LED</p>
</section>"""
    },
    {
        "mcode": "RM-CCT-16",
        "code": "Taś000749",
        "ean": "5905475368318",
        "name": "Rozdzielacz mocy LED Prescot CCT 1-6 DC 5-48V 25A",
        "cat": "Akcesoria do zasilaczy i taśm LED/Rozdzielacze mocy",
        "price": 19.50,
        "stock": 96,
        "unit": "szt",
        "cn": "8536 90 10",
        "img": "https://prescot.com.pl/userdata/gfx/69343.jpg",
        "desc": """<section>
<h2>Efektywna dystrybucja zasilania i sterowania dla taśm LED CCT o zmiennej barwie</h2>
<p>Dedykowany rozdzielacz mocy Prescot do taśm LED CCT umożliwia precyzyjne rozdzielenie sygnału zasilania i sterowania barwą bieli z jednego sterownika lub zasilacza na 6 niezależnych linii oświetleniowych. Znacząco upraszcza montaż rozbudowanych instalacji multi-zone i sufitów napinanych.</p>
<p>Dzięki grubym ścieżkom miedzianym i nośności do 25 A w zakresie 5–48V DC, rozdzielacz zapewnia idealną synchronizację temperatury barwowej i jednakową jasność diod na wszystkich podłączonych odcinkach taśmy.</p>
<h3>Dlaczego warto:</h3>
<p>- Wygodny podział zasilania i 2-kanałowego sterowania CCT na 6 niezależnych obwodów</p>
<p>- Pełna kompatybilność z instalacjami 12V, 24V oraz 48V DC o dużym obciążeniu prądowym</p>
<p>- Estetyczny porządek w przewodach bez skręcania wielu kabli pod pojedynczymi zaciskami sterownika</p>
<p>- Trwała konstrukcja gwarantująca długoletnią i bezawaryjną pracę</p>
<h3>Zastosowanie i miejsce montażu:</h3>
<p>- Systemy oświetlenia ze zmienną temperaturą barwową bieli (ciepła/zimna) w domach, biurach i salonach</p>
<p>- Montaż w zabudowach gipsowo-kartonowych, puszkach instalacyjnych i profilach architektonicznych</p>
</section>"""
    },
    {
        "mcode": "RM-RGB-14",
        "code": "Taś000750",
        "ean": "5905475368325",
        "name": "Rozdzielacz mocy LED Prescot RGB 1-4 DC 5-48V 25A",
        "cat": "Akcesoria do zasilaczy i taśm LED/Rozdzielacze mocy",
        "price": 20.00,
        "stock": 100,
        "unit": "szt",
        "cn": "8536 90 10",
        "img": "https://prescot.com.pl/userdata/gfx/69338.jpg",
        "desc": """<section>
<h2>Centralny rozdział mocy i sygnału dla wielostrefowych instalacji LED RGB</h2>
<p>Zaawansowany moduł rozdzielczy Prescot zaprojektowany do jednoczesnego zasilania i przekazywania sygnałów sterujących RGB do 4 niezależnych linii taśm kolorowych. Pozwala na bezproblemowe zasilenie wielu odcinków bez przeciążania pojedynczych zacisków sterownika.</p>
<p>Obsługa prądu ciągłego do 25 A przy napięciu od 5V do 48V DC zapobiega spadkom napięć i niepożądanym przekłamaniom barw na końcach długich obwodów oświetlenia dekoracyjnego.</p>
<h3>Dlaczego warto:</h3>
<p>- Równoległe podłączenie do 4 odcinków taśmy RGB z zachowaniem pełnej synchronizacji kolorów</p>
<p>- Wysoka wydajność prądowa do 25 A umożliwiająca obsługę gęstych taśm wielobarwnych</p>
<p>- Czysty i profesjonalny montaż elektroinstalacyjny z czytelnym oznaczeniem biegunów i kanałów barwnych</p>
<p>- Brak strat sygnału i spadków napięcia na poszczególnych liniach</p>
<h3>Zastosowanie i miejsce montażu:</h3>
<p>- Oświetlenie dekoracyjne RGB w sufitach podwieszanych, strefach relaksu, barach i klubach</p>
<p>- Wielopunktowe instalacje podświetlenia mebli, gzymsów oraz witryn reklamowych</p>
</section>"""
    },
    {
        "mcode": "RM-RGBW-13",
        "code": "Taś000751",
        "ean": "5905475368332",
        "name": "Rozdzielacz mocy LED Prescot RGBW 1-3 DC 5-48V 25A",
        "cat": "Akcesoria do zasilaczy i taśm LED/Rozdzielacze mocy",
        "price": 20.50,
        "stock": 98,
        "unit": "szt",
        "cn": "8536 90 10",
        "img": "https://prescot.com.pl/userdata/gfx/69333.jpg",
        "desc": """<section>
<h2>Stabilna dystrybucja zasilania i sygnału dla 4-kanałowych taśm LED RGBW</h2>
<p>Specjalistyczny rozdzielacz mocy Prescot do zaawansowanych instalacji 4-kanałowych RGB+W (kolor + niezależna biel). Umożliwia łatwe podpięcie do 3 niezależnych linii oświetleniowych z zachowaniem pełnej zgodności polaryzacji i czystości montażu.</p>
<p>Wzmocniona architektura miedziana z prądem maksymalnym 25 A zapewnia niezmienną dynamikę kolorów i pełną moc białego światła zadaniowego na każdym podłączonym segmencie.</p>
<h3>Dlaczego warto:</h3>
<p>- Bezpieczne rozgałęzienie 5-żyłowej instalacji RGBW na 3 osobne ciągi świetlne</p>
<p>- Zdolność przewodzenia prądu do 25 A dostosowana do najbardziej wymagających taśm dużej mocy</p>
<p>- Przejrzysty układ zacisków eliminujący ryzyko pomyłki montażowej przy podłączaniu 5 linii</p>
<p>- Niezawodna praca w instalacjach domowych, komercyjnych i hotelowych</p>
<h3>Zastosowanie i miejsce montażu:</h3>
<p>- Hybrydowe instalacje oświetleniowe łączące nastrojowy kolor z funkcjonalną bielą</p>
<p>- Zabudowy sufitowe, korytarze, hotele, restauracje oraz przestrzenie eventowe</p>
</section>"""
    },
    {
        "mcode": "RM-RGBCW-13",
        "code": "Taś000757",
        "ean": "5905475368387",
        "name": "Rozdzielacz mocy LED Prescot RGBCW 1-3 DC 5-48V 25A",
        "cat": "Akcesoria do zasilaczy i taśm LED/Rozdzielacze mocy",
        "price": 20.50,
        "stock": 95,
        "unit": "szt",
        "cn": "8536 90 10",
        "img": "https://prescot.com.pl/userdata/gfx/69327.jpg",
        "desc": """<section>
<h2>Precyzyjny podział zasilania dla 5-kanałowych instalacji LED RGBCCT / RGBCW</h2>
<p>Najbardziej rozbudowany rozdzielacz mocy Prescot dla 5-kanałowych systemów LED (RGB + ciepła biel + zimna biel). Umożliwia podłączenie 3 niezależnych linii oświetleniowych, zapewniając pełną synchronizację barwną oraz swobodną regulację odcienia bieli.</p>
<p>Moduł przystosowany jest do napięć 5–48V DC i obciążenia do 25 A, gwarantując bezpieczny przesył prądu bez nagrzewania styków i bez widocznych różnic w barwie świecenia między odcinkami.</p>
<h3>Dlaczego warto:</h3>
<p>- Prosty i bezbłędny podział 6-żyłowej instalacji RGBCCT na 3 niezależne strefy</p>
<p>- Potężna obciążalność do 25 A zapewniająca bezproblemową pracę taśm o wysokiej gęstości diod</p>
<p>- Wyraźne oznaczenia zacisków ułatwiające pracę instalatorowi</p>
<p>- Zabezpieczenie przed spadkami napięć w rozległych instalacjach inteligentnego domu</p>
<h3>Zastosowanie i miejsce montażu:</h3>
<p>- Zaawansowane inteligentne systemy oświetlenia Smart Home (Tuya, Zigbee, DALI, KNX)</p>
<p>- Nowoczesne wnętrza mieszkalne i komercyjne wymagające pełnej elastyczności scen świetlnych</p>
</section>"""
    }
]

def add_products(fpath):
    if not os.path.exists(fpath):
        print(f"File not found: {fpath}")
        return

    print(f"Dodawanie rozdzielaczy mocy do {fpath}...")
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(name="Arial", size=9, bold=True, color="006100")
    std_font = Font(name="Arial", size=9)
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    # Check if any of these mcodes already exist
    existing_mcodes = set()
    for r in range(2, ws.max_row + 1):
        mc = str(ws.cell(r, 2).value or "").strip()
        if mc:
            existing_mcodes.add(mc.upper())

    added_cnt = 0
    for p in PRODUCTS:
        if p["mcode"].upper() in existing_mcodes:
            print(f"  {p['mcode']} already exists in file, skipping.")
            continue

        r = ws.max_row + 1
        lp = r - 1

        row_vals = [
            lp,
            p["mcode"],
            p["code"],
            p["ean"],
            p["name"],
            p["cat"],
            p["price"],
            "",
            "PLN",
            23,
            p["stock"],
            p["unit"],
            "",
            1,
            1,
            "",
            p["cn"],
            "",
            "TAK",
            "Prescot Sp. z o.o., ul. Wileńska 1, 11-500 Giżycko, mail: komponenty@prescot.pl, tel: 877776482",
            "24h",
            p["desc"],
            p["img"],
            "",
            ""
        ]

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = std_font
            cell.border = thin_border
            if col_idx == 2:  # kod producenta -> green
                cell.fill = green_fill
                cell.font = green_font
            elif col_idx in [1, 9, 10, 11, 12, 14, 15, 17, 19, 21]:
                cell.alignment = Alignment(horizontal='center')
            elif col_idx in [7]:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00'

        added_cnt += 1
        print(f"  + Dodano wiersz {r} (Lp {lp}): {p['mcode']} | {p['name']}")

    # Ensure contiguous Lp renumbering
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).value = r - 1

    wb.save(fpath)
    print(f"Zapisano {fpath} (łącznie produktów: {ws.max_row - 1}, dodano: {added_cnt}).\n")

for f in FILES:
    add_products(f)
